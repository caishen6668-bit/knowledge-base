import sys, json, math, re
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
from collections import defaultdict
import openpyxl
from copy import copy as pycopy
from alibabacloud_quickbi_public20220101.client import Client
from alibabacloud_quickbi_public20220101 import models as qb_models
from alibabacloud_tea_openapi import models as open_api_models

# 定位仓库根目录，加载基础设施
_repo = Path(__file__).resolve()
while not (_repo / ".git").exists() and _repo != _repo.parent:
    _repo = _repo.parent
sys.path.insert(0, str(_repo))
from 基础设施.配置.env import require

MEXICO_APPS = {'AndaLana', 'Cridit', 'Kredizo', 'ServiCash', 'TruCred'}

AK_ID = require("QBI_ACCESS_KEY")
AK_SECRET = require("QBI_SECRET_KEY")
config = open_api_models.Config(access_key_id=AK_ID, access_key_secret=AK_SECRET, region_id='cn-hangzhou')
config.endpoint = 'quickbi-public.cn-hangzhou.aliyuncs.com'
config.read_timeout = 180000; config.connect_timeout = 30000
client = Client(config)

# ============================================================
# 1. Due cases (c2f93e0fa45b) for 6.11-6.17
# ============================================================
print('Fetching due cases...')
req = qb_models.QueryDataServiceRequest(api_id='c2f93e0fa45b')
resp = client.query_data_service(req)
due_data = resp.body.result.values

target_dates = ['20260611','20260612','20260613','20260614','20260615','20260616','20260617']
due_by_date = {d: [0,0,0,0] for d in target_dates}  # 分期新客,分期老客,单期新客,单期老客

for r in due_data:
    app = r.get('app','')
    if app not in MEXICO_APPS: continue
    d = str(r.get('due_date',''))
    if d not in target_dates: continue
    cust = r.get('cust_type','')
    order = r.get('order','')
    cases = int(float(r.get('case',0) or 0))
    is_installment = order in ('借款分期', '展期分期', '展期N期')
    is_single = order in ('非分期',)
    if is_installment and cust == '新客': due_by_date[d][0] += cases
    elif is_installment and cust == '老客': due_by_date[d][1] += cases
    elif is_single and cust == '新客': due_by_date[d][2] += cases
    elif is_single and cust == '老客': due_by_date[d][3] += cases

# ============================================================
# 2. Recovery rates (524c3ccd429c) for due_day=20260528 (mature)
# ============================================================
print('Fetching recovery rates...')
req2 = qb_models.QueryDataServiceRequest(api_id='524c3ccd429c', conditions=json.dumps({'due_week':'2026-23'}))
resp2 = client.query_data_service(req2)
rate_data = resp2.body.result.values

def parse(s):
    if not s or s == '-': return 0
    return float(str(s).replace(',',''))

agg = defaultdict(lambda: defaultdict(float))
for r in rate_data:
    app = r.get('app_name','')
    if app not in MEXICO_APPS: continue
    cust = r.get('cust_type','')
    order = r.get('order_type','')
    # Group: 借款分期→分期, 非分期→单期, 展期分期→分期, 展期N期→单期
    if order in ('借款分期', '展期分期', '展期N期'): product = '分期'
    elif order in ('非分期',): product = '单期'
    else: continue
    key = (cust, product)
    for stage in ['D_2','D_1','D0','D1','S1','S2']:
        agg[key][stage+'_pay'] += parse(r.get(stage+'_pay_amt'))
        agg[key][stage+'_due'] += parse(r.get(stage+'_due_amt'))
    agg[key]['D_3_pay'] += parse(r.get('D_3_pay_amt'))

rates = {}
for (cust, product), d in agg.items():
    total_all = d['D_3_pay'] + sum(d[s+'_due'] for s in ['D_2','D_1','D0','D1','S1','S2'])
    rates[(cust, product, 'pre')] = d['D_3_pay']/total_all if total_all>0 else 0
    for stage, s in [('D-2','D_2'),('D-1','D_1'),('D0','D0'),('D1','D1'),('S1','S1'),('S2','S2')]:
        rates[(cust, product, stage)] = d[s+'_pay']/d[s+'_due'] if d[s+'_due']>0 else 0
    print(f'{cust} {product}: pre={rates[(cust,product,"pre")]:.2%} D-2={rates[(cust,product,"D-2")]:.2%} D-1={rates[(cust,product,"D-1")]:.2%} D0={rates[(cust,product,"D0")]:.2%} D1={rates[(cust,product,"D1")]:.2%} S1={rates[(cust,product,"S1")]:.2%} S2={rates[(cust,product,"S2")]:.2%}')

# ============================================================
# 3. Waterfall model
# ============================================================
MEXICO_APPS = {'AndaLana', 'Cridit', 'Kredizo', 'ServiCash', 'TruCred'}

# Constants
MEXICO_APPS = {'AndaLana', 'Cridit', 'Kredizo', 'ServiCash', 'TruCred'}
AI_RATIO = 1.0; MAN_RATIO = 0.0
SECOND_NEW = 0.2; SECOND_OLD = 0.3
HC_D0_D1 = 35; HC_S1 = 80; HC_S2 = 100

segments = [
    ('分期新客', [due_by_date[d][0] for d in target_dates], '新客', '分期'),
    ('分期老客', [due_by_date[d][1] for d in target_dates], '老客', '分期'),
    ('单期新客', [due_by_date[d][2] for d in target_dates], '新客', '单期'),
    ('单期老客', [due_by_date[d][3] for d in target_dates], '老客', '单期'),
]

results = {}
for seg_name, vals, cust, prod in segments:
    a = sum(vals) / len(vals)
    r = rates

    pre_out = a * (1 - r.get((cust, prod, 'pre'), 0.15))
    d2_out = pre_out * (1 - r.get((cust, prod, 'D-2'), 0.06))
    d1_ai_out = d2_out * AI_RATIO * (1 - r.get((cust, prod, 'D-1'), 0.18))
    d1_man_out = d2_out * MAN_RATIO * (1 - 0)
    d0_out = (d1_ai_out + d1_man_out) * (1 - r.get((cust, prod, 'D0'), 0.36))
    d1_out = d0_out * (1 - r.get((cust, prod, 'D1'), 0.06))
    s1_out = d1_out * (1 - r.get((cust, prod, 'S1'), 0.025))
    s2_out = s1_out * (1 - r.get((cust, prod, 'S2'), 0.015))

    results[seg_name] = {
        'avg': a, 'pre_out': pre_out, 'd2_out': d2_out,
        'd1_ai_out': d1_ai_out, 'd1_man_out': d1_man_out,
        'd0_out': d0_out, 'd1_out': d1_out,
        's1_out': s1_out, 's2_out': s2_out,
    }

fn = results['分期新客']
fo = results['分期老客']
dn = results['单期新客']
do = results['单期老客']

# Staffing (新客 = 分期新客 + 单期新客, 老客 = 分期老客 + 单期老客)
new_d0_in = (fn['d1_ai_out']+fn['d1_man_out']+dn['d1_ai_out']+dn['d1_man_out']) * (1-SECOND_NEW)
new_d1_in = fn['d0_out'] + dn['d0_out']
old_d0_in = (fo['d1_ai_out']+fo['d1_man_out']+do['d1_ai_out']+do['d1_man_out']) * (1-SECOND_OLD)
old_d1_in = fo['d0_out'] + do['d0_out']
s1_in = (fn['d1_out']+fo['d1_out']+dn['d1_out']+do['d1_out']) * 2
s2_in = (fn['s1_out']+fo['s1_out']+dn['s1_out']+do['s1_out']) * 3

staffing = {
    '新客 D0': (new_d0_in, HC_D0_D1),
    '新客 D1': (new_d1_in, HC_D0_D1),
    '老客 D0': (old_d0_in, HC_D0_D1),
    '老客 D1': (old_d1_in, HC_D0_D1),
    'S1': (s1_in, HC_S1),
    'S2': (s2_in, HC_S2),
}

print('\n=== 人力预估 ===')
for label, (inflow, cap) in staffing.items():
    need = math.ceil(inflow/cap) if inflow > 0 else 0
    est = math.ceil(need/7 + need) if need > 0 else 0
    print(f'{label}: {inflow:.0f}案  人均={cap}  需要={need}  预估={est}')

# ============================================================
# 4. Write to Excel
# ============================================================
print('\nWriting to Excel...')
wb = openpyxl.load_workbook(r'Desktop\周报_6.8-6.10.xlsx')
ws = wb['人数预估']

# Clear old O-R data
for r in range(10, 17):
    ws.cell(row=r, column=15).value = None
    ws.cell(row=r, column=16).value = None
    ws.cell(row=r, column=17).value = None
    ws.cell(row=r, column=18).value = None

# Write due cases for 6.11-6.17 (O=分期新客, P=分期老客, Q=单期新客, R=单期老客)
for i, d in enumerate(target_dates):
    r = 10 + i
    ws.cell(row=r, column=15).value = due_by_date[d][0]  # 分期新客
    ws.cell(row=r, column=16).value = due_by_date[d][1]  # 分期老客
    ws.cell(row=r, column=17).value = due_by_date[d][2]  # 单期新客
    ws.cell(row=r, column=18).value = due_by_date[d][3]  # 单期老客

# Update recovery rates in B-D-F-H columns
# Row 11: D-2之前, Row 12: D-2, Row 13: D-1AI, Row 14: D-1人工, Row 15: D0, Row 16: D1, Row 17: S1, Row 18: S2
rate_rows = {
    11: 'pre', 12: 'D-2', 13: 'D-1', 14: 'D-1', 15: 'D0', 16: 'D1', 17: 'S1', 18: 'S2',
}
for r, stage in rate_rows.items():
    # B=分期新客, D=分期老客, F=单期新客, H=单期老客
    ws.cell(row=r, column=2).value = rates.get(('新客', '分期', stage), 0)
    ws.cell(row=r, column=4).value = rates.get(('老客', '分期', stage), 0)
    ws.cell(row=r, column=6).value = rates.get(('新客', '单期', stage), 0)
    ws.cell(row=r, column=8).value = rates.get(('老客', '单期', stage), 0)

# Update J column in 新客/老客 sheets
def safe_set(ws, r, c, v):
    try: ws.cell(row=r, column=c).value = v
    except AttributeError: pass

for idx, (label, (inflow, cap)) in enumerate([
    ('新客', (new_d0_in, HC_D0_D1)),
    ('新客', (new_d1_in, HC_D0_D1)),
    ('老客', (old_d0_in, HC_D0_D1)),
    ('老客', (old_d1_in, HC_D0_D1)),
]):
    sn = wb.sheetnames[1] if idx < 2 else wb.sheetnames[2]
    ws2 = wb[sn]
    jr = 4 + (idx % 2)  # row 4 or 5
    need = math.ceil(inflow/cap) if inflow > 0 else 0
    est = math.ceil(need/7 + need) if need > 0 else 0
    safe_set(ws2, jr, 11, round(inflow, 0))  # K: 预估入催
    safe_set(ws2, jr, 12, need)              # L: 需要
    safe_set(ws2, jr, 13, est)               # M: 预估

# Also update 整体 J columns
ws_all = wb['整体']
# Find 整体 新客 section J rows
for r in range(3, 7):
    safe_set(ws_all, r, 14, 35)  # N: 人均持案

wb.save(r'Desktop\周报_6.8-6.10.xlsx')
print('Done! 人数预估已更新。')
