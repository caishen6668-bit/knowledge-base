import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl, math
from collections import defaultdict

wb = openpyxl.load_workbook(r'Desktop\周报_6.1-6.7.xlsx', data_only=True)
ws = wb['整体']

# === 1. Parse employee data ===
new_emps = {}  # name -> {催回率, stages, groups}
old_emps = {}

# Parse 新客 section (rows 4-26)
for row in ws.iter_rows(min_row=4, max_row=26, values_only=True):
    stage, name, status, group, attend, rec_rate, achieve, tenure = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
    if not name or str(name).strip() == '小计':
        continue
    name = str(name).strip()
    if status and str(status).strip() == '离职':
        continue
    rec_rate = float(rec_rate) if rec_rate else 0
    stage = str(stage).strip().replace('(new)', '')
    group = str(group).strip() if group else ''

    if name not in new_emps:
        new_emps[name] = {'rec_rates': {}, 'groups': set()}
    new_emps[name]['rec_rates'][stage] = rec_rate
    new_emps[name]['groups'].add(group)

# Parse 老客 section (rows 31-54)
for row in ws.iter_rows(min_row=31, max_row=54, values_only=True):
    stage, name, status, group, attend, rec_rate, achieve, tenure = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]
    if not name or str(name).strip() == '小计':
        continue
    name = str(name).strip()
    if status and str(status).strip() == '离职':
        continue
    rec_rate = float(rec_rate) if rec_rate else 0
    stage = str(stage).strip().replace('(new)', '')
    group = str(group).strip() if group else ''

    if name not in old_emps:
        old_emps[name] = {'rec_rates': {}, 'groups': set()}
    old_emps[name]['rec_rates'][stage] = rec_rate
    old_emps[name]['groups'].add(group)

# === 2. Sort by best 催回率 ===
def sort_by_rate(emps):
    result = []
    for name, data in emps.items():
        best_stage = max(data['rec_rates'], key=data['rec_rates'].get)
        best_rate = data['rec_rates'][best_stage]
        groups = ', '.join(sorted(data['groups']))
        result.append((name, best_rate, best_stage, groups))
    result.sort(key=lambda x: x[1], reverse=True)
    return result

new_sorted = sort_by_rate(new_emps)
old_sorted = sort_by_rate(old_emps)

print('=== 新客组在职员工 (按最佳催回率排序) ===')
for i, (name, rate, stage, groups) in enumerate(new_sorted):
    stages_str = ', '.join(f'{s}:{r:.2%}' for s, r in new_emps[name]['rec_rates'].items())
    print(f'  {i+1}. {name}  best={rate:.2%} @{stage}  all=[{stages_str}]  groups={groups}')

print(f'\n=== 老客组在职员工 (按最佳催回率排序) ===')
for i, (name, rate, stage, groups) in enumerate(old_sorted):
    stages_str = ', '.join(f'{s}:{r:.2%}' for s, r in old_emps[name]['rec_rates'].items())
    print(f'  {i+1}. {name}  best={rate:.2%} @{stage}  all=[{stages_str}]  groups={groups}')

# === 3. Headcount requirements ===
# From 人数预估 model:
# 新客 D0:7 D1:5
# 老客 D0:4 D1:3
# S1:5-6 S2:6-7 (pooled)

new_d0_need = 7
new_d1_need = 5
old_d0_need = 4
old_d1_need = 3
s1_need = 5
s2_need = 6

# === 4. Assignment logic ===
# D0 gets best performers, then D1, then S1, then S2
# S1/S2 pool from both customer types
# But D0/D1 are customer-specific

def assign(emps_sorted, d0_n, d1_n):
    """Assign employees to D0 and D1. Return remaining for S1/S2."""
    d0 = emps_sorted[:d0_n]
    remaining = emps_sorted[d0_n:]
    d1 = remaining[:d1_n]
    rest = remaining[d1_n:]
    return d0, d1, rest

new_d0, new_d1, new_rest = assign(new_sorted, new_d0_need, new_d1_need)
old_d0, old_d1, old_rest = assign(old_sorted, old_d0_need, old_d1_need)

print('\n=== 分配结果 ===')

print('\n【新客 D0】(需要{}人)'.format(new_d0_need))
for name, rate, stage, groups in new_d0:
    print(f'  {name} ({rate:.2%}) - {groups}')

print('\n【新客 D1】(需要{}人)'.format(new_d1_need))
for name, rate, stage, groups in new_d1:
    print(f'  {name} ({rate:.2%}) - {groups}')

print('\n【老客 D0】(需要{}人)'.format(old_d0_need))
for name, rate, stage, groups in old_d0:
    print(f'  {name} ({rate:.2%}) - {groups}')

print('\n【老客 D1】(需要{}人)'.format(old_d1_need))
for name, rate, stage, groups in old_d1:
    print(f'  {name} ({rate:.2%}) - {groups}')

# S1/S2: pool from both + sort
s_pool = new_rest + old_rest
s_pool.sort(key=lambda x: x[1], reverse=True)
s1 = s_pool[:s1_need]
s2 = s_pool[s1_need:s1_need + s2_need]

print('\n【S1】(需要{}人, 不限客群)'.format(s1_need))
for name, rate, stage, groups in s1:
    print(f'  {name} ({rate:.2%}) - {groups}')

print('\n【S2】(需要{}人, 不限客群)'.format(s2_need))
for name, rate, stage, groups in s2:
    print(f'  {name} ({rate:.2%}) - {groups}')

print('\n=== 未分配 ===')
total_assigned = len(new_d0) + len(new_d1) + len(old_d0) + len(old_d1) + len(s1) + len(s2)
total_available = len(new_sorted) + len(old_sorted)
print(f'  已分配: {total_assigned}  可用: {total_available}')
if total_assigned < total_available:
    unassigned = s_pool[s1_need + s2_need:]
    for name, rate, stage, groups in unassigned:
        print(f'  {name} ({rate:.2%}) - {groups}')
