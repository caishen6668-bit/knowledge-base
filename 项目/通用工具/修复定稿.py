import sys, unicodedata
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from copy import copy

wb = openpyxl.load_workbook(r'Desktop\周报_6.1-6.7.xlsx')

thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
center = Alignment(horizontal='center', vertical='center')
hdr_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
bold_font = Font(bold=True, size=11, name='Microsoft YaHei')
warn_font = Font(bold=True, size=11, color='CC0000', name='Microsoft YaHei')

stage_colors = {
    'D0': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'D1': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'S1': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'S2': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
}

ws_all = wb['整体']
ws_new = wb[wb.sheetnames[1]]
ws_old = wb[wb.sheetnames[2]]

# ============================================================
# 1. Unmerge everything in 整体
# ============================================================
for mr in list(ws_all.merged_cells.ranges):
    ws_all.unmerge_cells(str(mr))

# ============================================================
# 2. Copy 新客 R1-30 → 整体 R1-30  (values only)
# ============================================================
for r in range(1, 31):
    for c in range(1, 15):
        ws_all.cell(row=r, column=c).value = ws_new.cell(row=r, column=c).value

# Re-apply merges from 新客 to 整体 (rows 1-30)
for mr in ws_new.merged_cells.ranges:
    if mr.max_row <= 30:
        ws_all.merge_cells(str(mr))

# ============================================================
# 3. Empty 3 rows: R31-R33
# ============================================================
for r in range(31, 34):
    for c in range(1, 15):
        ws_all.cell(row=r, column=c).value = None

# ============================================================
# 4. Copy 老客 R1-30 → 整体 R34-63  (values only)
# ============================================================
for r in range(1, 31):
    dst_row = 33 + r
    for c in range(1, 15):
        ws_all.cell(row=dst_row, column=c).value = ws_old.cell(row=r, column=c).value

# Re-apply merges from 老客 to 整体 (shifted down by 33 rows)
for mr in ws_old.merged_cells.ranges:
    new_min_row = mr.min_row + 33
    new_max_row = mr.max_row + 33
    new_range = f'{mr.coord.split(":")[0]}:{mr.coord.split(":")[1]}'
    # We need to shift the row numbers
    # This is tricky — let's just manually re-merge known areas
    pass

# Manual merges for the shifted 老客 section
# 老客 R1→整体R34: title row, merge A34:H35
ws_all.merge_cells('A34:H35')
# 老客 R1→整体R34: J1 label, merge J34:N34
ws_all.merge_cells('J34:N34')
# 老客 R8→整体R41: 老客组员, merge K41:M41
ws_all.merge_cells('K41:M41')
# 老客 R29→整体R62: warning, merge A62:I63
ws_all.merge_cells('A62:I63')
# 老客 R11→整体R44: 小计 row for D0
# 老客 R16→整体R49: 小计 row for D1
# 老客 R22→整体R55: 小计 row for S1
# 老客 R27→整体R60: 小计 row for S2

# Fix title merger: A1:H2 already done above from 新客 copy

# ============================================================
# 5. Fix I column in 整体 (adjustment suggestion)
# ============================================================
def norm(name):
    return unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')

# Build employee→stage from both sheets' K-M
emp_stage = {}
for ws_src in [ws_new, ws_old]:
    for r in range(10, 29):
        n = ws_src.cell(row=r, column=11).value
        s = ws_src.cell(row=r, column=12).value
        if n and s:
            emp_stage[norm(str(n))] = str(s)

# Apply to 整体
for r in range(1, ws_all.max_row + 1):
    name = ws_all.cell(row=r, column=2).value
    if name and str(name).strip() != '小计' and str(name).strip() not in ('员工名称', '组员姓名'):
        n = norm(str(name))
        if n in emp_stage:
            ws_all.cell(row=r, column=9).value = emp_stage[n]
            s = emp_stage[n]
            if s in stage_colors:
                try: ws_all.cell(row=r, column=9).fill = stage_colors[s]
                except: pass

# ============================================================
# 6. Clean unused rows
# ============================================================
for r in range(64, ws_all.max_row + 1):
    for c in range(1, 15):
        try: ws_all.cell(row=r, column=c).value = None
        except AttributeError: pass

wb.save(r'Desktop\周报_6.1-6.7.xlsx')
print('Done! 整体 = 新客(1-30) + 空3行(31-33) + 老客(34-63)')
