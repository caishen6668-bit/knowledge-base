import sys, unicodedata
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Border, Side

wb = openpyxl.load_workbook(r'Desktop\周报_6.1-6.7.xlsx')

def norm(name):
    return unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')

# ============================================================
# 1. Build employee → stage mapping from K-M columns
# ============================================================
emp_stage = {}  # normalized_name → stage
emp_orig = {}   # normalized_name → original_name

for idx in [1, 2]:  # 新客, 老客
    ws = wb[wb.sheetnames[idx]]
    for r in range(10, 29):
        name = ws.cell(row=r, column=11).value  # K
        stage = ws.cell(row=r, column=12).value  # L
        if name and stage:
            n = norm(str(name))
            emp_stage[n] = str(stage)
            emp_orig[n] = str(name)

print(f'Built mapping: {len(emp_stage)} employees → stages')

# ============================================================
# 2. Fill column I (调整建议) in 新客, 老客, 整体
# ============================================================
def fill_col_i(ws, start_row, end_row):
    count = 0
    for r in range(start_row, end_row + 1):
        name = ws.cell(row=r, column=2).value  # B column
        if name and str(name).strip() != '小计':
            n = norm(str(name))
            if n in emp_stage:
                ws.cell(row=r, column=9).value = emp_stage[n]  # I column
                count += 1
            else:
                ws.cell(row=r, column=9).value = None
        else:
            ws.cell(row=r, column=9).value = None
    return count

# 新客 sheet
c1 = fill_col_i(wb[wb.sheetnames[1]], 3, 28)
# 老客 sheet
c2 = fill_col_i(wb[wb.sheetnames[2]], 3, 28)
# 整体 sheet
c3 = fill_col_i(wb['整体'], 3, 55)

print(f'Filled I column: 新客={c1} 老客={c2} 整体={c3}')

# ============================================================
# 3. Sync 整体 sheet data from 新客 + 老客
# ============================================================
# 整体 structure:
#   Row 1-2: title
#   Row 3: header
#   Rows 4-26: 新客 data (copy from 新客 sheet)
#   Rows 27-29: separator + 警告信
#   Row 30: title for 老客
#   Rows 31-54: 老客 data (copy from 老客 sheet)

# Copy 新客 data to 整体
ws_all = wb['整体']
ws_new = wb[wb.sheetnames[1]]
ws_old = wb[wb.sheetnames[2]]

# Stage colors for I column
stage_colors = {
    'D0': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'D1': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'S1': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'S2': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
}

def copy_rows(src_ws, dst_ws, src_start, dst_start, count):
    """Copy A-I values from src to dst"""
    for i in range(count):
        src_row = src_start + i
        dst_row = dst_start + i
        for col in range(1, 10):  # A-I
            val = src_ws.cell(row=src_row, column=col).value
            dst_ws.cell(row=dst_row, column=col).value = val
        # Color I column by stage
        i_val = dst_ws.cell(row=dst_row, column=9).value
        if i_val and str(i_val).strip() in stage_colors:
            dst_ws.cell(row=dst_row, column=9).fill = stage_colors[str(i_val).strip()]

# Copy 新客 rows 4-28 to 整体 rows 4-28
copy_rows(ws_new, ws_all, 4, 4, 25)
# Copy 老客 rows 4-27 to 整体 rows 31-54
copy_rows(ws_old, ws_all, 4, 31, 24)

# Also copy column I to row 30+ in 整体 from 老客
# (already handled by the copy above)

print('Synced 整体 sheet from 新客 + 老客')

# Also add K-M to 整体 for VLOOKUP reference (combine from both sheets)
# Clear 整体 K-M first
for r in range(10, 55):
    for c in [11, 12, 13]:
        ws_all.cell(row=r, column=c).value = None

# Write ALL assignments to 整体 K-M (D0 new, D1 new, D0 old, D1 old, S1 all, S2 all)
all_assign = []
for idx in [1, 2]:
    ws = wb[wb.sheetnames[idx]]
    for r in range(10, 29):
        name = ws.cell(row=r, column=11).value
        stage = ws.cell(row=r, column=12).value
        if name and stage:
            all_assign.append((str(name), str(stage)))

row = 10
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
for name, stage in all_assign:
    fill = stage_colors.get(stage, PatternFill(fill_type=None))
    ck = ws_all.cell(row=row, column=11); ck.value = name; ck.fill = fill; ck.border = thin
    cl = ws_all.cell(row=row, column=12); cl.value = stage; cl.fill = fill; cl.border = thin
    cm = ws_all.cell(row=row, column=13)
    cm.value = '=VLOOKUP(K:K,B:D,3,0)'; cm.fill = fill; cm.border = thin
    row += 1

print(f'Wrote {row-10} entries to 整体 K-M')

wb.save(r'Desktop\周报_6.1-6.7.xlsx')
print('\nDone! I列已填 + 整体已同步')
