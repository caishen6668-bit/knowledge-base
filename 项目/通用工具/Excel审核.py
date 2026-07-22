import openpyxl, sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook('日报6.8-6.13.xlsx')

# ============ 1. 行结构 ============
print("="*60)
print("1. 行结构检查")
print("="*60)

expected_accs = {
    3: ('D-1','YNT101','新客'), 4: ('D-1','YNT101','老客'), 5: ('D-1','YNT101','汇总'),
    6: ('D-1','YNT102','新客'), 7: ('D-1','YNT102','老客'), 8: ('D-1','YNT102','汇总'),
    9: ('D0','YND001','新客'), 10: ('D0','YND001','老客'), 11: ('D0','YND001','汇总'),
    12: ('D0','YND002','新客'), 13: ('D0','YND002','老客'), 14: ('D0','YND002','汇总'),
    15: ('S1','ZYS101','新客'), 16: ('S1','ZYS101','老客'), 17: ('S1','ZYS101','汇总'),
    18: ('S1','YNS101','新客'), 19: ('S1','YNS101','老客'), 20: ('S1','YNS101','汇总'),
    21: ('S1','RECS101','新客'), 22: ('S1','RECS101','老客'), 23: ('S1','RECS101','汇总'),
    24: ('S2','RECS201','新客'), 25: ('S2','RECS201','老客'), 26: ('S2','RECS201','汇总'),
    27: ('S2','S2beiyong','新客'), 28: ('S2','S2beiyong','老客'), 29: ('S2','S2beiyong','汇总'),
}

en_map = {'D-1':'D-1','D0':'D0','S1':'S1','S2':'S2','新客':'New','老客':'Old','汇总':'Summary'}
rows_ok = True
for sn in ['案件周度','金额周度','案件周度英语','金额周度英语']:
    ws = wb[sn]
    is_en = sn.endswith('英语')
    for r, (stage, acct, ctype) in expected_accs.items():
        a = ws.cell(r,1).value
        b = ws.cell(r,2).value
        c = ws.cell(r,3).value
        if is_en:
            exp_a = en_map.get(stage, stage)
            exp_c = en_map.get(ctype, ctype)
            if (a,b,c) != (exp_a, acct, exp_c):
                print(f'  ⚠ {sn} Row{r}: ({exp_a},{acct},{exp_c}) != actual')
                rows_ok = False
        else:
            if (a,b) != (stage, acct):
                print(f'  ⚠ {sn} Row{r}: ({stage},{acct}) != ({a},{b})')
                rows_ok = False

if rows_ok: print('  [OK] 所有Sheet行结构正确')

# ============ 2. 公式行号引用 ============
print("\n2. 公式行号检查")

date_map = {
    '13': ('G','H','I','J'), '12': ('K','L','M','N'), '11': ('O','P','Q','R'),
    '10': ('S','T','U','V'), '09': ('W','X','Y','Z'), '08': ('AA','AB','AC','AD'),
}
sum_rows = {5,8,11,14,17,20,23,26,29}
sum_map = {5:(3,4),8:(6,7),11:(9,10),14:(12,13),17:(15,16),
           20:(18,19),23:(21,22),26:(24,25),29:(27,28)}

errors = 0
for sn in ['案件周度','金额周度']:
    ws = wb[sn]
    for label, (case_c, recv_c, target_c, rate_c) in date_map.items():
        for r in range(3, 30):
            # 回收率: =IF({target}{r}=0,"",{recv}{r}/{target}{r})
            rate_f = ws[f'{rate_c}{r}'].value
            if rate_f and str(rate_f).startswith('='):
                expected_rate = f'=IF({target_c}{r}=0,"",{recv_c}{r}/{target_c}{r})'
                if str(rate_f) != expected_rate:
                    print(f'  ⚠ {sn} Row{r} {rate_c}(回收率): 期望{expected_rate}')
                    print(f'     实际{rate_f}')
                    errors += 1

            if r in sum_rows:
                nr, or_ = sum_map[r]
                for col in [case_c, recv_c, target_c]:
                    f = ws[f'{col}{r}'].value
                    expected = f'={col}{nr}+{col}{or_}'
                    if f != expected:
                        print(f'  ⚠ {sn} Row{r} {col}: 期望{expected} 实际{f}')
                        errors += 1

if errors == 0: print(f'  [OK] 所有{sn}公式正确')
else: print(f'  [ERR] {errors} 个公式错误')

# ============ 3. VLOOKUP范围 ============
print("\n3. VLOOKUP参数范围检查")

target_ranges = {
    '13': ('I','AN','AO'), '12': ('M','AH','AI'), '11': ('Q','AB','AC'),
    '10': ('U','V','W'), '09': ('Y','P','Q'), '08': ('AC','J','K'),
}

ok = True
for sheet, param in [('案件周度','案件参数'), ('金额周度','金额参数')]:
    ws = wb[sheet]
    for label, (tc, vc1, vc2) in target_ranges.items():
        f = ws[f'{tc}3'].value
        expected_range = f'${vc1}$2:${vc2}$11'
        if f and str(f).startswith('=') and expected_range not in str(f):
            print(f'  ❌ {sheet} 6/{label} {tc}3: 缺{expected_range}')
            print(f'     实际: {f}')
            ok = False

if ok: print('  [OK] 所有VLOOKUP范围正确')

# ============ 4. 参数表数据 ============
print("\n4. 参数表数据对比")
for sn in ['案件参数','金额参数']:
    ws = wb[sn]
    ranges = ['J:K','P:Q','V:W','AB:AC','AH:AI','AN:AO']
    ref = None
    diffs = []
    for rng in ranges:
        c1, c2 = rng.split(':')
        data = {}
        for r in range(2,12):
            k = ws[f'{c1}{r}'].value
            v = ws[f'{c2}{r}'].value
            if k: data[k] = v
        if ref is None: ref = data
        elif data != ref:
            for k in ref:
                if data.get(k) != ref[k]:
                    diffs.append(f'{rng} {k}: {ref[k]} → {data[k]}')
    if diffs:
        print(f'  ⚠ {sn}: {len(diffs)} 处不同')
        for d in diffs: print(f'    {d}')
    else:
        print(f'  [OK] {sn}: 所有范围一致')

# ============ 5. 英文版公式完整性 ============
print("\n5. 英文版公式完整性")
en_errors = 0
for en_sn, ch_sn in [('案件周度英语','案件周度'), ('金额周度英语','金额周度')]:
    ws_en = wb[en_sn]
    for label, (case_c, recv_c, target_c, rate_c) in date_map.items():
        for r in range(3, 30):
            for col in [case_c, recv_c, target_c, rate_c]:
                f = ws_en[f'{col}{r}'].value
                if f is None:
                    en_errors += 1
                    if en_errors <= 3:
                        print(f'  ⚠ {en_sn} Row{r} {col}: 空公式')
                elif not (str(f).startswith('=') and ch_sn in str(f)):
                    en_errors += 1
                    if en_errors <= 3:
                        print(f'  ⚠ {en_sn} Row{r} {col}: {f}')

if en_errors == 0: print('  [OK] 英文版公式完整')
else: print(f'  [ERR] {en_errors} 个问题')

wb.close()
print("\n" + "="*60)
print("检查完成")
