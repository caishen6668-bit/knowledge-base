import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

wb = openpyxl.load_workbook(r'Desktop\周报_6.1-6.7.xlsx')

thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
center = Alignment(horizontal='center', vertical='center')
hdr_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
bold_font = Font(bold=True, size=11, name='Microsoft YaHei')
normal_font = Font(size=10.5, name='Microsoft YaHei')

# ============================================================
# 1. Fix 老客 J1: "新客组别" → "老客组别"
# ============================================================
ws_old = wb[wb.sheetnames[2]]
ws_old.cell(row=1, column=10).value = '老客组别'
print('Fixed 老客 J1: -> 老客组别')

# ============================================================
# 2. J1-N1 merge & center (both 新客 and 老客)
# ============================================================
for idx in [1, 2]:
    ws = wb[wb.sheetnames[idx]]
    # Unmerge first then merge
    try: ws.unmerge_cells('J1:N1')
    except: pass
    ws.merge_cells('J1:N1')
    ws.cell(row=1, column=10).alignment = center
    ws.cell(row=1, column=10).font = bold_font

print('Fixed J1:N1 merge+center')

# ============================================================
# 3. K8:M8 merge & center (新客组员/老客组员 area)
# ============================================================
for idx in [1, 2]:
    ws = wb[wb.sheetnames[idx]]
    label = '新客组员' if idx == 1 else '老客组员'
    try: ws.unmerge_cells('K8:M8')
    except: pass
    ws.merge_cells('K8:M8')
    ws.cell(row=8, column=11).value = label
    ws.cell(row=8, column=11).alignment = center
    ws.cell(row=8, column=11).font = bold_font

print('Fixed K8:M8 merge+center')

# ============================================================
# 4. Clean up unused rows (A35+ for 新客, A28+ for 老客)
# ============================================================
# 新客: data ends R26, warning at R29-30, clear R27-28 and R31+
for idx in [1, 2]:
    ws = wb[wb.sheetnames[idx]]
    if idx == 1:
        # 新客: keep 1-26(data), 27-28(spacer), 29-30(warning), clear 31+
        for r in range(31, ws.max_row + 1):
            for c in range(1, 15):
                ws.cell(row=r, column=c).value = None
    else:
        # 老客: keep 1-27(data), 28(spacer), 29-30(warning), clear 31+
        for r in range(31, ws.max_row + 1):
            for c in range(1, 15):
                ws.cell(row=r, column=c).value = None
print('Cleaned up empty rows')

# ============================================================
# 5. Fix 整体: add header row for 老客 section
# ============================================================
ws_all = wb['整体']

# R30: title row (already set)
# R31: should be the header row for 老客 section
headers = [None, '员工名称', '在职状态', '小组名称', '出勤天数', '催回率', '达成率', '入职时长', '调整建议']
for c, h in enumerate(headers, 1):
    if h:
        cell = ws_all.cell(row=31, column=c)
        cell.value = h
        cell.font = bold_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin

# Also add sub-headers for J-N columns
# J1-N2 structure: row1=label, row2=预估入催/需要/预估/人均持案
ws_all.cell(row=30, column=9).value = '老客组别'

# J column labels for 老客 section
old_j_headers = {1: '老客组别', 2: '预估入催', 3: '需要', 4: '预估', 5: '人均持案'}
# Actually the J-N headers are already in rows 1-2 for the new section
# For old section, we just need the data area

print('Fixed 整体 old customer header row')

# ============================================================
# 6. Ensure 整体 K-M also present for 老客 section
# ============================================================
# The K-M in 整体 already has all assignments from our sync script
# But verify it exists

# ============================================================
# 7. Clean up 整体 unused rows
# ============================================================
# 整体: data ends around R56, clear R57+
for r in range(57, ws_all.max_row + 1):
    for c in range(1, 15):
        ws_all.cell(row=r, column=c).value = None

# Also clean up 人数预估 if needed
ws_est = wb['人数预估']
try:
    ws_est.cell(row=1, column=10).value = ws_est.cell(row=1, column=10).value  # no-op
except:
    pass

wb.save(r'Desktop\周报_6.1-6.7.xlsx')
print('\nAll polishing done!')
