import sys, unicodedata, math
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font

wb = openpyxl.load_workbook(r'Desktop\周报_6.1-6.7.xlsx')

def norm(name):
    return unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')

# ============================================================
# 1. Update 人数预估 sheet: D0/D1 人均持案 → 35
# ============================================================
ws_est = wb['人数预估']
# Cells with 人均持案 for D0/D1: F23, F24, F29, F30
# New customer D0 (F23), D1 (F24); Old customer D0 (F29), D1 (F30)
ws_est['F23'] = 35  # 新客 D0
ws_est['F24'] = 35  # 新客 D1
ws_est['F29'] = 35  # 老客 D0
ws_est['F30'] = 35  # 老客 D1
print('Updated 人数预估: D0/D1 人均持案 → 35')

# ============================================================
# 2. Update J columns in 新客/老客 sheets
# ============================================================
# Recalculate M column (预估) for D0/D1
# 新客 D0: 186/35=5.31, need=6, est=CEILING(6/7+6)=7
# 新客 D1: 148/35=4.23, need=5, est=CEILING(5/7+5)=6
# 老客 D0: 118/35=3.37, need=4, est=CEILING(4/7+4)=5
# 老客 D1: 71/35=2.03, need=3, est=CEILING(3/7+3)=4

ws_new = wb[wb.sheetnames[1]]  # 新客
ws_old = wb[wb.sheetnames[2]]  # 老客

# 新客 J4 (D0): K=预估入催, L=需要, M=预估, N=人均
ws_new['K4'] = 186; ws_new['L4'] = 6; ws_new['M4'] = 7; ws_new['N4'] = 35
# 新客 J5 (D1)
ws_new['K5'] = 148; ws_new['L5'] = 5; ws_new['M5'] = 6; ws_new['N5'] = 35

# 老客 J4 (D0)
ws_old['K4'] = 118; ws_old['L4'] = 4; ws_old['M4'] = 5; ws_old['N4'] = 35
# 老客 J5 (D1)
ws_old['K5'] = 71; ws_old['L5'] = 3; ws_old['M5'] = 4; ws_old['N5'] = 35

print('Updated J columns with new headcount numbers')

# ============================================================
# 3. Re-assign employees with new targets
# ============================================================
emp_lookup = {}
emp_rated = {}
ws_all = wb['整体']
for row in ws_all.iter_rows(min_row=4, max_row=54, values_only=True):
    name = str(row[1]).strip() if row[1] else ''
    group = str(row[3]).strip() if row[3] else ''
    status = str(row[2]).strip() if row[2] else ''
    rec_rate = float(row[5]) if row[5] else 0
    if name and name != '小计' and status != '离职':
        n = norm(name)
        emp_lookup[n] = {'original': name, 'group': group}
        if n not in emp_rated or rec_rate > emp_rated[n][0]:
            emp_rated[n] = (rec_rate, group)

new_groups = {'A_Antonio', 'C_Hector'}
old_groups = {'D_Daniela', 'E_Martin'}

new_sorted = sorted([(n, r, g) for n, (r, g) in emp_rated.items() if g in new_groups],
                    key=lambda x: x[1], reverse=True)
old_sorted = sorted([(n, r, g) for n, (r, g) in emp_rated.items() if g in old_groups],
                    key=lambda x: x[1], reverse=True)

# New targets (人均=35): 新客 D0=7, D1=6 | 老客 D0=5, D1=4 | S1=6, S2=7
new_d0_n, new_d1_n = 7, 6
old_d0_n, old_d1_n = 5, 4
s1_n, s2_n = 6, 7

def get_orig(names):
    return [emp_lookup[n]['original'] for n in names if n in emp_lookup]

new_d0_ascii = [n for n, r, g in new_sorted[:new_d0_n]]
new_d1_ascii = [n for n, r, g in new_sorted[new_d0_n:new_d0_n+new_d1_n]]
old_d0_ascii = [n for n, r, g in old_sorted[:old_d0_n]]
old_d1_ascii = [n for n, r, g in old_sorted[old_d0_n:old_d0_n+old_d1_n]]

assigned = set(new_d0_ascii + new_d1_ascii + old_d0_ascii + old_d1_ascii)
pool = [(n, r, g) for n, r, g in new_sorted + old_sorted if n not in assigned]
pool.sort(key=lambda x: x[1], reverse=True)

s1_ascii = [n for n, r, g in pool[:s1_n]]
s2_ascii = [n for n, r, g in pool[s1_n:s1_n+s2_n]]

new_d0 = get_orig(new_d0_ascii)
new_d1 = get_orig(new_d1_ascii)
old_d0 = get_orig(old_d0_ascii)
old_d1 = get_orig(old_d1_ascii)
s1_all = get_orig(s1_ascii)
s2_all = get_orig(s2_ascii)

def split_group(names):
    n, o = [], []
    for name in names:
        g = emp_lookup.get(norm(name), {}).get('group', '')
        if g in new_groups: n.append(name)
        else: o.append(name)
    return n, o

s1_new, s1_old = split_group(s1_all)
s2_new, s2_old = split_group(s2_all)

print(f'\n=== 人均35 分配结果 ===')
print(f'新客 D0({new_d0_n}): {new_d0}')
print(f'新客 D1({new_d1_n}): {new_d1}')
print(f'老客 D0({old_d0_n}): {old_d0}')
print(f'老客 D1({old_d1_n}): {old_d1}')
print(f'S1({s1_n}): 新{s1_new} 老{s1_old}')
print(f'S2({len(s2_all)}/{s2_n}): 新{s2_new} 老{s2_old}')

# ============================================================
# 4. Write to sheets with colors
# ============================================================
stage_colors = {
    'D0': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'D1': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'S1': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'S2': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
}
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def write_sheet(ws, d0, d1, s1, s2):
    for r in range(10, 29):
        for c in [11, 12, 13]:
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)

    row = 10
    for stage_label, name_list in [('D0', d0), ('D1', d1), ('S1', s1), ('S2', s2)]:
        fill = stage_colors[stage_label]
        for name in name_list:
            ck = ws.cell(row=row, column=11); ck.value = name; ck.fill = fill; ck.border = thin
            cl = ws.cell(row=row, column=12); cl.value = stage_label; cl.fill = fill; cl.border = thin
            cm = ws.cell(row=row, column=13)
            cm.value = '=VLOOKUP(K:K,B:D,3,0)'; cm.fill = fill; cm.border = thin
            row += 1
    return row - 10

n1 = write_sheet(ws_new, new_d0, new_d1, s1_new, s2_new)
n2 = write_sheet(ws_old, old_d0, old_d1, s1_old, s2_old)
print(f'\n写入: 新客={n1}  老客={n2}  合计={n1+n2}')

# ============================================================
# 5. Verify
# ============================================================
print('\n=== 与目标对比 ===')
print(f'新客 D0: 目标=7  安排={len(new_d0)}  {"OK" if len(new_d0)==7 else "GAP"}')
print(f'新客 D1: 目标=6  安排={len(new_d1)}  {"OK" if len(new_d1)==6 else "GAP"}')
print(f'老客 D0: 目标=5  安排={len(old_d0)}  {"OK" if len(old_d0)==5 else "GAP"}')
print(f'老客 D1: 目标=4  安排={len(old_d1)}  {"OK" if len(old_d1)==4 else "GAP"}')
print(f'S1:      目标=6  安排={len(s1_all)}  {"OK" if len(s1_all)==6 else "GAP"}')
gap_s2 = s2_n - len(s2_all)
print(f'S2:      目标=7  安排={len(s2_all)}  {"GAP "+str(gap_s2)+"人" if gap_s2>0 else "OK"}')

wb.save(r'Desktop\周报_6.1-6.7.xlsx')
print('\nSaved!')
