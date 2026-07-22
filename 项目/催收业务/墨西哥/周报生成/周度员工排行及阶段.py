# 周报生成 - 从模板复制格式，只替换数据
import sys, requests, json, hmac, hashlib, base64, time, uuid, math, pandas as pd
from pathlib import Path
from urllib.parse import quote
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy as cp
import shutil, copy

sys.stdout.reconfigure(encoding='utf-8')

# 定位仓库根目录，加载基础设施
_repo = Path(__file__).resolve()
while not (_repo / ".git").exists() and _repo != _repo.parent:
    _repo = _repo.parent
sys.path.insert(0, str(_repo))
from 基础设施.配置.env import require

# ============================================================
AK=require("QBI_ACCESS_KEY");SK=require("QBI_SECRET_KEY")
API='https://quickbi-public.cn-hangzhou.aliyuncs.com/'
A_DUE='c2f93e0fa45b';A_ATT='7f9969dc9020';A_COL='c4f429db60b3';A_REC='524c3ccd429c'
DS='20260713';DE='20260719';PERIOD='7.13-7.19'
TODAY=pd.Timestamp('2026-07-20');DUE_WEEK='2026-29'
ABC=['A_Antonio','A_Group','B_Pavel','C_Hector','B_Hiromi']
DEF=['D_Daniela','E_Martin','F_Frida']
VA=['Cridit','Kredizo','ServiCash','TruCred']
FQ_O=['借款分期','展期N期','展期分期']
TMPL=r'C:\Users\Administrator\Desktop\周报_6.1-6.7.xlsx'
OUT=r'C:\Users\Administrator\Desktop\周报_7.13-7.19.xlsx'
CPP_D0,CPP_D1,CPP_S1,CPP_S2=35,35,80,100;CPP_D_1=40
D2_BEFORE={'分期新客':0.15,'分期老客':0.19,'单期新客':0.10,'单期老客':0.10}

def call_api(aid,cond=None):
    p={'Action':'QueryDataService','Format':'JSON','Version':'2022-01-01','AccessKeyId':AK,'SignatureMethod':'HMAC-SHA1','Timestamp':time.strftime('%Y-%m-%dT%H:%M:%SZ',time.gmtime()),'SignatureVersion':'1.0','SignatureNonce':str(uuid.uuid4()),'ApiId':aid}
    if cond:p.update(cond)
    sp=sorted(p.items());cq='&'.join([quote(k,safe='')+'='+quote(str(v),safe='') for k,v in sp])
    s=base64.b64encode(hmac.new((SK+'&').encode(),('POST&%2F&'+quote(cq,safe='')).encode(),hashlib.sha1).digest()).decode()
    p['Signature']=s
    return pd.DataFrame(requests.post(API,data=p,timeout=90).json()['Result']['Values'])

def to_num(df,cols):
    for c in cols:
        if c in df.columns:df[c]=pd.to_numeric(df[c].astype(str).str.replace(',','').str.replace('-','0'),errors='coerce').fillna(0)

# ============================================================
# 数据拉取 (同之前)
# ============================================================
print('[1/4] 回收+出勤...')
col_parts=[]
for dd in pd.date_range(pd.Timestamp(DS),pd.Timestamp(DE)):
    col_parts.append(call_api(A_COL,{'Conditions':'{"statis_date":"%s"}'%dd.strftime('%Y%m%d')}))
df=pd.concat(col_parts)

# 仅保留PL团队
df = df[df['dept_name'] == 'PL'].copy()

print("\n===== PL过滤后 =====")
print(df['dept_name'].value_counts())
print(df['coll_group_name'].value_counts())

df['d']=pd.to_datetime(df['statis_date'],format='%Y%m%d',errors='coerce')
df=df[(df['d']>=DS)&(df['d']<=DE)].copy()
to_num(df,['today_repay_amount','today_enter_collect_amount','target'])
df['rate']=df['today_repay_amount']/df['today_enter_collect_amount'].replace(0,float('nan'))
df['rate']=df['rate'].fillna(0)
df['achieve']=df['rate']/df['target'].replace(0,float('nan'))
df['achieve']=df['achieve'].fillna(0)

att=call_api(A_ATT);att['day']=pd.to_datetime(att['day'],format='%Y%m%d',errors='coerce')
att['join']=pd.to_datetime(att['join_date'],format='%Y%m%d',errors='coerce')
att['tenure']=(TODAY-att['join']).dt.days
ast=att[['name','day','working_status']].rename(columns={'name':'staff_name','day':'d'})
df=pd.merge(df,ast,on=['staff_name','d'],how='left')
df['working_status']=df['working_status'].fillna('上班')
df=df[df['working_status']=='上班'].copy()
le=att.sort_values('day').groupby('name').last().reset_index()
ei=le[['name','group','is_resigned','tenure']].rename(columns={'name':'staff_name','group':'g0','is_resigned':'status'})
df=pd.merge(df,ei,on='staff_name',how='left')
df['tenure']=df['tenure'].fillna(0).astype(int)
df['status']=df['status'].fillna('在职').astype(str)

def get_g(r):
    g=r.get('g0','')
    if pd.notna(g) and g not in ['-','','未知','nan']:return g
    dn=str(r.get('dept_name',''))
    if dn in ABC or dn in DEF:return dn
    # g0缺失时用coll_group_name判定，不再硬编码C_Hector
    cg=str(r.get('coll_group_name',''))
    if cg in ABC or cg in DEF:return cg
    if 'PL' in dn:return 'C_Hector'
    return None

df['grp']=df.apply(get_g,axis=1)
df['cust']=df['grp'].apply(lambda g:'new' if g in ABC else ('old' if g in DEF else None))
df=df[df['cust'].notna()].copy()
df=df[df['overdue_level'].isin(['D0','D1','S1','S1(new)','S2','S2(new)'])].copy()

g=df.groupby(['cust','overdue_level','staff_name','grp','status','tenure'],sort=False).agg(
    days=('d','nunique'),
    repay_amt=('today_repay_amount','sum'),
    enter_amt=('today_enter_collect_amount','sum'),
    target=('target','mean')
).reset_index()

g['rate']=g['repay_amt']/g['enter_amt'].replace(0,float('nan'))
g['rate']=g['rate'].fillna(0)
g['ach']=g['rate']/g['target'].replace(0,float('nan'))
g['ach']=g['ach'].fillna(0)
g.rename(columns={'overdue_level':'stage','staff_name':'name','grp':'team'},inplace=True)
g=g.sort_values(['cust','stage','rate'],ascending=[True,True,False]).reset_index(drop=True)
new=g[g['cust']=='new'];old=g[g['cust']=='old']
print(f'  新客:{new["name"].nunique()}人 老客:{old["name"].nunique()}人 共{len(df)}条')

print('[2/4] 到期单量...')
start_d=pd.Timestamp(DE)+pd.Timedelta(days=1)
# A_DUE单次调用返回14天数据，只取一次避免重复
due=call_api(A_DUE,{'Conditions':'{"due_date":"%s"}'%start_d.strftime('%Y%m%d')})
due['d']=pd.to_datetime(due['due_date'],format='%Y%m%d',errors='coerce')
due['case']=pd.to_numeric(due['case'],errors='coerce');due=due[due['app']!='Instamonei'].copy()
is_fq=due['order'].isin(FQ_O);is_x=due['cust_type']=='新客';is_l=due['cust_type']=='老客'
daily_due=[]
for i in range(7):
    d=start_d+pd.Timedelta(days=i)
    o=int(due[(due['d']==d)&is_fq&is_x]['case'].sum())
    p=int(due[(due['d']==d)&is_fq&is_l]['case'].sum())
    q=int(due[(due['d']==d)&(~is_fq)&is_x]['case'].sum())
    r=int(due[(due['d']==d)&(~is_fq)&is_l]['case'].sum())
    daily_due.append((d,o,p,q,r))
avg_o=sum(x[1] for x in daily_due)/7;avg_p=sum(x[2] for x in daily_due)/7
avg_q=sum(x[3] for x in daily_due)/7;avg_r=sum(x[4] for x in daily_due)/7

print(f'\n===== A_DUE调试 =====')
print(f'start_d(下周首日)={start_d.strftime("%Y-%m-%d")}, 拉取范围={DS}~{pd.Timestamp(DE)+pd.Timedelta(days=7):%Y%m%d}')
print(f'due总行数={len(due)}, due列={due.columns.tolist()}')
print(f'due日期分布:')
print(due['d'].value_counts().sort_index())
print(f'\n每日到期(下周7天):')
for d,o,p,q,r_v in daily_due:
    print(f'  {d.strftime("%m/%d")} 分期新={o} 分期老={p} 单期新={q} 单期老={r_v}')
print(f'日均: 分期新={avg_o:.1f} 分期老={avg_p:.1f} 单期新={avg_q:.1f} 单期老={avg_r:.1f}')
print(f'========================\n')

print('[3/4] 回收率 (due_week)...')
rec_w=call_api(A_REC);rec_w=rec_w[rec_w['due_week']==DUE_WEEK].copy()
acols=['due_amt','D_2_pay_amt','D_2_due_amt','D_1_pay_amt','D_1_due_amt','D0_pay_amt','D0_due_amt',
       'D1_pay_amt','D1_due_amt','S1_pay_amt','S1_due_amt','S2_pay_amt','S2_due_amt']
to_num(rec_w,acols);rec_w['mn']=pd.to_numeric(rec_w['mult_no'],errors='coerce').fillna(0)
rec_w['is_fq']=(rec_w['app_name'].isin(VA))&(rec_w['order_type'].isin(FQ_O))&(rec_w['mn'].isin([1,2,3,4]))
r_fq=rec_w[rec_w['is_fq']].copy();r_dq=rec_w[(~rec_w['is_fq'])&rec_w['app_name'].isin(VA)].copy()

print(rec_w.columns.tolist())

print("\n===== 分期新客验证 =====")

s = r_fq[r_fq['cust_type']=='新客'].copy()

print("总到期金额:", round(s['due_amt'].sum(),2))
print("D0到期金额:", round(s['D0_due_amt'].sum(),2))
print("D1到期金额:", round(s['D1_due_amt'].sum(),2))
print("S1到期金额:", round(s['S1_due_amt'].sum(),2))

print("D-2回款:", round(s['D_2_pay_amt'].sum(),2))
print("D-1回款:", round(s['D_1_pay_amt'].sum(),2))
print("D0回款:", round(s['D0_pay_amt'].sum(),2))
print("D1回款:", round(s['D1_pay_amt'].sum(),2))
print("S1回款:", round(s['S1_pay_amt'].sum(),2))

print("\n========== 调试 ==========")

print("due_week分布:")
print(rec_w['due_week'].value_counts().sort_index())

print("\nmult_no分布:")
print(rec_w['mult_no'].value_counts().sort_index())

print("\nAPP分布:")
print(rec_w['app_name'].value_counts())

print("========== 调试结束 ==========\n")

def stage_rates(sub,ct):
    rates={};pairs=[('D-2','D_2'),('D-1','D_1'),('D0','D0'),('D1','D1'),('S1','S1'),('S2','S2')]
    for st,pk in pairs:
        s=sub[sub['cust_type']==ct];pay=float(s[f'{pk}_pay_amt'].sum());due_v=float(s[f'{pk}_due_amt'].sum())
        rates[st]=round(pay/due_v,4) if due_v>0 else 0
    return rates

print("===== 分期新客 =====")

s = r_fq[r_fq['cust_type']=='新客']

for st in [
    ('D-2','D_2'),
    ('D-1','D_1'),
    ('D0','D0'),
    ('D1','D1'),
    ('S1','S1'),
    ('S2','S2')
]:
    name,pk = st

    pay = s[f'{pk}_pay_amt'].sum()
    due = s[f'{pk}_due_amt'].sum()

    print(
        name,
        "pay=",round(pay,2),
        "due=",round(due,2),
        "rate=",round(pay/due*100,2) if due else 0
    )

nf_rates=stage_rates(r_fq,'新客');of_rates=stage_rates(r_fq,'老客')
nd_rates=stage_rates(r_dq,'新客');od_rates=stage_rates(r_dq,'老客')

print("\n===== 回收率口径 =====")
print(f"{'阶段':<6} {'分期新客':<12} {'分期老客':<12} {'单期新客':<12} {'单期老客':<12}")
for st in ['D-2','D-1','D0','D1','S1','S2']:
    print(f"{st:<6} {nf_rates.get(st,0):<12.4f} {of_rates.get(st,0):<12.4f} {nd_rates.get(st,0):<12.4f} {od_rates.get(st,0):<12.4f}")

print('[4/4] Waterfall+排人...')
def wf(due_v,rates,ai_r=1):
    # rates:
    # [D2_BEFORE,D-2,D-1,AI,D0,D1,S1,S2]
    # 当前模型只使用到 S1，S2回收率未参与计算（沿用原模板逻辑）
    c11=due_v*(1-rates[0]);c12=c11*(1-rates[1]);c13=c12*ai_r*(1-rates[2]);c14=c12*(1-ai_r)*(1-rates[3])
    d0_in=c13+c14;d1_out=d0_in*(1-rates[4]);s1_out=d1_out*(1-rates[5]);s2_out=s1_out*(1-rates[6])
    return d0_in,d1_out,s1_out*2,s2_out*3

nf_=wf(avg_o,[D2_BEFORE['分期新客'],nf_rates.get('D-2',0),nf_rates.get('D-1',0),0,nf_rates.get('D0',0),nf_rates.get('D1',0),nf_rates.get('S1',0),nf_rates.get('S2',0)])
nq_=wf(avg_q,[D2_BEFORE['单期新客'],nd_rates.get('D-2',0),nd_rates.get('D-1',0),0,nd_rates.get('D0',0),nd_rates.get('D1',0),nd_rates.get('S1',0),nd_rates.get('S2',0)])
of_=wf(avg_p,[D2_BEFORE['分期老客'],of_rates.get('D-2',0),of_rates.get('D-1',0),0,of_rates.get('D0',0),of_rates.get('D1',0),of_rates.get('S1',0),of_rates.get('S2',0)])
oq_=wf(avg_r,[D2_BEFORE['单期老客'],od_rates.get('D-2',0),od_rates.get('D-1',0),0,od_rates.get('D0',0),od_rates.get('D1',0),od_rates.get('S1',0),od_rates.get('S2',0)])

# 策略系数（和模板公式一致: *(1-N23) for 新客, *(1-N24) for 老客）
STRAT_NEW=0.2;STRAT_OLD=0.3
nd0_in=(nf_[0]+nq_[0])*(1-STRAT_NEW);nd1_in=nf_[1]+nq_[1]
od0_in=(of_[0]+oq_[0])*(1-STRAT_OLD);od1_in=of_[1]+oq_[1]
s1_total=nf_[2]+nq_[2]+of_[2]+oq_[2];s2_total=nf_[3]+nq_[3]+of_[3]+oq_[3]

nd0_n=max(1,math.ceil(nd0_in/CPP_D0));nd1_n=max(1,math.ceil(nd1_in/CPP_D1))
od0_n=max(1,math.ceil(od0_in/CPP_D0));od1_n=max(1,math.ceil(od1_in/CPP_D1))
s1_n=max(1,math.ceil(s1_total/CPP_S1));s2_n=max(1,math.ceil(s2_total/CPP_S2))

print(f'\n===== Waterfall调试 =====')
print(f'分期新客wf: d0_in={nf_[0]:.1f} d1_out={nf_[1]:.1f} s1_out={nf_[2]:.1f} s2_out={nf_[3]:.1f}')
print(f'单期新客wf: d0_in={nq_[0]:.1f} d1_out={nq_[1]:.1f} s1_out={nq_[2]:.1f} s2_out={nq_[3]:.1f}')
print(f'分期老客wf: d0_in={of_[0]:.1f} d1_out={of_[1]:.1f} s1_out={of_[2]:.1f} s2_out={of_[3]:.1f}')
print(f'单期老客wf: d0_in={oq_[0]:.1f} d1_out={oq_[1]:.1f} s1_out={oq_[2]:.1f} s2_out={oq_[3]:.1f}')
print(f'策略后: nd0_in={nd0_in:.1f} nd1_in={nd1_in:.1f} od0_in={od0_in:.1f} od1_in={od1_in:.1f}')
print(f's1_total={s1_total:.1f} s2_total={s2_total:.1f}')
print(f'CPP前: nd0={nd0_in/CPP_D0:.1f} nd1={nd1_in/CPP_D1:.1f} od0={od0_in/CPP_D0:.1f} od1={od1_in/CPP_D1:.1f}')
print(f'CPP后: nd0_n={nd0_n} nd1_n={nd1_n} od0_n={od0_n} od1_n={od1_n} s1_n={s1_n} s2_n={s2_n}')
print(f'========================\n')

buf_new_d0=nd0_n+math.ceil(nd0_n/7);buf_new_d1=nd1_n+math.ceil(nd1_n/7)
buf_old_d0=od0_n+math.ceil(od0_n/7);buf_old_d1=od1_n+math.ceil(od1_n/7)
buf_s1=s1_n+math.ceil(s1_n/7);buf_s2=s2_n+math.ceil(s2_n/7)

def assign_all(data,d0_est,d1_est,s1_est,s2_est):
    """统一按优先级分配: D0→D1→S1→S2, 前面不够从后面抢人"""
    result={'D0':[],'D1':[],'S1':[],'S2':[]};used=set()
    pool=data[data['status']!='离职'].copy()
    pool=pool.sort_values('rate',ascending=False).drop_duplicates('name',keep='first')
    # D0-stage优先排D0
    d0_pool=pool[pool['stage'].str.contains('D0',na=False)]
    for _,row in d0_pool.iterrows():
        if len(result['D0'])<d0_est:
            result['D0'].append((row['name'],row['team'],row['rate']));used.add(row['name'])
    # D0不满→从剩余所有人补
    remain=pool[~pool['name'].isin(used)]
    for _,row in remain.iterrows():
        if len(result['D0'])>=d0_est:break
        result['D0'].append((row['name'],row['team'],row['rate']));used.add(row['name'])
    # D1→S1→S2 按序填充
    for stage,limit in [('D1',d1_est),('S1',s1_est),('S2',s2_est)]:
        remain=pool[~pool['name'].isin(used)]
        for _,row in remain.iterrows():
            if len(result[stage])>=limit:break
            result[stage].append((row['name'],row['team'],row['rate']));used.add(row['name'])
    # 空槽
    for stage,limit in [('D0',d0_est),('D1',d1_est),('S1',s1_est),('S2',s2_est)]:
        while len(result[stage])<limit:result[stage].append(('','',0))
    return result

# Assign per group: S1/S2 quota split by remaining headcount proportion
# Count total people per group for S1/S2 split
new_pool=new[new['status']!='离职'].drop_duplicates('name')
old_pool=old[old['status']!='离职'].drop_duplicates('name')
new_total_people=len(new_pool);old_total_people=len(old_pool)
total_people=new_total_people+old_total_people
if total_people>0:
    new_ratio=new_total_people/total_people;old_ratio=old_total_people/total_people
else:
    new_ratio=0.5;old_ratio=0.5

new_s1_n=max(0,round(buf_s1*new_ratio));old_s1_n=max(0,buf_s1-new_s1_n)
new_s2_n=max(0,round(buf_s2*new_ratio));old_s2_n=max(0,buf_s2-new_s2_n)

new_assigned=assign_all(new,buf_new_d0,buf_new_d1,new_s1_n,new_s2_n)
old_assigned=assign_all(old,buf_old_d0,buf_old_d1,old_s1_n,old_s2_n)

# Report how many are actually assigned
new_total=sum(len(v) for v in new_assigned.values())
old_total=sum(len(v) for v in old_assigned.values())
print(f'  新客D0={nd0_n}(预{buf_new_d0}) D1={nd1_n}(预{buf_new_d1}) S1={new_s1_n}/{buf_s1} S2={new_s2_n}/{buf_s2} 排{new_total}人')
print(f'  老客D0={od0_n}(预{buf_old_d0}) D1={od1_n}(预{buf_old_d1}) S1={old_s1_n}/{buf_s1} S2={old_s2_n}/{buf_s2} 排{old_total}人')

# ============================================================
# 从模板复制，只替换数据
# ============================================================
print('生成Excel(从模板复制)...')
shutil.copy2(TMPL, OUT)
wb = load_workbook(OUT)
# 图片样式预定义
stage_fills = {'D0':'FFC6EFCE','D1':'FFBDD7EE','S1':'FFFCE4D6','S2':'FFD9D9D9'}

def fill_ranking_sheet(ws, data, cust_label, assigned, d0in, d1in, d0n, d1n, d0_est, d1_est, row_offset=0):
    """填充排名sheet：清数据→写新数据，保留模板格式，支持row_offset偏移"""
    # 找到模板的现有区域边界
    # 查找A列"警告信名单"的位置来确定排名数据结束行
    warning_row = None
    for r in range(4, ws.max_row+1):
        v = ws.cell(row=r, column=1).value
        if v and '警告信' in str(v):
            warning_row = r
            break
    if not warning_row:
        warning_row = ws.max_row + 1

    # === 更新标题 ===
    ws.cell(row=1+row_offset, column=1).value = f'{PERIOD} PL组员回收情况汇总'
    ws.cell(row=1+row_offset, column=10).value = f'{cust_label}组别'

    # === 更新J4-N6: D-1/D0/D1 预估数字 ===
    ws.cell(row=4+row_offset, column=10).value = 'D0'
    ws.cell(row=4+row_offset, column=11).value = round(d0in, 1)
    ws.cell(row=4+row_offset, column=12).value = d0n
    ws.cell(row=4+row_offset, column=13).value = d0_est
    ws.cell(row=4+row_offset, column=14).value = CPP_D0
    ws.cell(row=5+row_offset, column=10).value = 'D1'
    ws.cell(row=5+row_offset, column=11).value = round(d1in, 1)
    ws.cell(row=5+row_offset, column=12).value = d1n
    ws.cell(row=5+row_offset, column=13).value = d1_est
    ws.cell(row=5+row_offset, column=14).value = CPP_D1
    ws.cell(row=6+row_offset, column=12).value = d0n + d1n
    ws.cell(row=6+row_offset, column=13).value = d0_est + d1_est

    # === A-I: 排名数据 ===
    # Build the ranking rows first
    ranking_rows = []  # [(stage, name, status, team, days, rate, ach, tenure, is_subtot, is_off), ...]

    for (st,), gr in data.groupby(['stage'], sort=False):
        gs = gr.sort_values('rate', ascending=False)
        for _, row in gs.iterrows():
            rate_v = row['rate']; status = str(row['status']); off = status == '离职'
            ranking_rows.append((st, row['name'], status, row['team'], int(row['days']),
                                round(rate_v, 5), round(row['ach'], 5), int(row['tenure']),
                                False, off))
        # 小计行
        subtotal_enter = gs['enter_amt'].sum()
        subtotal_repay = gs['repay_amt'].sum()
        subtotal_rate = subtotal_repay / subtotal_enter if subtotal_enter > 0 else 0
        subtotal_target = gs['target'].iloc[0]
        subtotal_ach = subtotal_rate / subtotal_target if subtotal_target > 0 else 0
        ranking_rows.append((st, '小计', '', '', int(gs['days'].sum()),
                            round(subtotal_rate, 5), round(subtotal_ach, 5), '',
                            True, False))

    # === K-M: 人员安排 ===
    staff_rows = []  # [(name, stage, team), ...]
    for st in ['D0', 'D1', 'S1', 'S2']:
        for name, team, rate in assigned.get(st, []):
            staff_rows.append((name, st, team))

    # === 找到模板的"组员姓名"行来定位人员安排起始位置 ===
    staff_header_row = None
    for r in range(1, ws.max_row+1):
        v = ws.cell(row=r, column=11).value
        if v and '组员姓名' in str(v):
            staff_header_row = r
            break
    if not staff_header_row:
        staff_header_row = 9 + row_offset

    # === 算出需要的行数 ===
    need_ranking_rows = len(ranking_rows)
    need_staff_rows = len(staff_rows)

    # 排名从行4开始，占 need_ranking_rows 行
    ranking_end = 4 + need_ranking_rows - 1

    # 人员安排从 staff_header_row+1 开始(组员姓名下一行)，占 need_staff_rows 行
    staff_start = staff_header_row + 1
    staff_end = staff_start + need_staff_rows - 1

    # 警告信所在行 = max(ranking_end, staff_end) + 2
    new_warning_row = max(ranking_end, staff_end) + 2

    # === 写排名数据 A-I (行4开始), I列=VLOOKUP ===
    from openpyxl.styles import PatternFill, Font
    FILL_SUB=PatternFill('solid',fgColor='D9E2F3');FILL_RED=PatternFill('solid',fgColor='FF6666')
    FILL_W=PatternFill('solid',fgColor='FFFFFF')
    FONT_B=Font(name='微软雅黑',size=10,bold=True);FONT_N=Font(name='微软雅黑',size=10)
    rn = 4 + row_offset
    for i, (st, name, status, team, days, rate_v, ach_v, tenure_v, is_sub, is_off) in enumerate(ranking_rows):
        vals = [st, name, status, team, days, rate_v, ach_v, tenure_v]
        # 确定这行的填充和字体
        if is_sub:
            row_fill,row_font=FILL_SUB,FONT_B
        elif is_off:
            row_fill,row_font=FILL_RED,FONT_N
        else:
            row_fill,row_font=FILL_W,FONT_N
        for ci, v in enumerate(vals, 1):
            cell=ws.cell(row=rn, column=ci)
            try:cell.value=v;cell.fill=row_fill;cell.font=row_font
            except:pass
        # I列：VLOOKUP查K-M人员安排的阶段（小计留空）
        try:
            if not is_sub:
                ws.cell(row=rn, column=9).value = f'=IFERROR(VLOOKUP(B{rn},K:L,2,0),\"\")'
            else:
                ws.cell(row=rn, column=9).value = None
            ws.cell(row=rn, column=9).fill = row_fill
            ws.cell(row=rn, column=9).font = row_font
        except: pass
        # 百分比格式
        if rate_v is not None and not is_sub:
            ws.cell(row=rn, column=6).number_format = '0.00%'
        if ach_v is not None and not is_sub:
            ws.cell(row=rn, column=7).number_format = '0.00%'
        rn += 1

    # Clear old ranking data
    for r in range(rn, warning_row):
        for c in range(1, 10):
            try:
                ws.cell(row=r, column=c).value = None
                ws.cell(row=r, column=c).fill = FILL_W
            except: pass

    # === 写人员安排 K-M ===
    from openpyxl.styles import PatternFill
    FILL_MAP={'D0':PatternFill('solid',fgColor='C6EFCE'),'D1':PatternFill('solid',fgColor='BDD7EE'),
           'S1':PatternFill('solid',fgColor='FCE4D6'),'S2':PatternFill('solid',fgColor='D9D9D9')}
    WHITE_FILL=PatternFill('solid',fgColor='FFFFFF')

    # 先清空 K-M 的旧数据
    clear_end = max(warning_row, staff_end + 5)
    for r in range(10 + row_offset, clear_end):
        for c in range(11, 14):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = WHITE_FILL

    # 写入新的人员安排（空槽只写阶段颜色和标签，留空姓名）
    sr = staff_start
    for name, stage, team in staff_rows:
        f=FILL_MAP.get(stage, WHITE_FILL)
        if name:  # 有人
            ws.cell(row=sr, column=11).value = name;ws.cell(row=sr, column=11).fill = f
            ws.cell(row=sr, column=12).value = stage;ws.cell(row=sr, column=12).fill = f
            ws.cell(row=sr, column=13).value = f'=VLOOKUP(K{sr},B:D,3,0)';ws.cell(row=sr, column=13).fill = f
        else:  # 空槽：只染色和标阶段
            ws.cell(row=sr, column=11).value = None;ws.cell(row=sr, column=11).fill = f
            ws.cell(row=sr, column=12).value = stage;ws.cell(row=sr, column=12).fill = f
            ws.cell(row=sr, column=13).value = None;ws.cell(row=sr, column=13).fill = f
        sr += 1

    # === 警告信：出勤>=3天 且 达成率<=40% ===
    warning_names = []
    for name, grp in data[data['status']!='离职'].groupby('name'):
        total_days = grp['days'].max()
        avg_ach = grp['ach'].mean()
        if total_days >= 3 and avg_ach <= 0.40:
            warning_names.append(name)
    warning_text = f'警告信名单：{", ".join(warning_names)}' if warning_names else '警告信名单：'
    # 使用前面计算好的警告信位置
    warn_row = new_warning_row

    # 清理模板里遗留的警告信
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and '警告信名单' in str(v):
            try:
                ws.unmerge_cells(f'A{r}:I{r+1}')
            except:
                pass
            for rr in range(r, min(r+2, ws.max_row+1)):
                for cc in range(1, 10):
                    ws.cell(rr, cc).value = None

    # 写新的警告信
    ws.cell(row=warn_row, column=1).value = warning_text
    try:
        ws.merge_cells(f'A{warn_row}:I{warn_row+1}')
    except:
        pass

    return ws

# ============================================================
# 填充三个排名Sheet
# ============================================================
print('  填充新客...')
fill_ranking_sheet(wb['新客'], new, '新客', new_assigned, nd0_in, nd1_in, nd0_n, nd1_n, buf_new_d0, buf_new_d1)
print('  填充老客...')
fill_ranking_sheet(wb['老客'], old, '老客', old_assigned, od0_in, od1_in, od0_n, od1_n, buf_old_d0, buf_old_d1)

# 整体Sheet: 复制新客Sheet + 粘贴老客内容
print('  更新整体...')
if '整体' in wb.sheetnames: del wb['整体']
ws_all = wb.copy_worksheet(wb['新客'])
ws_all.title = '整体'
ws_old = wb['老客']

# 新客数据末行
new_last = max(r for r in range(1, ws_all.max_row+1)
               if any(ws_all.cell(row=r,column=c).value is not None for c in range(1,15)))
gap = 3

# 粘贴老客
for r in range(1, ws_old.max_row+1):
    dr = new_last + gap + r
    for c in range(1, 15):
        sv = ws_old.cell(row=r, column=c).value
        if sv is not None:
            try:
                dst = ws_all.cell(row=dr, column=c)
                dst.value = sv
                dst.fill = copy.copy(ws_old.cell(row=r, column=c).fill)
                dst.font = copy.copy(ws_old.cell(row=r, column=c).font)
                dst.number_format = ws_old.cell(row=r, column=c).number_format
                dst.alignment = copy.copy(ws_old.cell(row=r, column=c).alignment)
            except: pass

# 修复老客区VLOOKUP公式行号
import re
for r in range(new_last+gap+1, ws_all.max_row+1):
    i9 = ws_all.cell(row=r, column=9).value
    m13 = ws_all.cell(row=r, column=13).value
    if isinstance(i9, str) and 'VLOOKUP' in i9:
        ws_all.cell(row=r, column=9).value = f'=IFERROR(VLOOKUP(B{r},K:L,2,0),\"\")'
    if isinstance(m13, str) and 'VLOOKUP' in m13:
        ws_all.cell(row=r, column=13).value = f'=VLOOKUP(K{r},B:D,3,0)'

# 复制老客合并单元格
for mc_str in [str(m) for m in ws_old.merged_cells.ranges]:
    m = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', mc_str)
    if m:
        c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        nr1, nr2 = new_last+gap+r1, new_last+gap+r2
        try: ws_all.merge_cells(f'{c1}{nr1}:{c2}{nr2}')
        except: pass

# 排序: 整体, 新客, 老客, 人数预估
order = ['整体','新客','老客','人数预估']
for i, name in enumerate(order):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

# ============================================================
# 人数预估: 只替换回收率数值和到期量数据，公式不动
# ============================================================
print('  更新人数预估...')
ws4 = wb['人数预估']

# 回收率 (B/D/F/H columns, rows 11-18)
rate_data_est = [
    [D2_BEFORE['分期新客'],nf_rates.get('D-2',0),nf_rates.get('D-1',0),0, nf_rates.get('D0',0),nf_rates.get('D1',0),nf_rates.get('S1',0),nf_rates.get('S2',0)],
    [D2_BEFORE['分期老客'],of_rates.get('D-2',0),of_rates.get('D-1',0),0, of_rates.get('D0',0),of_rates.get('D1',0),of_rates.get('S1',0),of_rates.get('S2',0)],
    [D2_BEFORE['单期新客'],nd_rates.get('D-2',0),nd_rates.get('D-1',0),0, nd_rates.get('D0',0),nd_rates.get('D1',0),nd_rates.get('S1',0),nd_rates.get('S2',0)],
    [D2_BEFORE['单期老客'],od_rates.get('D-2',0),od_rates.get('D-1',0),0, od_rates.get('D0',0),od_rates.get('D1',0),od_rates.get('S1',0),od_rates.get('S2',0)],
]
rate_cols_est = [2, 4, 6, 8]  # B, D, F, H

for i in range(8):  # 8 rows
    for bi, col in enumerate(rate_cols_est):
        ws4.cell(row=11+i, column=col).value = round(rate_data_est[bi][i], 4)

# 到期量 (O-R columns, rows 10-16)
for i, (d, o, p, q, r) in enumerate(daily_due):
    from datetime import datetime
    ws4.cell(row=10+i, column=14).value = datetime(d.year, d.month, d.day)
    ws4.cell(row=10+i, column=15).value = o
    ws4.cell(row=10+i, column=16).value = p
    ws4.cell(row=10+i, column=17).value = q
    ws4.cell(row=10+i, column=18).value = r

# 策略参数
ws4.cell(row=23, column=14).value = 0.2  # N23
ws4.cell(row=24, column=14).value = 0.3  # N24

# 标题日期
for sname in ['整体','新客','老客']:
    ws = wb[sname]
    ws.cell(row=1, column=1).value = f'{PERIOD} PL组员回收情况汇总'

# ============================================================
wb.save(OUT)
print(f'\n✅ 已生成: {OUT}')
print(f'  从模板复制格式, 仅替换数据值')
print(f'  新客:{new["name"].nunique()}人 老客:{old["name"].nunique()}人')
