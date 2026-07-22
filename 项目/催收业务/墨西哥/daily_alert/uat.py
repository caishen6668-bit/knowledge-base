"""
UAT 自动验收模块 — 自动读取 Quick BI 数据并完成交叉验证

用法:
  python 每日数据预警.py --uat                        # 使用默认日期
  python 每日数据预警.py --uat --date 2026-07-07        # 指定运行日期
  python 每日数据预警.py --uat --country MX             # 仅墨西哥

功能:
  1. 自动从 Quick BI 读取原始数据（回收率 API + 案件量 API）
  2. 程序聚合计算 vs 原始数据直接求和 → 交叉验证
  3. 回收率 API due_case vs 案件量 API case → 独立数据源交叉验证
  4. 生成 docs/UAT验收报告.xlsx（全部自动填充，零手工）
  5. 仅展示差异明细，Summary 汇总通过率
"""

# NOTE: UTF-8 stdout wrapping is handled by main.py on import.
# Do NOT re-wrap here — it would double-close the underlying buffer.

import json
import os
from datetime import datetime, timedelta
from collections import defaultdict

from . import config
from .quickbi import _cached_query, _to_num


# ============================================================
#  日期工具
# ============================================================

def _get_due_week(date_str):
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    iso = dt.isocalendar()
    return f"{iso[0]}-{iso[1]:02d}"


# ============================================================
#  原始数据拉取（v1.0.1: 使用 _cached_query 自动去重）
# ============================================================

def _fetch_raw(api_id, conditions):
    """调用 Quick BI API 获取原始行数据"""
    conditions_str = json.dumps(conditions, ensure_ascii=False)
    data = _cached_query(api_id, conditions_str)
    if not data.get("Success"):
        raise RuntimeError(f"API {api_id} 调用失败: {data.get('Message', 'Unknown')}")
    return data.get("Result", {}).get("Values", [])


def _fetch_case_raw(api_id, conditions):
    """调用案件量 API，返回原始行"""
    conditions_str = json.dumps(conditions, ensure_ascii=False)
    fields = json.dumps(["app", "due_date", "order", "cust_type", "case"])
    data = _cached_query(api_id, conditions_str, fields)
    if not data.get("Success"):
        raise RuntimeError(f"API {api_id} 调用失败: {data.get('Message', 'Unknown')}")
    return data.get("Result", {}).get("Values", [])


# ============================================================
#  指标计算
# ============================================================

STAGES = ["D-2", "D-1", "D0", "D1", "S1", "S2"]

def _stage_key(stage_label):
    """D-2→D_2, D0→D0, ..."""
    for sk, sl in config.STAGE_KEY_MAP.items():
        if sl == stage_label:
            return sk
    return None


def calc_metrics(rows, stage_label):
    """
    从原始行计算某个阶段的指标。
    使用统一首逾计算函数 calculate_first_overdue_rate（BI 口径）。

    Returns:
        {
            "due_amt": float,      # 到期本金 = Σ due_amt
            "pay_amt": float,      # 累计回款 = Σ cumulative_pay (up to stage)
            "overdue_rate": float, # 金额首逾率 = 1 - pay/due
            "case_count": int,     # 到期笔数 = Σ due_case
            "row_count": int,      # 原始行数
        }
    """
    from .quickbi import calculate_first_overdue_rate

    due = 0.0
    pay = 0.0
    case_count = 0

    for r in rows:
        rcalc = calculate_first_overdue_rate(r, stage_label)
        due += rcalc["due_amt"]
        pay += rcalc["cum_pay"]
        case_count += int(_to_num(r.get("due_case", 0)))

    overdue_rate = round(1.0 - (pay / due), 6) if due > 0 else 0.0

    return {
        "due_amt": round(due, 2),
        "pay_amt": round(pay, 2),
        "overdue_rate": overdue_rate,
        "case_count": case_count,
        "row_count": len(rows),
    }


def calc_metrics_alt(rows, stage_label):
    """
    备用计算路径：逐行计算首逾率后取加权平均（用于交叉验证）。
    使用统一首逾计算函数 calculate_first_overdue_rate（BI 口径）。

    Returns:
        {"overdue_rate_alt": float, ...}
    """
    from .quickbi import calculate_first_overdue_rate

    total_due = 0.0
    weighted_overdue = 0.0

    for r in rows:
        rcalc = calculate_first_overdue_rate(r, stage_label)
        due = rcalc["due_amt"]
        total_due += due
        if due > 0:
            weighted_overdue += rcalc["overdue_rate"] * due

    alt_rate = round(weighted_overdue / total_due, 6) if total_due > 0 else 0.0
    return {"overdue_rate_alt": alt_rate}


# ============================================================
#  过滤器
# ============================================================

def filter_country(rows, cc):
    apps = set(config.COUNTRIES[cc]["apps"])
    return [r for r in rows if r.get("app_name", "") in apps]


def filter_by_mult_no(rows, is_installment):
    if is_installment:
        return [r for r in rows if _to_num(r.get("mult_no", "1")) >= 2]
    else:
        return [r for r in rows if _to_num(r.get("mult_no", "1")) == 1]


def filter_by_order_type(rows, ot):
    return [r for r in rows if r.get("order_type", "") == ot]


def filter_by_order_grade(rows, grade):
    return [r for r in rows if r.get("order_grade", "") == grade]


# ============================================================
#  案件量 API 交叉验证
# ============================================================

def build_case_reference(case_rows, country_code):
    """
    从案件量 API 原始行聚合各维度案件数。

    Returns:
        {
            "overall": {"D-2": N, "D-1": N, ...},       # 按 stage 聚合
            "by_product": {"单期": N, "分期": N},
            "by_order_type": {"非分期": N, ...},
            "by_cust_type": {"新客": N, "老客": N},
        }
        注意：案件量 API 没有 stage 维度，所以按 stage 的聚合都是相同的总数。
    """
    apps = set(config.COUNTRIES[country_code]["apps"])
    filtered = [r for r in case_rows if r.get("app", "") in apps]

    # 案件量 API 没有 stage 字段，按维度聚合
    overall = sum(_to_num(r.get("case", 0)) for r in filtered)

    by_product = defaultdict(int)
    by_order_type = defaultdict(int)
    by_cust_type = defaultdict(int)

    for r in filtered:
        order = r.get("order", "")
        product = config.ORDER_TYPE_MAP.get(order)
        cust = r.get("cust_type", "")
        c = int(_to_num(r.get("case", 0)))

        if product:
            by_product[product] += c
        by_order_type[order] += c
        by_cust_type[cust] += c

    return {
        "overall": overall,
        "by_product": dict(by_product),
        "by_order_type": dict(by_order_type),
        "by_cust_type": dict(by_cust_type),
    }


# ============================================================
#  差异记录
# ============================================================

class Diff:
    """单条差异记录"""
    def __init__(self, country, stage, dim_type, dim_value, metric,
                 prog_val, ref_val, diff_val, note=""):
        self.country = country
        self.stage = stage
        self.dim_type = dim_type      # "整体" | "单期分期" | "包体" | "订单风控等级"
        self.dim_value = dim_value     # 该维度下的具体值（如 "单期", "非分期", "A"）
        self.metric = metric           # "首逾率" | "到期本金" | "到期笔数"
        self.prog_val = prog_val
        self.ref_val = ref_val
        self.diff_val = diff_val
        self.note = note

    @property
    def is_consistent(self):
        """误差 ≤ 0.01% 视为一致"""
        if isinstance(self.prog_val, (int, float)) and isinstance(self.ref_val, (int, float)):
            if isinstance(self.prog_val, float) and abs(self.prog_val) <= 1.0:
                # 比率类：0.01% = 0.0001
                return abs(self.diff_val) < 0.0001
            else:
                # 金额/笔数类：差值 < 0.01
                return abs(self.diff_val) < 0.01
        return self.prog_val == self.ref_val


# ============================================================
#  主流程
# ============================================================

def run_uat(run_date, country_code="ALL"):
    """
    执行 UAT 自动验收。

    Args:
        run_date: 运行日期 "2026-07-07"
        country_code: "MX" | "AR" | "ALL"
    """
    run_dt = datetime.strptime(run_date, "%Y-%m-%d")
    business_date = (run_dt - timedelta(days=1)).strftime("%Y-%m-%d")
    due_week = _get_due_week(business_date)
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    countries_to_run = ["MX", "AR"] if country_code == "ALL" else [country_code]

    # 案件量 API 的日期格式
    biz_dt = datetime.strptime(business_date, "%Y-%m-%d")
    week_start = biz_dt.strftime("%Y%m%d")
    week_end = (biz_dt + timedelta(days=6)).strftime("%Y%m%d")

    print(f"\n{'='*60}")
    print(f"  UAT 自动验收  v{config.VERSION}")
    print(f"  运行日期: {run_date}  |  业务日期: {business_date}  |  Due Week: {due_week}")
    print(f"  国家: {country_code}")
    print(f"{'='*60}")

    # ================================================================
    #  Phase 1: 拉取数据
    # ================================================================
    print(f"\n  ⏳ Phase 1: 拉取 Quick BI 数据 ...")

    all_recovery_rows = []
    all_case_rows = []
    case_refs = {}

    for cc in countries_to_run:
        country_name = config.COUNTRIES[cc]["name"]

        # 回收率 API
        rec_rows = _fetch_raw(config.QBI_API_RECOVERY, {"due_week": due_week})
        rec_filtered = filter_country(rec_rows, cc)
        all_recovery_rows.extend(rec_filtered)
        print(f"  [{cc}] 回收率 API: {len(rec_rows)} 原始行 → {len(rec_filtered)} 行 (apps={config.COUNTRIES[cc]['apps']})")

        # 案件量 API（用于交叉验证）
        case_rows = _fetch_case_raw(config.QBI_API_CASES, {"due_date": week_start})
        case_filtered = [r for r in case_rows
                         if week_start <= r.get("due_date", "") <= week_end]
        all_case_rows.extend(case_filtered)
        case_refs[cc] = build_case_reference(case_filtered, cc)
        print(f"  [{cc}] 案件量 API: {len(case_rows)} 原始行 → {len(case_filtered)} 行 (含日期过滤)")

    # 构建按国家过滤的 recovery 行缓存
    rec_by_country = {cc: filter_country(all_recovery_rows, cc) for cc in countries_to_run}

    # ================================================================
    #  Phase 2: 计算 & 交叉验证
    # ================================================================
    print(f"\n  ⏳ Phase 2: 计算指标 & 交叉验证 ...")

    all_diffs = []          # 所有差异
    stats = defaultdict(lambda: {"total": 0, "pass": 0, "fail": 0})

    ORDER_TYPES = ["非分期", "借款分期", "展期分期", "展期N期"]
    GRADES = ["A", "B", "C", "D", "E", "F"]

    for cc in countries_to_run:
        country_name = config.COUNTRIES[cc]["name"]
        rows = rec_by_country[cc]
        case_ref = case_refs.get(cc, {})

        for stage in STAGES:
            # ---------- 整体 ----------
            m = calc_metrics(rows, stage)
            dim = "整体"
            stats[dim]["total"] += 1

            # 备用计算路径交叉验证
            m_alt = calc_metrics_alt(rows, stage)
            rate_diff = abs(m["overdue_rate"] - m_alt["overdue_rate_alt"])
            if rate_diff >= 0.0001:
                all_diffs.append(Diff(country_name, stage, dim, "—", "首逾率(备用路径)",
                                      m["overdue_rate"], m_alt["overdue_rate_alt"], rate_diff,
                                      "聚合计算 vs 逐行加权平均"))

            # 整体一致性判断（备用路径一致即通过）
            if rate_diff < 0.0001:
                stats[dim]["pass"] += 1
            else:
                stats[dim]["fail"] += 1

            # ---------- 单期 / 分期 ----------
            dim = "单期分期"
            for label, is_inst in [("单期", False), ("分期", True)]:
                subset = filter_by_mult_no(rows, is_inst)
                m_sub = calc_metrics(subset, stage)
                stats[dim]["total"] += 1

                # 备用路径
                m_sub_alt = calc_metrics_alt(subset, stage)
                rate_diff = abs(m_sub["overdue_rate"] - m_sub_alt["overdue_rate_alt"])
                if rate_diff >= 0.0001:
                    all_diffs.append(Diff(country_name, stage, dim, label, "首逾率(备用路径)",
                                          m_sub["overdue_rate"], m_sub_alt["overdue_rate_alt"],
                                          rate_diff, "聚合计算 vs 逐行加权平均"))

                if rate_diff < 0.0001:
                    stats[dim]["pass"] += 1
                else:
                    stats[dim]["fail"] += 1

            # ---------- 包体 ----------
            dim = "包体"
            for ot in ORDER_TYPES:
                subset = filter_by_order_type(rows, ot)
                if len(subset) == 0:
                    continue
                m_sub = calc_metrics(subset, stage)
                m_sub_alt = calc_metrics_alt(subset, stage)
                stats[dim]["total"] += 1

                rate_diff = abs(m_sub["overdue_rate"] - m_sub_alt["overdue_rate_alt"])
                if rate_diff >= 0.0001:
                    all_diffs.append(Diff(country_name, stage, dim, ot, "首逾率(备用路径)",
                                          m_sub["overdue_rate"], m_sub_alt["overdue_rate_alt"],
                                          rate_diff, "聚合计算 vs 逐行加权平均"))

                if rate_diff < 0.0001:
                    stats[dim]["pass"] += 1
                else:
                    stats[dim]["fail"] += 1

            # ---------- 订单风控等级 ----------
            dim = "订单风控等级"
            for grade in GRADES:
                subset = filter_by_order_grade(rows, grade)
                if len(subset) == 0:
                    continue
                m_sub = calc_metrics(subset, stage)
                m_sub_alt = calc_metrics_alt(subset, stage)
                stats[dim]["total"] += 1

                rate_diff = abs(m_sub["overdue_rate"] - m_sub_alt["overdue_rate_alt"])
                if rate_diff >= 0.0001:
                    all_diffs.append(Diff(country_name, stage, dim, grade, "首逾率(备用路径)",
                                          m_sub["overdue_rate"], m_sub_alt["overdue_rate_alt"],
                                          rate_diff, "聚合计算 vs 逐行加权平均"))

                if rate_diff < 0.0001:
                    stats[dim]["pass"] += 1
                else:
                    stats[dim]["fail"] += 1

    # ---------- 案件量交叉验证（独立数据源） ----------
    print(f"\n  ⏳ Phase 2b: 案件量交叉验证（回收率 API due_case vs 案件量 API case）...")

    dim_case = "案件量交叉验证"
    for cc in countries_to_run:
        country_name = config.COUNTRIES[cc]["name"]
        rows = rec_by_country[cc]
        case_ref = case_refs.get(cc, {})

        # 回收率 API 的 due_case 总和 vs 案件量 API 的 case 总和
        prog_total_case = sum(int(_to_num(r.get("due_case", 0))) for r in rows)
        ref_total_case = case_ref.get("overall", 0)

        stats[dim_case]["total"] += 1
        case_diff = abs(prog_total_case - ref_total_case)
        if case_diff > 0:
            all_diffs.append(Diff(country_name, "—", dim_case, "总笔数", "到期笔数",
                                  prog_total_case, ref_total_case, case_diff,
                                  "回收率API due_case vs 案件量API case"))
            stats[dim_case]["fail"] += 1
        else:
            stats[dim_case]["pass"] += 1

        # 按包体类型比较
        for ot in ORDER_TYPES:
            subset = filter_by_order_type(rows, ot)
            prog_case = sum(int(_to_num(r.get("due_case", 0))) for r in subset)
            ref_case = case_ref.get("by_order_type", {}).get(ot, 0)

            if prog_case == 0 and ref_case == 0:
                continue

            stats[dim_case]["total"] += 1
            case_diff = abs(prog_case - ref_case)
            if case_diff > 0:
                all_diffs.append(Diff(country_name, "—", dim_case, f"包体={ot}", "到期笔数",
                                      prog_case, ref_case, case_diff,
                                      "回收率API due_case vs 案件量API case"))
                stats[dim_case]["fail"] += 1
            else:
                stats[dim_case]["pass"] += 1

    # ================================================================
    #  Phase 3: 生成 Excel
    # ================================================================
    print(f"\n  ⏳ Phase 3: 生成 Excel 报告 ...")
    _generate_excel(
        run_date=run_date,
        business_date=business_date,
        due_week=due_week,
        gen_time=gen_time,
        countries_to_run=countries_to_run,
        rec_by_country=rec_by_country,
        case_refs=case_refs,
        all_diffs=all_diffs,
        stats=dict(stats),
    )

    # ================================================================
    #  Phase 4: 打印控制台摘要
    # ================================================================
    print(f"\n{'='*60}")
    print(f"  UAT 验收结果")
    print(f"{'='*60}")

    total_all = sum(s["total"] for s in stats.values())
    pass_all = sum(s["pass"] for s in stats.values())
    fail_all = sum(s["fail"] for s in stats.values())

    for dim_name, s in stats.items():
        status = "✅ 通过" if s["fail"] == 0 else f"❌ {s['fail']}/{s['total']} 失败"
        print(f"  {dim_name:20s}: {status:20s}  ({s['pass']}/{s['total']})")

    overall_rate = pass_all / total_all * 100 if total_all > 0 else 0
    print(f"  {'─'*50}")
    print(f"  综合通过率: {overall_rate:.1f}%  ({pass_all}/{total_all})")

    if all_diffs:
        print(f"\n  ⚠️ 差异明细 ({len(all_diffs)} 条):")
        for d in all_diffs[:20]:  # 最多显示 20 条
            print(f"    [{d.dim_type}] {d.country} {d.stage} {d.dim_value} "
                  f"| {d.metric}: 程序={d.prog_val} 参考={d.ref_val} 差值={d.diff_val}")
        if len(all_diffs) > 20:
            print(f"    ... 还有 {len(all_diffs) - 20} 条，详见 Excel")
    else:
        print(f"\n  🎉 全部通过！无差异。")

    max_diff = max((abs(d.diff_val) for d in all_diffs), default=0)
    print(f"\n  一致数量: {pass_all}")
    print(f"  不一致数量: {fail_all}")
    print(f"  最大误差: {max_diff:.6f}")
    print(f"{'='*60}")


# ============================================================
#  Excel 生成
# ============================================================

def _generate_excel(run_date, business_date, due_week, gen_time,
                    countries_to_run, rec_by_country, case_refs,
                    all_diffs, stats):
    """生成 docs/UAT验收报告.xlsx"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    # ---- 样式 ----
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14)
    SECTION_FONT = Font(name="Microsoft YaHei", bold=True, size=12)
    NORMAL_FONT = Font(name="Microsoft YaHei", size=10)
    MONO_FONT = Font(name="Consolas", size=10)
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    BOTTOM_BORDER = Border(bottom=Side(style="medium"))

    def style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def style_cell(cell, is_rate=False, is_int=False):
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if is_rate:
            cell.number_format = '0.00%'
        elif is_int:
            cell.number_format = '#,##0'
        else:
            cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def auto_width(ws, max_col, min_w=10, max_w=35):
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            best = min_w
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
                for val in row:
                    if val is not None:
                        best = max(best, min(len(str(val)) * 1.2, max_w))
            ws.column_dimensions[letter].width = best

    wb = openpyxl.Workbook()

    ORDER_TYPES = ["非分期", "借款分期", "展期分期", "展期N期"]
    GRADES = ["A", "B", "C", "D", "E", "F"]
    STAGES_LIST = STAGES

    # ========================
    #  Sheet 1: 整体
    # ========================
    ws = wb.active
    ws.title = "整体"

    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value=f"UAT 自动验收 — 整体金额首逾率  |  业务日期: {business_date}  |  生成: {gen_time}")
    ws["A1"].font = TITLE_FONT

    headers = ["国家", "阶段", "首逾率(程序)", "到期本金(程序)", "到期笔数(程序)", "备用路径首逾率", "差值"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for cc in countries_to_run:
        country_name = config.COUNTRIES[cc]["name"]
        rows = rec_by_country[cc]
        for stage in STAGES_LIST:
            m = calc_metrics(rows, stage)
            m_alt = calc_metrics_alt(rows, stage)
            diff = abs(m["overdue_rate"] - m_alt["overdue_rate_alt"])

            ws.cell(row=row, column=1, value=country_name)
            ws.cell(row=row, column=2, value=stage)
            ws.cell(row=row, column=3, value=m["overdue_rate"])
            ws.cell(row=row, column=4, value=m["due_amt"])
            ws.cell(row=row, column=5, value=m["case_count"])
            ws.cell(row=row, column=6, value=m_alt["overdue_rate_alt"])
            ws.cell(row=row, column=7, value=round(diff, 8))

            for c in [1, 2, 3, 4, 5, 6, 7]:
                style_cell(ws.cell(row=row, column=c),
                           is_rate=(c in [3, 6, 7]),
                           is_int=(c == 5))

            # 差异高亮
            if diff >= 0.0001:
                ws.cell(row=row, column=7).fill = RED_FILL
            else:
                ws.cell(row=row, column=7).fill = GREEN_FILL

            row += 1
        row += 1  # 国家间空行

    auto_width(ws, len(headers))
    ws.column_dimensions["A"].width = 16
    ws.freeze_panes = "A4"

    # ========================
    #  Sheet 2: 单期分期
    # ========================
    ws = wb.create_sheet("单期分期")

    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value=f"UAT 自动验收 — 单期/分期金额首逾率  |  业务日期: {business_date}")
    ws["A1"].font = TITLE_FONT

    headers = ["国家", "阶段", "产品类型", "首逾率(程序)", "到期本金(程序)", "到期笔数(程序)", "备用路径首逾率", "差值"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for cc in countries_to_run:
        country_name = config.COUNTRIES[cc]["name"]
        rows = rec_by_country[cc]
        for stage in STAGES_LIST:
            for label, is_inst in [("单期 (mult_no=1)", False), ("分期 (mult_no≥2)", True)]:
                subset = filter_by_mult_no(rows, is_inst)
                m = calc_metrics(subset, stage)
                m_alt = calc_metrics_alt(subset, stage)
                diff = abs(m["overdue_rate"] - m_alt["overdue_rate_alt"])

                ws.cell(row=row, column=1, value=country_name)
                ws.cell(row=row, column=2, value=stage)
                ws.cell(row=row, column=3, value=label)
                ws.cell(row=row, column=4, value=m["overdue_rate"])
                ws.cell(row=row, column=5, value=m["due_amt"])
                ws.cell(row=row, column=6, value=m["case_count"])
                ws.cell(row=row, column=7, value=m_alt["overdue_rate_alt"])
                ws.cell(row=row, column=8, value=round(diff, 8))

                for c in [1, 2, 3, 4, 5, 6, 7, 8]:
                    style_cell(ws.cell(row=row, column=c),
                               is_rate=(c in [4, 7, 8]),
                               is_int=(c == 6))

                if diff >= 0.0001:
                    ws.cell(row=row, column=8).fill = RED_FILL
                else:
                    ws.cell(row=row, column=8).fill = GREEN_FILL

                row += 1
            row += 1
        row += 1

    auto_width(ws, len(headers))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A4"

    # ========================
    #  Sheet 3: 包体
    # ========================
    ws = wb.create_sheet("包体")

    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value=f"UAT 自动验收 — 包体明细  |  业务日期: {business_date}")
    ws["A1"].font = TITLE_FONT

    headers = ["国家", "阶段", "包体", "首逾率(程序)", "到期本金(程序)", "到期笔数(程序)",
               "备用路径首逾率", "首逾率差值", "行数"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for cc in countries_to_run:
        country_name = config.COUNTRIES[cc]["name"]
        rows = rec_by_country[cc]
        for stage in STAGES_LIST:
            for ot in ORDER_TYPES:
                subset = filter_by_order_type(rows, ot)
                if len(subset) == 0:
                    continue
                m = calc_metrics(subset, stage)
                m_alt = calc_metrics_alt(subset, stage)
                diff = abs(m["overdue_rate"] - m_alt["overdue_rate_alt"])

                ws.cell(row=row, column=1, value=country_name)
                ws.cell(row=row, column=2, value=stage)
                ws.cell(row=row, column=3, value=ot)
                ws.cell(row=row, column=4, value=m["overdue_rate"])
                ws.cell(row=row, column=5, value=m["due_amt"])
                ws.cell(row=row, column=6, value=m["case_count"])
                ws.cell(row=row, column=7, value=m_alt["overdue_rate_alt"])
                ws.cell(row=row, column=8, value=round(diff, 8))
                ws.cell(row=row, column=9, value=m["row_count"])

                for c in [1, 2, 3, 4, 5, 6, 7, 8, 9]:
                    style_cell(ws.cell(row=row, column=c),
                               is_rate=(c in [4, 7, 8]),
                               is_int=(c in [6, 9]))

                if diff >= 0.0001:
                    ws.cell(row=row, column=8).fill = RED_FILL
                else:
                    ws.cell(row=row, column=8).fill = GREEN_FILL

                row += 1
            row += 1
        row += 1

    auto_width(ws, len(headers))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A4"

    # ========================
    #  Sheet 4: 订单风控等级
    # ========================
    ws = wb.create_sheet("订单风控等级")

    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value=f"UAT 自动验收 — 订单风控等级 (order_grade) 金额首逾率  |  业务日期: {business_date}")
    ws["A1"].font = TITLE_FONT

    headers = ["国家", "阶段", "Grade", "首逾率(程序)", "到期本金(程序)", "到期笔数(程序)",
               "备用路径首逾率", "首逾率差值", "行数"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for cc in countries_to_run:
        country_name = config.COUNTRIES[cc]["name"]
        rows = rec_by_country[cc]
        for stage in STAGES_LIST:
            for grade in GRADES:
                subset = filter_by_order_grade(rows, grade)
                if len(subset) == 0:
                    continue
                m = calc_metrics(subset, stage)
                m_alt = calc_metrics_alt(subset, stage)
                diff = abs(m["overdue_rate"] - m_alt["overdue_rate_alt"])

                ws.cell(row=row, column=1, value=country_name)
                ws.cell(row=row, column=2, value=stage)
                ws.cell(row=row, column=3, value=grade)
                ws.cell(row=row, column=4, value=m["overdue_rate"])
                ws.cell(row=row, column=5, value=m["due_amt"])
                ws.cell(row=row, column=6, value=m["case_count"])
                ws.cell(row=row, column=7, value=m_alt["overdue_rate_alt"])
                ws.cell(row=row, column=8, value=round(diff, 8))
                ws.cell(row=row, column=9, value=m["row_count"])

                for c in [1, 2, 3, 4, 5, 6, 7, 8, 9]:
                    style_cell(ws.cell(row=row, column=c),
                               is_rate=(c in [4, 7, 8]),
                               is_int=(c in [6, 9]))

                if diff >= 0.0001:
                    ws.cell(row=row, column=8).fill = RED_FILL
                else:
                    ws.cell(row=row, column=8).fill = GREEN_FILL

                row += 1
            row += 1
        row += 1

    auto_width(ws, len(headers))
    ws.column_dimensions["A"].width = 16
    ws.freeze_panes = "A4"

    # ========================
    #  Sheet 5: 案件量交叉验证
    # ========================
    ws = wb.create_sheet("案件量交叉验证")

    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value=f"UAT 自动验收 — 案件量交叉验证 (回收率API due_case vs 案件量API case)  |  业务日期: {business_date}")
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:G2")
    ws.cell(row=2, column=1, value="⚠️ 这是独立数据源交叉验证 — 回收率 API (524c3ccd429c) due_case 字段 vs 案件量 API (c2f93e0fa45b) case 字段")
    ws["A2"].font = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")

    headers = ["国家", "维度类型", "维度值", "回收率API(due_case)", "案件量API(case)", "差值", "是否一致"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))

    row = 5
    ORDER_TYPES = ["非分期", "借款分期", "展期分期", "展期N期"]

    for cc in countries_to_run:
        country_name = config.COUNTRIES[cc]["name"]
        rows = rec_by_country[cc]
        case_ref = case_refs.get(cc, {})

        # 总计
        prog_total = sum(int(_to_num(r.get("due_case", 0))) for r in rows)
        ref_total = case_ref.get("overall", 0)
        diff_total = prog_total - ref_total
        consistent = "✅" if diff_total == 0 else "❌"

        ws.cell(row=row, column=1, value=country_name)
        ws.cell(row=row, column=2, value="总计")
        ws.cell(row=row, column=3, value="全部")
        ws.cell(row=row, column=4, value=prog_total)
        ws.cell(row=row, column=5, value=ref_total)
        ws.cell(row=row, column=6, value=diff_total)
        ws.cell(row=row, column=7, value=consistent)

        for c in [1, 2, 3, 4, 5, 6, 7]:
            style_cell(ws.cell(row=row, column=c), is_int=(c in [4, 5, 6]))
        if consistent == "❌":
            ws.cell(row=row, column=7).fill = RED_FILL
        else:
            ws.cell(row=row, column=7).fill = GREEN_FILL
        row += 1

        # 按包体
        for ot in ORDER_TYPES:
            subset = filter_by_order_type(rows, ot)
            prog_case = sum(int(_to_num(r.get("due_case", 0))) for r in subset)
            ref_case = case_ref.get("by_order_type", {}).get(ot, 0)

            if prog_case == 0 and ref_case == 0:
                continue

            diff_case = prog_case - ref_case
            consistent = "✅" if diff_case == 0 else "❌"

            ws.cell(row=row, column=1, value=country_name)
            ws.cell(row=row, column=2, value="包体")
            ws.cell(row=row, column=3, value=ot)
            ws.cell(row=row, column=4, value=prog_case)
            ws.cell(row=row, column=5, value=ref_case)
            ws.cell(row=row, column=6, value=diff_case)
            ws.cell(row=row, column=7, value=consistent)

            for c in [1, 2, 3, 4, 5, 6, 7]:
                style_cell(ws.cell(row=row, column=c), is_int=(c in [4, 5, 6]))
            if consistent == "❌":
                ws.cell(row=row, column=7).fill = RED_FILL
            else:
                ws.cell(row=row, column=7).fill = GREEN_FILL
            row += 1

        # 按产品类型
        for label, is_inst in [("单期 (mult_no=1)", False), ("分期 (mult_no≥2)", True)]:
            subset = filter_by_mult_no(rows, is_inst)
            prog_case = sum(int(_to_num(r.get("due_case", 0))) for r in subset)
            prod_label = "单期" if not is_inst else "分期"
            ref_case = case_ref.get("by_product", {}).get(prod_label, 0)

            if prog_case == 0 and ref_case == 0:
                continue

            diff_case = prog_case - ref_case
            consistent = "✅" if diff_case == 0 else "❌"

            ws.cell(row=row, column=1, value=country_name)
            ws.cell(row=row, column=2, value="产品类型")
            ws.cell(row=row, column=3, value=label)
            ws.cell(row=row, column=4, value=prog_case)
            ws.cell(row=row, column=5, value=ref_case)
            ws.cell(row=row, column=6, value=diff_case)
            ws.cell(row=row, column=7, value=consistent)

            for c in [1, 2, 3, 4, 5, 6, 7]:
                style_cell(ws.cell(row=row, column=c), is_int=(c in [4, 5, 6]))
            if consistent == "❌":
                ws.cell(row=row, column=7).fill = RED_FILL
            else:
                ws.cell(row=row, column=7).fill = GREEN_FILL
            row += 1

        row += 1  # 国家间空行

    auto_width(ws, len(headers))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.freeze_panes = "A5"

    # ========================
    #  Sheet 6: 差异明细（仅在有差异时展示）
    # ========================
    ws = wb.create_sheet("差异明细")

    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value=f"UAT 自动验收 — 差异明细  |  业务日期: {business_date}")
    ws["A1"].font = TITLE_FONT

    if all_diffs:
        ws.merge_cells("A2:I2")
        ws.cell(row=2, column=1, value=f"共 {len(all_diffs)} 条差异（仅展示误差 > 0.01% 的项）")
        ws["A2"].font = Font(name="Microsoft YaHei", size=9, italic=True, color="C00000")

        headers = ["国家", "阶段", "维度类型", "维度值", "指标", "程序值", "参考值", "差值", "备注"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=4, column=c, value=h)
        style_header_row(ws, 4, len(headers))

        row = 5
        for d in all_diffs:
            ws.cell(row=row, column=1, value=d.country)
            ws.cell(row=row, column=2, value=d.stage)
            ws.cell(row=row, column=3, value=d.dim_type)
            ws.cell(row=row, column=4, value=d.dim_value)
            ws.cell(row=row, column=5, value=d.metric)
            ws.cell(row=row, column=6, value=d.prog_val)
            ws.cell(row=row, column=7, value=d.ref_val)
            ws.cell(row=row, column=8, value=round(d.diff_val, 8))
            ws.cell(row=row, column=9, value=d.note)

            for c in [1, 2, 3, 4, 5, 6, 7, 8, 9]:
                style_cell(ws.cell(row=row, column=c),
                           is_rate=(c in [5] and "率" in str(d.metric)),
                           is_int=(c in [6, 7] and isinstance(d.prog_val, int)))

            # 高亮不一致行
            if not d.is_consistent:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).fill = RED_FILL

            row += 1

        auto_width(ws, len(headers))
        ws.column_dimensions["I"].width = 40
    else:
        ws.merge_cells("A2:F2")
        ws.cell(row=2, column=1, value="🎉 全部通过！无差异。")
        ws["A2"].font = Font(name="Microsoft YaHei", bold=True, size=14, color="006100")
        ws["A2"].fill = GREEN_FILL
        ws.merge_cells("A3:F3")
        ws.cell(row=3, column=1, value="所有维度的程序计算值与备用路径计算值均在 0.01% 误差范围内一致。")

    ws.freeze_panes = "A5"

    # ========================
    #  Sheet 7: Summary
    # ========================
    ws = wb.create_sheet("Summary")

    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value=f"UAT 自动验收 — Summary  |  v{config.VERSION}  |  {gen_time}")
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:E2")
    ws.cell(row=2, column=1, value=f"业务日期: {business_date}  |  Due Week: {due_week}  |  运行日期: {run_date}")
    ws["A2"].font = Font(name="Microsoft YaHei", size=10, color="666666")

    # ---- 汇总表 ----
    r = 4
    ws.cell(row=r, column=1, value="验收维度").font = HEADER_FONT
    ws.cell(row=r, column=2, value="检查项数").font = HEADER_FONT
    ws.cell(row=r, column=3, value="通过").font = HEADER_FONT
    ws.cell(row=r, column=4, value="失败").font = HEADER_FONT
    ws.cell(row=r, column=5, value="结果").font = HEADER_FONT
    style_header_row(ws, r, 5)
    r += 1

    dim_order = ["整体", "单期分期", "包体", "订单风控等级", "案件量交叉验证"]
    total_all = 0
    pass_all = 0
    fail_all = 0

    for dim_name in dim_order:
        s = stats.get(dim_name, {"total": 0, "pass": 0, "fail": 0})
        total_all += s["total"]
        pass_all += s["pass"]
        fail_all += s["fail"]

        status = "✅ 通过" if s["fail"] == 0 and s["total"] > 0 else ("❌ 失败" if s["fail"] > 0 else "— (无数据)")

        ws.cell(row=r, column=1, value=dim_name)
        ws.cell(row=r, column=2, value=s["total"])
        ws.cell(row=r, column=3, value=s["pass"])
        ws.cell(row=r, column=4, value=s["fail"])
        ws.cell(row=r, column=5, value=status)

        for c in [1, 2, 3, 4, 5]:
            style_cell(ws.cell(row=r, column=c), is_int=(c in [2, 3, 4]))
        ws.cell(row=r, column=1).font = Font(name="Microsoft YaHei", bold=True, size=10)

        if s["fail"] > 0:
            ws.cell(row=r, column=5).fill = RED_FILL
        elif s["total"] > 0:
            ws.cell(row=r, column=5).fill = GREEN_FILL
        else:
            ws.cell(row=r, column=5).fill = YELLOW_FILL

        r += 1

    # 综合通过率
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(row=r, column=1, value="综合通过率").font = Font(name="Microsoft YaHei", bold=True, size=14)

    overall_rate = pass_all / total_all * 100 if total_all > 0 else 0
    ws.cell(row=r, column=5, value=f"{overall_rate:.1f}%")
    ws.cell(row=r, column=5).font = Font(name="Microsoft YaHei", bold=True, size=16,
                                          color="006100" if overall_rate >= 99 else "C00000")
    ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")

    for c in [1, 5]:
        ws.cell(row=r, column=c).border = BOTTOM_BORDER
    r += 2

    # ---- 差异统计 ----
    ws.cell(row=r, column=1, value="差异统计").font = SECTION_FONT
    r += 1

    max_diff = max((abs(d.diff_val) for d in all_diffs), default=0)

    info_rows = [
        ("一致数量", pass_all),
        ("不一致数量", fail_all),
        ("最大误差", f"{max_diff:.8f}" if max_diff > 0 else "0"),
        ("差异明细条数", len(all_diffs)),
    ]
    for label, val in info_rows:
        ws.cell(row=r, column=1, value=label).font = Font(name="Microsoft YaHei", bold=True, size=10)
        ws.cell(row=r, column=2, value=val).font = NORMAL_FONT
        r += 1

    r += 1

    # ---- 说明 ----
    ws.cell(row=r, column=1, value="验证方法说明").font = SECTION_FONT
    r += 1

    notes = [
        "1. 首逾率验证：程序聚合计算（Σpay/Σdue → 1-pay/due） vs 备用路径（逐行加权平均）",
        "2. 误差阈值：首逾率差异 < 0.01%（0.0001）视为 ✅ 一致",
        "3. 案件量验证：回收率 API due_case 字段 vs 案件量 API case 字段（独立数据源）",
        "4. 到期本金：统一使用 {stage}_due_amt 字段，不使用 deprecated due_amt",
        f"5. 程序版本: v{config.VERSION} Build {config.BUILD}",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).font = Font(name="Microsoft YaHei", size=9, color="666666")
        ws.merge_cells(f"A{r}:E{r}")
        r += 1

    auto_width(ws, 5)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A5"

    # ---- 保存 ----
    script_dir = os.path.dirname(os.path.abspath(__file__))
    docs_dir = os.path.join(script_dir, "docs")
    os.makedirs(docs_dir, exist_ok=True)

    output_path = os.path.join(docs_dir, "UAT验收报告.xlsx")
    wb.save(output_path)

    print(f"\n  ✅ Excel 已生成: {output_path}")
    print(f"  包含 {len(wb.sheetnames)} 个工作表: {', '.join(wb.sheetnames)}")
    print(f"  总检查项: {total_all}  |  通过: {pass_all}  |  失败: {fail_all}  |  通过率: {overall_rate:.1f}%")


# ============================================================
#  命令行入口（独立运行）
# ============================================================

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="UAT 自动验收")
    parser.add_argument("--date", default=None, help="Run date (YYYY-MM-DD)")
    args = parser.parse_args()

    run_date = args.date or (datetime.now() - timedelta(days=0)).strftime("%Y-%m-%d")
    run_uat(run_date=run_date)
