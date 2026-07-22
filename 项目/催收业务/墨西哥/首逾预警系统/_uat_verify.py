"""
UAT 验收 — 生成 Excel 报告

用法:
  python _uat_verify.py

输出: docs/UAT验收.xlsx
"""
import io as _io
import sys as _sys
if _sys.platform == "win32" and hasattr(_sys.stdout, 'buffer'):
    _sys.stdout = _io.TextIOWrapper(_sys.stdout.buffer, encoding="utf-8", errors="replace")

import json
import os
from datetime import datetime, timedelta
from collections import defaultdict

_sys.path.insert(0, "D:/knowledge-base/scripts")
from daily_alert.config import (
    QBI_API_RECOVERY, COUNTRIES, STAGE_KEY_MAP, ORDER_TYPE_MAP, VERSION,
)
from daily_alert.quickbi import _sign_and_call, _to_num

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

# ============================================================
#  日期
# ============================================================
BUSINESS_DATE = "2026-07-06"
biz_dt = datetime.strptime(BUSINESS_DATE, "%Y-%m-%d")
DUE_WEEK = f"{biz_dt.isocalendar()[0]}-{biz_dt.isocalendar()[1]:02d}"
RUN_DATE = (biz_dt + timedelta(days=1)).strftime("%Y-%m-%d")
GEN_TIME = datetime.now().strftime("%Y-%m-%d %H:%M")

STAGES = ["D-2", "D-1", "D0", "D1", "S1", "S2"]

print(f"UAT 验收 — 生成 Excel")
print(f"  业务日期: {BUSINESS_DATE}  |  Due Week: {DUE_WEEK}")
print(f"  运行日期: {RUN_DATE}  |  生成时间: {GEN_TIME}")

# ============================================================
#  数据拉取
# ============================================================
print("拉取数据...", end=" ", flush=True)
extra = {"ApiId": QBI_API_RECOVERY, "Conditions": json.dumps({"due_week": DUE_WEEK}, ensure_ascii=False)}
data = _sign_and_call("QueryDataService", extra)
all_rows = data.get("Result", {}).get("Values", [])
print(f"{len(all_rows)} 行")

# ============================================================
#  辅助函数
# ============================================================

def filter_country(rows, cc):
    apps = set(COUNTRIES[cc]["apps"])
    return [r for r in rows if r.get("app_name", "") in apps]

def filter_by_mult_no(rows, is_installment):
    if is_installment:
        return [r for r in rows if _to_num(r.get("mult_no", "1")) >= 2]
    else:
        return [r for r in rows if _to_num(r.get("mult_no", "1")) == 1]

def filter_by_order_type(rows, ot):
    return [r for r in rows if r.get("order_type", "") == ot]

def filter_by_order_grade(rows, grade):
    return [r for r in rows if r.get("order_grade", "") == grade]

def calc_stage_metrics(rows, stage_label):
    stage_key = None
    for sk, sl in STAGE_KEY_MAP.items():
        if sl == stage_label:
            stage_key = sk
            break
    if stage_key is None:
        return {"overdue_rate": 0, "due_amt": 0, "pay_amt": 0, "case_count": 0, "row_count": 0}

    due = sum(_to_num(r.get(f"{stage_key}_due_amt", 0)) for r in rows)
    pay = sum(_to_num(r.get(f"{stage_key}_pay_amt", 0)) for r in rows)
    case_count = sum(int(_to_num(r.get("due_case", 0))) for r in rows)

    return {
        "overdue_rate": round(1.0 - (pay / due), 6) if due > 0 else 0.0,
        "due_amt": round(due, 2),
        "pay_amt": round(pay, 2),
        "case_count": case_count,
        "row_count": len(rows),
    }

# ============================================================
#  Excel 样式
# ============================================================

HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14)
NORMAL_FONT = Font(name="Microsoft YaHei", size=10)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header(ws, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data_cell(cell, is_rate=False):
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    if is_rate:
        cell.number_format = '0.00%'
        cell.alignment = Alignment(horizontal="center")
    else:
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws, max_col, min_width=10, max_width=30):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for val in row:
                if val:
                    max_len = max(max_len, min(len(str(val)) * 1.3, max_width))
        ws.column_dimensions[letter].width = max_len

# ============================================================
#  公式: IF(BI="", "", IF(ABS(程序-BI)<0.0001, "✅", "❌"))
# ============================================================

def match_formula(bi_col_letter, prog_col_letter, row):
    """生成 Excel 一致性检查公式"""
    bi_cell = f"{bi_col_letter}{row}"
    prog_cell = f"{prog_col_letter}{row}"
    return f'=IF({bi_cell}="","",IF(ABS({prog_cell}-{bi_cell})<0.0001,"✅","❌"))'

# ============================================================
#  创建工作簿
# ============================================================

wb = openpyxl.Workbook()

# ========================
#  Sheet 1: 整体
# ========================
ws1 = wb.active
ws1.title = "整体"

# 标题行
ws1.merge_cells("A1:H1")
ws1.cell(row=1, column=1, value=f"UAT 验收 — 整体金额首逾率  |  业务日期: {BUSINESS_DATE}  |  生成: {GEN_TIME}")
ws1["A1"].font = TITLE_FONT

# 表头 (row 3)
headers1 = ["国家", "阶段", "BI-首逾率", "程序-首逾率", "差值", "是否一致", "到期本金(程序)", "到期笔数(程序)"]
for c, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=c, value=h)
style_header(ws1, len(headers1))

row = 4
for cc in ["MX", "AR"]:
    country_name = COUNTRIES[cc]["name"]
    rows = filter_country(all_rows, cc)
    for stage in STAGES:
        m = calc_stage_metrics(rows, stage)
        ws1.cell(row=row, column=1, value=country_name)
        ws1.cell(row=row, column=2, value=stage)
        # BI值留空
        ws1.cell(row=row, column=4, value=m["overdue_rate"])  # 程序-首逾率
        ws1.cell(row=row, column=5, value=None)  # 差值 (formula below)
        ws1.cell(row=row, column=7, value=m["due_amt"])
        ws1.cell(row=row, column=8, value=m["case_count"])

        for c in [1, 2, 3, 4, 5, 7, 8]:
            style_data_cell(ws1.cell(row=row, column=c))
        style_data_cell(ws1.cell(row=row, column=4), is_rate=True)

        # 差值公式 = 程序 - BI
        ws1.cell(row=row, column=5).value = f'=IF(C{row}="","",D{row}-C{row})'
        style_data_cell(ws1.cell(row=row, column=5), is_rate=True)

        # 是否一致公式
        ws1.cell(row=row, column=6).value = match_formula("C", "D", row)
        ws1.cell(row=row, column=6).font = Font(name="Microsoft YaHei", size=14)
        ws1.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws1.cell(row=row, column=6).border = THIN_BORDER

        row += 1

    # 空行分隔国家
    row += 1

auto_width(ws1, len(headers1))
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 14
ws1.column_dimensions["E"].width = 12
ws1.column_dimensions["F"].width = 12
ws1.freeze_panes = "A4"

# ========================
#  Sheet 2: 单期/分期
# ========================
ws2 = wb.create_sheet("单期分期")

ws2.merge_cells("A1:H1")
ws2.cell(row=1, column=1, value=f"UAT 验收 — 单期/分期金额首逾率  |  业务日期: {BUSINESS_DATE}")
ws2["A1"].font = TITLE_FONT

headers2 = ["国家", "阶段", "产品类型", "BI-首逾率", "程序-首逾率", "差值", "是否一致", "行数"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=c, value=h)
style_header(ws2, len(headers2))

row = 4
for cc in ["MX", "AR"]:
    country_name = COUNTRIES[cc]["name"]
    rows = filter_country(all_rows, cc)
    for stage in STAGES:
        for label, is_inst in [("单期 (mult_no=1)", False), ("分期 (mult_no>=2)", True)]:
            subset = filter_by_mult_no(rows, is_inst)
            m = calc_stage_metrics(subset, stage)
            ws2.cell(row=row, column=1, value=country_name)
            ws2.cell(row=row, column=2, value=stage)
            ws2.cell(row=row, column=3, value=label)
            ws2.cell(row=row, column=5, value=m["overdue_rate"])
            ws2.cell(row=row, column=8, value=m["row_count"])

            for c in [1, 2, 3, 4, 5, 6, 8]:
                style_data_cell(ws2.cell(row=row, column=c))
            style_data_cell(ws2.cell(row=row, column=5), is_rate=True)

            ws2.cell(row=row, column=6).value = f'=IF(E{row}="","",E{row}-D{row})'
            style_data_cell(ws2.cell(row=row, column=6), is_rate=True)

            ws2.cell(row=row, column=7).value = match_formula("D", "E", row)
            ws2.cell(row=row, column=7).font = Font(name="Microsoft YaHei", size=14)
            ws2.cell(row=row, column=7).alignment = Alignment(horizontal="center")
            ws2.cell(row=row, column=7).border = THIN_BORDER

            row += 1
        row += 1  # 产品类型间空行
    row += 1  # 国家间空行

auto_width(ws2, len(headers2))
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 12
ws2.freeze_panes = "A4"

# ========================
#  Sheet 3: 包体
# ========================
ws3 = wb.create_sheet("包体")

ws3.merge_cells("A1:L1")
ws3.cell(row=1, column=1, value=f"UAT 验收 — 包体明细  |  业务日期: {BUSINESS_DATE}")
ws3["A1"].font = TITLE_FONT

headers3 = ["国家", "阶段", "包体", "产品类型",
            "BI-首逾率", "程序-首逾率", "首逾率一致",
            "BI-到期本金", "程序-到期本金", "到期本金一致",
            "BI-到期笔数", "程序-到期笔数", "到期笔数一致"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=c, value=h)
style_header(ws3, len(headers3))

ORDER_TYPES = ["非分期", "借款分期", "展期分期", "展期N期"]

row = 4
for cc in ["MX", "AR"]:
    country_name = COUNTRIES[cc]["name"]
    rows = filter_country(all_rows, cc)
    for stage in STAGES:
        for ot in ORDER_TYPES:
            subset = filter_by_order_type(rows, ot)
            m = calc_stage_metrics(subset, stage)
            if m["row_count"] == 0:
                continue

            prod = ORDER_TYPE_MAP.get(ot, "?")
            ws3.cell(row=row, column=1, value=country_name)
            ws3.cell(row=row, column=2, value=stage)
            ws3.cell(row=row, column=3, value=ot)
            ws3.cell(row=row, column=4, value=prod)
            # BI-首逾率留空
            ws3.cell(row=row, column=6, value=m["overdue_rate"])
            # BI-到期本金留空
            ws3.cell(row=row, column=9, value=m["due_amt"])
            # BI-到期笔数留空
            ws3.cell(row=row, column=12, value=m["case_count"])

            for c in [1, 2, 3, 4, 5, 6, 8, 9, 11, 12]:
                style_data_cell(ws3.cell(row=row, column=c))
            style_data_cell(ws3.cell(row=row, column=6), is_rate=True)
            style_data_cell(ws3.cell(row=row, column=9), is_rate=False)
            ws3.cell(row=row, column=9).number_format = '#,##0'
            ws3.cell(row=row, column=12).number_format = '#,##0'

            # 首逾率一致
            ws3.cell(row=row, column=7).value = match_formula("E", "F", row)
            ws3.cell(row=row, column=7).font = Font(name="Microsoft YaHei", size=14)
            ws3.cell(row=row, column=7).alignment = Alignment(horizontal="center")
            ws3.cell(row=row, column=7).border = THIN_BORDER

            # 到期本金一致
            ws3.cell(row=row, column=10).value = f'=IF(H{row}="","",IF(ABS(I{row}-H{row})<0.01,"✅","❌"))'
            ws3.cell(row=row, column=10).font = Font(name="Microsoft YaHei", size=14)
            ws3.cell(row=row, column=10).alignment = Alignment(horizontal="center")
            ws3.cell(row=row, column=10).border = THIN_BORDER

            # 到期笔数一致
            ws3.cell(row=row, column=13).value = f'=IF(K{row}="","",IF(K{row}=L{row},"✅","❌"))'
            ws3.cell(row=row, column=13).font = Font(name="Microsoft YaHei", size=14)
            ws3.cell(row=row, column=13).alignment = Alignment(horizontal="center")
            ws3.cell(row=row, column=13).border = THIN_BORDER

            row += 1
        row += 1  # stage 间空行
    row += 1  # 国家间空行

auto_width(ws3, len(headers3))
for col_letter in ["E", "F", "H", "I", "K", "L"]:
    ws3.column_dimensions[col_letter].width = 16
ws3.freeze_panes = "A4"

# ========================
#  Sheet 4: 订单风控等级
# ========================
ws4 = wb.create_sheet("订单风控等级")

ws4.merge_cells("A1:H1")
ws4.cell(row=1, column=1, value=f"UAT 验收 — 订单风控等级 (order_grade) 金额首逾率  |  业务日期: {BUSINESS_DATE}")
ws4["A1"].font = TITLE_FONT

headers4 = ["国家", "阶段", "Grade", "BI-首逾率", "程序-首逾率", "差值", "是否一致", "行数"]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=c, value=h)
style_header(ws4, len(headers4))

GRADES = ["A", "B", "C", "D", "E", "F"]

row = 4
for cc in ["MX", "AR"]:
    country_name = COUNTRIES[cc]["name"]
    rows = filter_country(all_rows, cc)
    for stage in STAGES:
        for grade in GRADES:
            subset = filter_by_order_grade(rows, grade)
            m = calc_stage_metrics(subset, stage)
            if m["row_count"] == 0:
                continue

            ws4.cell(row=row, column=1, value=country_name)
            ws4.cell(row=row, column=2, value=stage)
            ws4.cell(row=row, column=3, value=grade)
            ws4.cell(row=row, column=5, value=m["overdue_rate"])
            ws4.cell(row=row, column=8, value=m["row_count"])

            for c in [1, 2, 3, 4, 5, 6, 8]:
                style_data_cell(ws4.cell(row=row, column=c))
            style_data_cell(ws4.cell(row=row, column=5), is_rate=True)

            ws4.cell(row=row, column=6).value = f'=IF(E{row}="","",E{row}-D{row})'
            style_data_cell(ws4.cell(row=row, column=6), is_rate=True)

            ws4.cell(row=row, column=7).value = match_formula("D", "E", row)
            ws4.cell(row=row, column=7).font = Font(name="Microsoft YaHei", size=14)
            ws4.cell(row=row, column=7).alignment = Alignment(horizontal="center")
            ws4.cell(row=row, column=7).border = THIN_BORDER

            row += 1
        row += 1
    row += 1

auto_width(ws4, len(headers4))
ws4.column_dimensions["D"].width = 14
ws4.column_dimensions["E"].width = 14
ws4.column_dimensions["F"].width = 12
ws4.column_dimensions["G"].width = 12
ws4.freeze_panes = "A4"

# ========================
#  Sheet 5: 字段更新 & 上线标准
# ========================
ws5 = wb.create_sheet("字段更新与上线标准")

ws5.merge_cells("A1:C1")
ws5.cell(row=1, column=1, value=f"数据字典更新 & 上线检查清单  |  v{VERSION}  |  {GEN_TIME}")
ws5["A1"].font = TITLE_FONT

# 字段更新
ws5.cell(row=3, column=1, value="字段更新").font = HEADER_FONT
ws5.merge_cells("A3:C3")

field_headers = ["字段", "状态", "说明"]
for c, h in enumerate(field_headers, 1):
    ws5.cell(row=4, column=c, value=h)
style_header(ws5, 3)

field_data = [
    ["due_case", "✅ 正式启用", "到期笔数，100%非空，在回收率API中直接提供"],
    ["mult_no", "✅ 正式启用", "分期期数。1=单期，≥2=分期。替代 ORDER_TYPE_MAP"],
    ["order_grade", "✅ 正式启用", "订单风控等级 A~F（6级），替代 cust_type 作为默认风控维度"],
    ["cust_type", "⚠️ 保留兼容", "仅保留向后兼容，不再作为主要分析维度"],
    ["due_amt", "❌ Deprecated", "= D_2_due_amt + D_3_pay_amt，不是到期本金。禁止用于权重/贡献/影响金额"],
    ["D_3_pay_amt", "📝 已记录", "D-3回款金额，与 D_2_due_amt 相加 = due_amt"],
    ["影响金额", "📝 规范", "统一使用 {stage}_due_amt: D0→D0_due_amt, D1→D1_due_amt..."],
]

for i, (fld, status, desc) in enumerate(field_data):
    r = 5 + i
    ws5.cell(row=r, column=1, value=fld)
    ws5.cell(row=r, column=2, value=status)
    ws5.cell(row=r, column=3, value=desc)
    for c in [1, 2, 3]:
        style_data_cell(ws5.cell(row=r, column=c))
    ws5.cell(row=r, column=1).font = Font(name="Consolas", size=10)
    ws5.cell(row=r, column=3).alignment = Alignment(wrap_text=True)

# 上线标准
check_start = 5 + len(field_data) + 2
ws5.cell(row=check_start, column=1, value="上线标准").font = HEADER_FONT
ws5.merge_cells(f"A{check_start}:C{check_start}")

check_headers = ["检查项", "状态", "备注"]
for c, h in enumerate(check_headers, 1):
    ws5.cell(row=check_start + 1, column=c, value=h)
style_header(ws5, 3)

check_data = [
    ["整体首逾率与 BI 一致", "☐", "待 BI 比对"],
    ["单期/分期首逾率与 BI 一致", "☐", "待 BI 比对"],
    ["各包体首逾率/到期本金/到期笔数与 BI 一致", "☐", "待 BI 比对"],
    ["各风控等级首逾率与 BI 一致", "☐", "待 BI 比对"],
    ["到期本金使用 {stage}_due_amt", "✅", "已验证"],
    ["到期笔数使用 due_case", "✅", "已验证"],
    ["分期判定使用 mult_no", "✅", "已验证"],
    ["风控维度使用 order_grade", "✅", "已验证"],
    ["due_amt 已标记 Deprecated", "✅", "已标记"],
    ["所有项 BI 一致", "☐", ""],
]

for i, (item, status, note) in enumerate(check_data):
    r = check_start + 2 + i
    ws5.cell(row=r, column=1, value=item)
    ws5.cell(row=r, column=2, value=status)
    ws5.cell(row=r, column=3, value=note)
    for c in [1, 2, 3]:
        style_data_cell(ws5.cell(row=r, column=c))

# 最终结论
final_row = check_start + 2 + len(check_data) + 1
ws5.cell(row=final_row, column=1, value="最终结论").font = HEADER_FONT
ws5.cell(row=final_row, column=2, value="☐ 通过  /  ☐ 不通过")
ws5.cell(row=final_row, column=2).font = Font(name="Microsoft YaHei", bold=True, size=12, color="C00000")

auto_width(ws5, 3)
ws5.column_dimensions["C"].width = 55
ws5.freeze_panes = "A5"

# ============================================================
#  保存
# ============================================================
script_dir = os.path.dirname(os.path.abspath(__file__))
docs_dir = os.path.join(script_dir, "docs")
os.makedirs(docs_dir, exist_ok=True)

output_path = os.path.join(docs_dir, "UAT验收.xlsx")
wb.save(output_path)

print(f"\n  Excel 已生成: {output_path}")
print(f"  包含 {len(wb.sheetnames)} 个工作表: {', '.join(wb.sheetnames)}")
print(f"\n  使用方法:")
print(f"  1. 打开 {output_path}")
print(f"  2. 在 BI值 列填入 BI 中对应的数值")
print(f"  3. 「是否一致」列自动显示 ✅ 或 ❌")
print(f"  4. 完成验收后勾选「上线标准」工作表")
