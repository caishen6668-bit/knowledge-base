"""
V2 阈值优化器 — 自动搜索最佳告警阈值

目标: 使 30 天告警率落在 25%~40% 之间。

搜索空间:
  - ALERT_TARGET_PP:  1.0 ~ 3.0pp（步长 0.5）
  - ALERT_MIN_IMPACT:  100k ~ 500k（步长 50k）
  - P1 变化阈值:       1.0 ~ 2.5pp（步长 0.5）
  - 连续恶化天数:      3 ~ 5 天

方法:
  1. 运行一次 30 天模拟，收集上游原始结果（TrendReport + RootCauseResult）
  2. 对每种阈值组合，仅回放 Action + Alert Decision（纯规则引擎，无 API 调用）
  3. 筛选 25%~40% 告警率组合
  4. 排名输出 Top 10 + 推荐组合

输出:
  - Threshold_Optimization_Report.xlsx

不修改任何业务逻辑。不修改正式配置。
"""

import io
import os
import pickle
import sys
import traceback
from collections import Counter, defaultdict
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Any, Tuple
from dataclasses import dataclass, field
from itertools import product

from .. import config as v1_config
from . import config as v2_config
from . import action_engine as _action_engine
from .trend_engine import compute_trends, reset_cache, _get_business_date
from .root_cause import analyze_root_cause
from .action_engine import analyze_actions
from .alert_decision import decide_alert
from .models import TrendReport, RootCauseResult, ActionResult, AlertDecision


# ============================================================
#  配置
# ============================================================

TARGET_ALERT_RATE_MIN = 0.25      # 目标告警率下限
TARGET_ALERT_RATE_MAX = 0.40      # 目标告警率上限

REPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "docs")
OPTIMIZATION_REPORT = "Threshold_Optimization_Report.xlsx"
CACHE_FILE = os.path.join(REPORT_DIR, ".simulation_cache.pkl")

# 搜索空间
SEARCH_TARGET_PP = [0.010, 0.015, 0.020, 0.025, 0.030]        # 1.0~3.0pp
SEARCH_MIN_IMPACT = list(range(100_000, 550_000, 50_000))       # 100k~500k
SEARCH_P1_THRESHOLD = [0.010, 0.015, 0.020, 0.025]              # 1.0~2.5pp
SEARCH_CONSECUTIVE_DAYS = [3, 4, 5]                             # 3~5天

DEFAULT_COUNTRIES = ["MX", "AR"]
DEFAULT_STAGES = ["D0"]
SIMULATION_DAYS = 30


# ============================================================
#  缓存数据结构
# ============================================================

@dataclass
class CachedDay:
    """单天上游结果的轻量缓存"""
    country_code: str
    country_name: str
    business_date: str
    stage: str
    has_data: bool = False
    # 核心指标（无需完整对象即可回放）
    overall_current_rate: float = 0.0
    overall_change_abs: float = 0.0
    overall_due_amount: float = 0.0
    overall_alert_level: str = "GREEN"
    # 根因路径
    worsening_paths: list = field(default_factory=list)
    improving_paths: list = field(default_factory=list)
    # 原始对象引用（仅内存使用，不序列化）
    trend_report: Any = field(default=None, repr=False)
    root_cause: Any = field(default=None, repr=False)


# ============================================================
#  Phase 1: 数据收集
# ============================================================

class SuppressOutput:
    """临时抑制 stdout"""
    def __init__(self, silent: bool = True):
        self.silent = silent
        self._old_stdout = None

    def __enter__(self):
        if self.silent:
            self._old_stdout = sys.stdout
            sys.stdout = io.StringIO()
        return self

    def __exit__(self, *args):
        if self.silent and self._old_stdout:
            sys.stdout = self._old_stdout


def collect_upstream_data(
    run_date: str = None,
    countries: List[str] = None,
    stages: List[str] = None,
    days: int = SIMULATION_DAYS,
    silent: bool = True,
) -> List[CachedDay]:
    """
    收集 30 天上游数据（Trend + Root Cause）。

    始终重新收集（API 缓存层保证重复请求不产生额外 HTTP 调用）。
    收集完成后保存轻量缓存到 pickle（仅元数据，不含原始对象）。
    """
    if run_date is None:
        run_date = date.today().strftime("%Y-%m-%d")
    if countries is None:
        countries = DEFAULT_COUNTRIES
    if stages is None:
        stages = DEFAULT_STAGES

    # 计算业务日期
    run_dt = datetime.strptime(run_date, "%Y-%m-%d")
    business_dates = []
    for i in range(days):
        bd = (run_dt - timedelta(days=i + 1)).strftime("%Y-%m-%d")
        business_dates.append(bd)
    business_dates.reverse()

    print(f"\n  📡 Phase 1: 收集上游数据 ({len(business_dates)}天 × {len(countries)}国 × {len(stages)}阶段)")
    print(f"     日期: {business_dates[0]} → {business_dates[-1]}")

    all_cached: List[CachedDay] = []
    total = len(business_dates) * len(countries) * len(stages)
    completed = 0

    for bd in business_dates:
        for cc in countries:
            for stage in stages:
                completed += 1
                cached = _collect_one_day(cc, bd, stage, silent)
                all_cached.append(cached)

                if completed % 10 == 0 or completed == total:
                    has = sum(1 for c in all_cached if c.has_data)
                    print(f"     [{completed}/{total}] {bd} {cc}/{stage}  "
                          f"data={cached.has_data}  |  累计: {has}天有数据")

    # 保存缓存
    os.makedirs(os.path.dirname(CACHE_FILE), exist_ok=True)
    # 序列化时排除 trend_report 和 root_cause（它们包含不可序列化的对象）
    to_save = []
    for c in all_cached:
        to_save.append(CachedDay(
            country_code=c.country_code,
            country_name=c.country_name,
            business_date=c.business_date,
            stage=c.stage,
            has_data=c.has_data,
            overall_current_rate=c.overall_current_rate,
            overall_change_abs=c.overall_change_abs,
            overall_due_amount=c.overall_due_amount,
            overall_alert_level=c.overall_alert_level,
            worsening_paths=c.worsening_paths,
            improving_paths=c.improving_paths,
        ))

    with open(CACHE_FILE, "wb") as f:
        pickle.dump(to_save, f)
    print(f"     💾 缓存已保存: {CACHE_FILE} ({len(to_save)} 条)")

    return all_cached


def _collect_one_day(
    country_code: str,
    business_date: str,
    stage: str,
    silent: bool = True,
) -> CachedDay:
    """收集单天的 Trend + RootCause 数据"""
    cached = CachedDay(
        country_code=country_code,
        country_name=v1_config.COUNTRIES[country_code]["name"],
        business_date=business_date,
        stage=stage,
    )

    run_date = (datetime.strptime(business_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    with SuppressOutput(silent):
        try:
            # Phase 1: Trend
            report = compute_trends(
                country_code=country_code,
                business_date=business_date,
                stage=stage,
                run_date=run_date,
            )
            if not report or not report.results:
                return cached

            cached.has_data = True
            cached.trend_report = report
            cached.overall_alert_level = report.worst_overall

            overall = report.get("overall", "total")
            if overall:
                cached.overall_current_rate = overall.current_value
                for c in overall.comparisons:
                    if c.method == "dod":
                        cached.overall_change_abs = c.change_abs
                        break

            # Phase 2: Root Cause (if triggered)
            anomalies = report.get_anomalies()
            triggered = [a for a in anomalies
                         if a.overall_judgment in ("ORANGE", "RED")]

            if triggered:
                try:
                    root_cause = analyze_root_cause(
                        trend_report=report,
                        country_code=country_code,
                        business_date=business_date,
                        stage=stage,
                    )
                    cached.root_cause = root_cause

                    if root_cause and root_cause.has_root_cause:
                        cached.overall_due_amount = root_cause.overall_due_amount
                        for p in root_cause.worsening:
                            cached.worsening_paths.append({
                                "label": p.path_label,
                                "change_pp": p.change_pp,
                                "impact_amount": p.impact_amount,
                            })
                        for p in root_cause.improving:
                            cached.improving_paths.append({
                                "label": p.path_label,
                                "change_pp": p.change_pp,
                                "impact_amount": p.impact_amount,
                            })
                except Exception:
                    pass

        except Exception:
            pass

    return cached


# ============================================================
#  Phase 2: 阈值回放
# ============================================================

# 需要重建成完整对象的路径
def _rebuild_root_cause_path(path_dict: dict, models_module):
    """从缓存字典重建 RootCausePath（最小化版本，仅供 Action 使用）"""
    RootCausePath = models_module.RootCausePath
    return RootCausePath(
        path_label=path_dict["label"],
        change_pp=path_dict["change_pp"],
        contribution_pp=path_dict["change_pp"],
        contribution_pct=0.0,
        impact_amount=path_dict["impact_amount"],
    )


def _rebuild_root_cause_result(cached: CachedDay, models_module) -> Optional[RootCauseResult]:
    """从缓存重建 RootCauseResult"""
    if not cached.root_cause:
        return None
    return cached.root_cause


def replay_with_thresholds(
    cached_days: List[CachedDay],
    p1_threshold: float,
    alert_target_pp: float,
    alert_min_impact: float,
    alert_consecutive_days: int,
    silent: bool = True,
) -> Dict[str, Any]:
    """
    对一组缓存的上游数据，用指定阈值回放 Action + Alert Decision。

    仅修改运行时模块常量，不修改源文件。
    运行结束后恢复原始值。

    Args:
        cached_days: collect_upstream_data() 的返回
        p1_threshold: P1_CHANGE_THRESHOLD（小数，如 0.020 = 2.0pp）
        alert_target_pp: ALERT_TARGET_PP（小数）
        alert_min_impact: ALERT_MIN_IMPACT（整数）
        alert_consecutive_days: ALERT_CONSECUTIVE_DAYS（整数）

    Returns:
        dict with alert_rate, p1_count, p2_count, p3_count, send_count, silent_count
    """
    # ---- 保存原始值 ----
    orig_p1_change = _action_engine.P1_CHANGE_THRESHOLD
    orig_p1_red = _action_engine.P1_RED_CHANGE_THRESHOLD
    orig_target_pp = v2_config.ALERT_TARGET_PP
    orig_min_impact = v2_config.ALERT_MIN_IMPACT
    orig_consec_days = v2_config.ALERT_CONSECUTIVE_DAYS

    # ---- 覆盖阈值 ----
    _action_engine.P1_CHANGE_THRESHOLD = p1_threshold
    _action_engine.P1_RED_CHANGE_THRESHOLD = p1_threshold / 2  # RED 阈值 = P1 的一半
    v2_config.ALERT_TARGET_PP = alert_target_pp
    v2_config.ALERT_MIN_IMPACT = int(alert_min_impact)
    v2_config.ALERT_CONSECUTIVE_DAYS = int(alert_consecutive_days)

    try:
        stats = _replay_all(cached_days, silent)
    finally:
        # ---- 恢复原始值 ----
        _action_engine.P1_CHANGE_THRESHOLD = orig_p1_change
        _action_engine.P1_RED_CHANGE_THRESHOLD = orig_p1_red
        v2_config.ALERT_TARGET_PP = orig_target_pp
        v2_config.ALERT_MIN_IMPACT = orig_min_impact
        v2_config.ALERT_CONSECUTIVE_DAYS = orig_consec_days

    return stats


def _replay_all(cached_days: List[CachedDay], silent: bool = True) -> Dict[str, Any]:
    """对缓存数据回放 Action + Alert Decision"""
    valid = [c for c in cached_days if c.has_data]
    p1_total = p2_total = p3_total = send_count = 0

    with SuppressOutput(silent):
        for cached in cached_days:
            if not cached.has_data or not cached.root_cause:
                continue

            # 重建 RootCauseResult
            root_cause = cached.root_cause

            # Phase 3: Action (使用覆盖后的阈值)
            try:
                action_result = analyze_actions(
                    trend_report=cached.trend_report,
                    root_cause=root_cause,
                )
            except Exception:
                continue

            if not action_result:
                continue

            p1_total += len(action_result.p1_actions)
            p2_total += len(action_result.p2_actions)
            p3_total += len(action_result.p3_actions)

            # Phase 3.5: Alert Decision (使用覆盖后的阈值)
            try:
                decision = decide_alert(
                    trend_report=cached.trend_report,
                    root_cause=root_cause,
                    action_result=action_result,
                    country_code=cached.country_code,
                )
                if decision.should_alert:
                    send_count += 1
            except Exception:
                pass

    total_valid = len(valid)
    alert_rate = send_count / total_valid if total_valid > 0 else 0.0

    return {
        "total_valid": total_valid,
        "send_count": send_count,
        "silent_count": total_valid - send_count,
        "alert_rate": alert_rate,
        "p1_total": p1_total,
        "p2_total": p2_total,
        "p3_total": p3_total,
    }


# ============================================================
#  Phase 3: 网格搜索
# ============================================================

@dataclass
class OptimizationResult:
    """单组参数的优化结果"""
    p1_threshold: float
    target_pp: float
    min_impact: float
    consec_days: int
    alert_rate: float
    send_count: int
    silent_count: int
    p1_total: int
    p2_total: int
    p3_total: int
    in_target: bool = False     # 是否在 25%-40% 目标区间

    @property
    def p1_display(self) -> str:
        return f"{self.p1_threshold * 100:.1f}pp"

    @property
    def target_display(self) -> str:
        return f"{self.target_pp * 100:.1f}pp"

    @property
    def impact_display(self) -> str:
        return f"{self.min_impact:,.0f}"


def run_optimization(
    run_date: str = None,
    countries: List[str] = None,
    stages: List[str] = None,
    days: int = SIMULATION_DAYS,
    silent: bool = True,
) -> List[OptimizationResult]:
    """
    运行完整的阈值优化流程。

    1. 收集/加载上游缓存
    2. 网格搜索所有参数组合
    3. 筛选、排序、输出
    """
    print(f"\n{'='*60}")
    print(f"  V2 阈值优化器 (Threshold Optimizer)")
    print(f"  目标告警率: {TARGET_ALERT_RATE_MIN:.0%} ~ {TARGET_ALERT_RATE_MAX:.0%}")
    print(f"{'='*60}")

    # ---- Step 1: 收集数据 ----
    cached = collect_upstream_data(run_date, countries, stages, days, silent)

    valid_days = [c for c in cached if c.has_data]
    print(f"\n  📡 Phase 2: 网格搜索")
    print(f"     有效数据: {len(valid_days)} 天")
    print(f"     搜索空间: "
          f"target_pp={len(SEARCH_TARGET_PP)} × "
          f"impact={len(SEARCH_MIN_IMPACT)} × "
          f"p1={len(SEARCH_P1_THRESHOLD)} × "
          f"consec={len(SEARCH_CONSECUTIVE_DAYS)} = "
          f"{len(SEARCH_TARGET_PP) * len(SEARCH_MIN_IMPACT) * len(SEARCH_P1_THRESHOLD) * len(SEARCH_CONSECUTIVE_DAYS)} 组合")

    # ---- Step 2: 网格搜索 ----
    all_combos = list(product(
        SEARCH_P1_THRESHOLD,
        SEARCH_TARGET_PP,
        SEARCH_MIN_IMPACT,
        SEARCH_CONSECUTIVE_DAYS,
    ))
    total_combos = len(all_combos)

    results: List[OptimizationResult] = []
    in_target_count = 0

    for idx, (p1_t, tgt_pp, min_imp, consec) in enumerate(all_combos):
        stats = replay_with_thresholds(
            cached, p1_t, tgt_pp, min_imp, consec, silent=True
        )

        in_target = TARGET_ALERT_RATE_MIN <= stats["alert_rate"] <= TARGET_ALERT_RATE_MAX
        if in_target:
            in_target_count += 1

        results.append(OptimizationResult(
            p1_threshold=p1_t,
            target_pp=tgt_pp,
            min_impact=min_imp,
            consec_days=consec,
            alert_rate=stats["alert_rate"],
            send_count=stats["send_count"],
            silent_count=stats["silent_count"],
            p1_total=stats["p1_total"],
            p2_total=stats["p2_total"],
            p3_total=stats["p3_total"],
            in_target=in_target,
        ))

        # 进度
        if (idx + 1) % 50 == 0 or (idx + 1) == total_combos:
            print(f"     [{idx + 1}/{total_combos}] "
                  f"已测 {idx + 1} 组合, "
                  f"命中目标区间: {in_target_count} 组合")

    # ---- Step 3: 排序 & 筛选 ----
    # 优先: 在目标区间内，按告警率排序（越接近 33% 越好）
    in_target_results = [r for r in results if r.in_target]
    target_center = (TARGET_ALERT_RATE_MIN + TARGET_ALERT_RATE_MAX) / 2

    # 目标区间内: 按告警率接近中心排序
    in_target_results.sort(key=lambda r: abs(r.alert_rate - target_center))

    # 目标区间外: 按告警率排序
    out_target_results = [r for r in results if not r.in_target]
    out_target_results.sort(key=lambda r: r.alert_rate)

    # Top 10 = 目标区间内前 10（不足则补区间外最近的）
    top10 = in_target_results[:10]
    if len(top10) < 10:
        # 从区间外补
        remaining = 10 - len(top10)
        top10.extend(out_target_results[:remaining])

    print(f"\n  📊 搜索完成")
    print(f"     总组合: {total_combos}")
    print(f"     目标区间内: {in_target_count} ({in_target_count/total_combos:.1%})")
    print(f"     Top 10 告警率范围: {top10[0].alert_rate:.1%} ~ {top10[-1].alert_rate:.1%}")

    return top10, results


# ============================================================
#  推荐
# ============================================================

def recommend_best(top10: List[OptimizationResult]) -> OptimizationResult:
    """
    从 Top 10 中推荐最佳参数。

    准则:
      1. 在目标区间内（25%-40%）
      2. 告警率接近 33%（区间中心）
      3. 倾向于较低的 P1 阈值（不过度过滤）
      4. 倾向于较低的 target_pp（不过度宽松）
    """
    if not top10:
        return None

    in_target = [r for r in top10 if r.in_target]
    candidates = in_target if in_target else top10

    # 评分: 越低越好
    def score(r: OptimizationResult) -> float:
        target_center = (TARGET_ALERT_RATE_MIN + TARGET_ALERT_RATE_MAX) / 2
        rate_score = abs(r.alert_rate - target_center) * 100  # 偏离中心的程度
        p1_penalty = r.p1_threshold * 20  # P1 阈值越高，越不敏感 → 少量惩罚
        return rate_score + p1_penalty

    candidates.sort(key=score)
    return candidates[0]


# ============================================================
#  打印 & Excel
# ============================================================

def print_optimization_summary(top10: List[OptimizationResult], best: OptimizationResult):
    """打印优化总结"""
    print(f"\n{'='*60}")
    print(f"  🏆 Top 10 参数组合")
    print(f"{'='*60}")

    header = (f"  {'排名':<5} {'P1阈值':<8} {'目标偏离':<10} {'影响金额':<12} "
              f"{'连续天数':<8} {'告警率':<8} {'SEND':<6} {'P1/P2/P3':<15} {'区间':<6}")
    print(header)
    print(f"  {'─'*58}")

    for i, r in enumerate(top10, 1):
        tag = "✅" if r.in_target else "⬜"
        print(f"  {i:<5} {r.p1_display:<8} {r.target_display:<10} {r.impact_display:<12} "
              f"{r.consec_days:<8} {r.alert_rate:<8.1%} {r.send_count:<6} "
              f"{r.p1_total}/{r.p2_total}/{r.p3_total:<13} {tag}")

    if best:
        print(f"\n  {'='*60}")
        print(f"  💡 推荐参数 (Recommended)")
        print(f"  {'─'*60}")
        print(f"  ALERT_TARGET_PP       = {best.target_pp}   # {best.target_display}")
        print(f"  ALERT_MIN_IMPACT      = {int(best.min_impact)}   # {best.impact_display}")
        print(f"  P1_CHANGE_THRESHOLD   = {best.p1_threshold}  # {best.p1_display}")
        print(f"  ALERT_CONSECUTIVE_DAYS = {best.consec_days}")
        print(f"  ─────────────────────────────────")
        print(f"  告警率: {best.alert_rate:.1%}  |  SEND: {best.send_count}  |  "
              f"P1={best.p1_total} P2={best.p2_total} P3={best.p3_total}")

    print(f"{'='*60}\n")


def generate_optimization_excel(
    top10: List[OptimizationResult],
    all_results: List[OptimizationResult],
    best: OptimizationResult,
    output_path: str = None,
):
    """生成 Threshold_Optimization_Report.xlsx"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if output_path is None:
        output_path = os.path.join(REPORT_DIR, OPTIMIZATION_REPORT)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()

    # 样式
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    target_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    title_font = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
    subtitle_font = Font(name="Microsoft YaHei", bold=True, size=12, color="2F5496")
    best_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ============================================
    #  Sheet 1: Top 10
    # ============================================
    ws1 = wb.active
    ws1.title = "Top 10"

    ws1.merge_cells("A1:J1")
    ws1.cell(row=1, column=1, value="阈值优化 Top 10 参数组合").font = title_font

    # 推荐参数
    if best:
        ws1.merge_cells("A3:J3")
        ws1.cell(row=3, column=1, value="💡 推荐参数").font = subtitle_font
        rec_data = [
            ("ALERT_TARGET_PP", best.target_display, best.target_pp),
            ("ALERT_MIN_IMPACT", best.impact_display, int(best.min_impact)),
            ("P1_CHANGE_THRESHOLD", best.p1_display, best.p1_threshold),
            ("ALERT_CONSECUTIVE_DAYS", str(best.consec_days), best.consec_days),
        ]
        for i, (name, display, value) in enumerate(rec_data):
            row = 4 + i
            ws1.cell(row=row, column=1, value=name).font = Font(bold=True)
            ws1.cell(row=row, column=1).fill = best_fill
            ws1.cell(row=row, column=2, value=display).fill = best_fill
            ws1.cell(row=row, column=3, value=value).fill = best_fill
        ws1.cell(row=8, column=1, value=f"预期告警率: {best.alert_rate:.1%}").font = Font(bold=True, color="2F5496")
        ws1.cell(row=9, column=1, value=f"SEND: {best.send_count} | P1={best.p1_total} P2={best.p2_total} P3={best.p3_total}")

    # Top 10 表格
    headers1 = ["排名", "P1变化阈值", "目标偏离阈值", "影响金额阈值", "连续恶化天数",
                 "告警率", "SEND次数", "P1次数", "P2次数", "P3次数", "在目标区间"]
    start_row = 11
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=start_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, r in enumerate(top10, 1):
        row = start_row + i
        vals = [i, r.p1_display, r.target_display, r.impact_display,
                r.consec_days, f"{r.alert_rate:.1%}", r.send_count,
                r.p1_total, r.p2_total, r.p3_total, "✅" if r.in_target else "⬜"]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if r.in_target:
                cell.fill = target_fill

    _auto_width(ws1)
    ws1.freeze_panes = f"A{start_row + 1}"

    # ============================================
    #  Sheet 2: Full Grid
    # ============================================
    ws2 = wb.create_sheet("Full Grid")

    ws2.merge_cells("A1:J1")
    ws2.cell(row=1, column=1, value=f"完整搜索网格 ({len(all_results)} 组合)").font = title_font

    headers2 = ["P1阈值", "目标偏离", "影响金额", "连续天数",
                 "告警率", "SEND", "P1", "P2", "P3", "在目标区间"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # 按告警率排序
    sorted_all = sorted(all_results, key=lambda r: r.alert_rate)
    for i, r in enumerate(sorted_all):
        row = 4 + i
        vals = [r.p1_display, r.target_display, r.impact_display,
                r.consec_days, f"{r.alert_rate:.1%}", r.send_count,
                r.p1_total, r.p2_total, r.p3_total, "✅" if r.in_target else ""]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if r.in_target:
                cell.fill = target_fill

    _auto_width(ws2)
    ws2.freeze_panes = "A4"

    # ============================================
    #  Sheet 3: Summary
    # ============================================
    ws3 = wb.create_sheet("Summary")

    ws3.merge_cells("A1:D1")
    ws3.cell(row=1, column=1, value="阈值优化总结").font = title_font

    in_target = [r for r in all_results if r.in_target]
    summary_rows = [
        ("搜索组合总数", str(len(all_results))),
        ("目标告警率区间", f"{TARGET_ALERT_RATE_MIN:.0%} ~ {TARGET_ALERT_RATE_MAX:.0%}"),
        ("命中目标区间", f"{len(in_target)} 组合 ({len(in_target)/len(all_results):.1%})"),
        ("", ""),
        ("搜索维度", ""),
        ("  P1 变化阈值", f"{SEARCH_P1_THRESHOLD[0]*100:.1f} ~ {SEARCH_P1_THRESHOLD[-1]*100:.1f}pp (步长 0.5pp)"),
        ("  目标偏离阈值", f"{SEARCH_TARGET_PP[0]*100:.1f} ~ {SEARCH_TARGET_PP[-1]*100:.1f}pp (步长 0.5pp)"),
        ("  影响金额阈值", f"{SEARCH_MIN_IMPACT[0]:,} ~ {SEARCH_MIN_IMPACT[-1]:,} (步长 50k)"),
        ("  连续恶化天数", f"{SEARCH_CONSECUTIVE_DAYS[0]} ~ {SEARCH_CONSECUTIVE_DAYS[-1]} 天"),
        ("", ""),
        ("推荐参数", ""),
    ]

    row = 3
    for label, value in summary_rows:
        ws3.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=row, column=2, value=value)
        row += 1

    if best:
        row += 1
        ws3.cell(row=row, column=1, value="推荐配置").font = subtitle_font
        row += 1
        rec = [
            ("ALERT_TARGET_PP", str(best.target_pp), best.target_display),
            ("ALERT_MIN_IMPACT", str(int(best.min_impact)), best.impact_display),
            ("P1_CHANGE_THRESHOLD", str(best.p1_threshold), best.p1_display),
            ("P1_RED_CHANGE_THRESHOLD", str(best.p1_threshold / 2), f"{best.p1_threshold / 2 * 100:.1f}pp"),
            ("ALERT_CONSECUTIVE_DAYS", str(best.consec_days), f"{best.consec_days} 天"),
        ]
        for name, val, display in rec:
            ws3.cell(row=row, column=1, value=name).font = Font(bold=True)
            ws3.cell(row=row, column=2, value=val)
            ws3.cell(row=row, column=3, value=f"({display})")
            row += 1

        row += 1
        ws3.cell(row=row, column=1, value="预期效果").font = subtitle_font
        row += 1
        effects = [
            ("告警率", f"{best.alert_rate:.1%}"),
            ("SEND 次数", f"{best.send_count}/{len([c for c in (all_results[0:60] if all_results else [])])}"),
            ("P1 / P2 / P3", f"{best.p1_total} / {best.p2_total} / {best.p3_total}"),
        ]
        for label, value in effects:
            ws3.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws3.cell(row=row, column=2, value=value)
            row += 1

    _auto_width(ws3)

    # 保存
    wb.save(output_path)
    print(f"  📊 Excel 报告已生成: {output_path}")
    return output_path


def _auto_width(ws):
    """自动列宽"""
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                char_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_len = max(max_len, min(char_len + 2, 50))
        ws.column_dimensions[col_letter].width = max(max_len, 8)
