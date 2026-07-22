"""
V2 模拟运行引擎 — 30 天回测验证

读取最近 30 天数据，逐天运行完整 V2 流程：
  Trend → Root Cause → Action → Alert Decision

统计告警频率、根因分布、行动建议分布，
生成 Simulation_Report.xlsx。

支持 --save-cache 将上游结果缓存为 pickle，供 Threshold Optimizer 回放。

不修改任何业务逻辑（Trend / RootCause / Action / AlertDecision）。
"""

import io
import os
import pickle
import sys
import traceback
from collections import Counter, defaultdict
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Any, Tuple

from .. import config as v1_config
from . import config as v2_config
from .trend_engine import compute_trends, reset_cache, _get_business_date
from .root_cause import analyze_root_cause
from .action_engine import analyze_actions
from .alert_decision import decide_alert
from .alert_decision_v2 import decide_alert_v2
from .models import TrendReport, RootCauseResult, ActionResult, AlertDecision


# ============================================================
#  配置
# ============================================================

SIMULATION_DAYS = 30                # 回溯天数
DEFAULT_COUNTRIES = ["MX", "AR"]    # 默认模拟国家
DEFAULT_STAGES = ["D0"]             # 默认模拟阶段

# 输出路径
REPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "docs")
REPORT_FILE = "Simulation_Report.xlsx"


# ============================================================
#  单日结果
# ============================================================

class DayResult:
    """单日单国家单阶段的模拟结果"""
    __slots__ = (
        "run_date", "business_date", "country_code", "country_name", "stage",
        "has_data", "has_anomaly", "has_root_cause",
        "overall_rate", "overall_change_pp", "overall_alert",
        "p1_count", "p2_count", "p3_count",
        "should_alert", "alert_reason", "trigger_source",
        "max_impact", "target_deviation_pp", "consecutive_worsening_days",
        "confidence_score", "confidence_breakdown",
        "due_amount", "case_count", "business_scale_skip",
        "worsening_paths", "improving_paths",
        "top_actions", "top_root_causes",
        "error",
    )

    def __init__(self):
        self.run_date = ""
        self.business_date = ""
        self.country_code = ""
        self.country_name = ""
        self.stage = ""
        self.has_data = False
        self.has_anomaly = False
        self.has_root_cause = False
        self.overall_rate = 0.0
        self.overall_change_pp = 0.0
        self.overall_alert = "GREEN"
        self.p1_count = 0
        self.p2_count = 0
        self.p3_count = 0
        self.should_alert = False
        self.alert_reason = ""
        self.trigger_source = ""
        self.max_impact = 0.0
        self.target_deviation_pp = 0.0
        self.consecutive_worsening_days = 0
        self.confidence_score = 0.0
        self.confidence_breakdown = ""
        self.due_amount = 0.0
        self.case_count = 0
        self.business_scale_skip = False
        self.worsening_paths: List[str] = []
        self.improving_paths: List[str] = []
        self.top_actions: List[str] = []
        self.top_root_causes: List[str] = []
        self.error = ""

    def to_row(self) -> list:
        """转为 Excel 行"""
        return [
            self.business_date,
            self.country_name,
            self.stage,
            "✅" if self.has_data else "❌",
            f"{self.overall_rate:.2%}" if self.has_data else "",
            f"{self.overall_change_pp * 100:+.2f}pp" if self.has_data else "",
            self.overall_alert,
            "Y" if self.has_root_cause else "N",
            self.p1_count, self.p2_count, self.p3_count,
            "📢 SEND" if self.should_alert else "🔇 SILENT",
            self.alert_reason,
            self.trigger_source,
            f"{self.max_impact:,.0f}" if self.max_impact > 0 else "",
            f"{self.target_deviation_pp * 100:+.2f}pp" if self.has_data else "",
            str(self.consecutive_worsening_days),
            f"{self.confidence_score:.0f}%" if self.has_data else "",
            self.confidence_breakdown,
            f"{self.due_amount:,.0f}" if self.due_amount > 0 else "",
            str(self.case_count) if self.case_count > 0 else "",
            "Y" if self.business_scale_skip else "",
            " | ".join(self.worsening_paths[:3]),
            " | ".join(self.improving_paths[:3]),
            " | ".join(self.top_actions[:3]),
            self.error,
        ]


# ============================================================
#  输出抑制
# ============================================================

class SuppressOutput:
    """临时抑制 stdout（不修改业务逻辑）"""
    def __init__(self, silent: bool = True):
        self.silent = silent
        self._buf = None
        self._old_stdout = None

    def __enter__(self):
        if self.silent:
            self._old_stdout = sys.stdout
            sys.stdout = io.StringIO()
        return self

    def __exit__(self, *args):
        if self.silent and self._old_stdout:
            sys.stdout = self._old_stdout


# ============================================================
#  单日运行
# ============================================================

def _run_one_day(
    country_code: str,
    business_date: str,
    stage: str,
    silent: bool = True,
    alert_version: str = "v1",
) -> DayResult:
    """
    运行单个 (country, date, stage) 的完整 V2 流程。

    Args:
        country_code: "MX" | "AR"
        business_date: "2026-07-06"
        stage: "D0" | "D1" | ...
        silent: 是否抑制控制台输出

    Returns:
        DayResult
    """
    result = DayResult()
    result.country_code = country_code
    result.country_name = v1_config.COUNTRIES[country_code]["name"]
    result.business_date = business_date
    result.run_date = (datetime.strptime(business_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    result.stage = stage

    with SuppressOutput(silent):
        try:
            # ---- Phase 1: Trend ----
            report = compute_trends(
                country_code=country_code,
                business_date=business_date,
                stage=stage,
                run_date=result.run_date,
            )

            if not report or not report.results:
                result.error = "无数据"
                return result

            result.has_data = True
            overall = report.get("overall", "total")
            if overall:
                result.overall_rate = overall.current_value
                result.overall_alert = report.worst_overall

                # 计算整体变化 (vs 昨天)
                if overall.comparisons:
                    for c in overall.comparisons:
                        if c.method == "dod":
                            result.overall_change_pp = c.change_abs
                            break

            # ---- Phase 2: Root Cause ----
            anomalies = report.get_anomalies()
            triggered = [a for a in anomalies
                         if a.overall_judgment in ("ORANGE", "RED")]
            result.has_anomaly = len(triggered) > 0

            root_cause = None
            if triggered:
                try:
                    root_cause = analyze_root_cause(
                        trend_report=report,
                        country_code=country_code,
                        business_date=business_date,
                        stage=stage,
                    )
                except Exception:
                    pass  # Root cause failure is non-fatal

            if root_cause and root_cause.has_root_cause:
                result.has_root_cause = True

                # 收集恶化/改善路径
                for p in root_cause.worsening:
                    result.worsening_paths.append(p.path_label)
                for p in root_cause.improving:
                    result.improving_paths.append(p.path_label)

                # 收集根因标签
                for p in root_cause.worsening[:3]:
                    grade = _extract_grade(p.path_label)
                    result.top_root_causes.append(
                        f"{p.path_label} ({p.change_pp * 100:+.2f}pp)"
                    )
                for p in root_cause.improving[:3]:
                    result.top_root_causes.append(
                        f"{p.path_label} ({p.change_pp * 100:+.2f}pp)"
                    )

            # ---- Phase 3: Action ----
            action_result = None
            if root_cause and root_cause.has_root_cause:
                try:
                    action_result = analyze_actions(
                        trend_report=report,
                        root_cause=root_cause,
                    )
                except Exception:
                    pass

            if action_result:
                result.p1_count = len(action_result.p1_actions)
                result.p2_count = len(action_result.p2_actions)
                result.p3_count = len(action_result.p3_actions)

                for item in action_result.all_actions:
                    result.top_actions.append(
                        f"[{item.priority}] {item.action} → {item.target_path}"
                    )

            # ---- Phase 3.5: Alert Decision ----
            if action_result:
                try:
                    if alert_version == "v2":
                        decision = decide_alert_v2(
                            trend_report=report,
                            root_cause=root_cause,
                            action_result=action_result,
                            country_code=country_code,
                        )
                    else:
                        decision = decide_alert(
                            trend_report=report,
                            root_cause=root_cause,
                            action_result=action_result,
                            country_code=country_code,
                        )
                    result.should_alert = decision.should_alert
                    result.alert_reason = decision.alert_reason
                    result.trigger_source = decision.trigger_source
                    result.max_impact = decision.max_impact
                    result.target_deviation_pp = decision.target_deviation_pp
                    result.consecutive_worsening_days = decision.consecutive_worsening_days
                    result.confidence_score = decision.confidence_score
                    result.confidence_breakdown = decision.confidence_breakdown
                    result.due_amount = decision.due_amount
                    result.case_count = decision.case_count
                    result.business_scale_skip = decision.business_scale_skip
                except Exception:
                    pass

        except Exception as e:
            result.error = str(e)[:200]

    return result


def _extract_grade(path_label: str) -> str:
    """从路径标签中提取风控等级，如 '单期 → 非分期 → C' → 'C'"""
    parts = path_label.split("→")
    if parts:
        return parts[-1].strip()
    return ""


# ============================================================
#  30 天模拟
# ============================================================

def run_simulation(
    run_date: str = None,
    countries: List[str] = None,
    stages: List[str] = None,
    days: int = SIMULATION_DAYS,
    silent: bool = True,
    alert_version: str = "v1",
) -> tuple:
    """
    运行 30 天模拟。

    Args:
        run_date: 运行日期（默认今天）
        countries: 国家列表（默认 MX, AR）
        stages: 阶段列表（默认 D0）
        days: 回溯天数（默认 30）
        silent: 是否抑制单日输出
        alert_version: "v1" | "v2" — Alert Decision 版本

    Returns:
        (List[DayResult], dict of statistics)
    """
    if run_date is None:
        run_date = date.today().strftime("%Y-%m-%d")
    if countries is None:
        countries = DEFAULT_COUNTRIES
    if stages is None:
        stages = DEFAULT_STAGES

    # 计算业务日期范围
    run_dt = datetime.strptime(run_date, "%Y-%m-%d")
    business_dates = []
    for i in range(days):
        bd = (run_dt - timedelta(days=i + 1)).strftime("%Y-%m-%d")
        business_dates.append(bd)
    business_dates.reverse()  # 从最早到最新

    print(f"\n{'='*60}")
    print(f"  V2 模拟运行 — 回测 {days} 天")
    print(f"  日期范围: {business_dates[0]} → {business_dates[-1]}")
    print(f"  国家: {', '.join(countries)}  |  阶段: {', '.join(stages)}")
    print(f"  总任务: {len(business_dates) * len(countries) * len(stages)}")
    print(f"{'='*60}")

    all_results: List[DayResult] = []

    total = len(business_dates) * len(countries) * len(stages)
    completed = 0

    for bd in business_dates:
        for cc in countries:
            for stage in stages:
                completed += 1
                day_result = _run_one_day(cc, bd, stage, silent=silent,
                                         alert_version=alert_version)
                all_results.append(day_result)

                # 进度（每 5 个打印一次）
                if completed % 5 == 0 or completed == total:
                    has_data = sum(1 for r in all_results if r.has_data)
                    has_alert = sum(1 for r in all_results if r.should_alert)
                    print(f"  [{completed}/{total}] {bd} {cc} {stage}  "
                          f"data={day_result.has_data} alert={day_result.should_alert}"
                          f"  |  累计: {has_data}天有数据, {has_alert}次告警")

    # ---- 统计 ----
    stats = _compute_statistics(all_results, business_dates, countries, stages)

    return all_results, stats


# ============================================================
#  V1 vs V2 比较
# ============================================================

def run_comparison(
    run_date: str = None,
    countries: List[str] = None,
    stages: List[str] = None,
    days: int = SIMULATION_DAYS,
    silent: bool = True,
) -> dict:
    """
    运行 V1 vs V2 Alert Decision 对比。

    先跑 V1（现有规则），再跑 V2（整体驱动），
    对比告警率、差异天数、每国每阶段详情。
    """
    if run_date is None:
        run_date = date.today().strftime("%Y-%m-%d")
    if countries is None:
        countries = DEFAULT_COUNTRIES
    if stages is None:
        stages = DEFAULT_STAGES

    print(f"\n{'='*60}")
    print(f"  V2 Alert Decision V1 vs V2 对比")
    print(f"  日期范围: {days} 天  |  国家: {', '.join(countries)}")
    print(f"{'='*60}")

    # ---- V1 ----
    print(f"\n  🔵 运行 V1 Alert Decision ...")
    v1_results, v1_stats = run_simulation(
        run_date=run_date, countries=countries, stages=stages,
        days=days, silent=silent,
    )

    # ---- V2 ----
    print(f"\n  🟢 运行 V2 Alert Decision ...")
    # 重新运行但使用 V2 alert decision
    v2_results, v2_stats = run_simulation(
        run_date=run_date, countries=countries, stages=stages,
        days=days, silent=silent, alert_version="v2",
    )

    # ---- 对比 ----
    print(f"\n  📊 对比分析 ...")

    comparison = {
        "v1": v1_stats,
        "v2": v2_stats,
        "v1_results": v1_results,
        "v2_results": v2_results,
    }

    # 逐日逐国对比
    diffs = []  # 差异记录
    v1_alerts = {(r.business_date, r.country_code): r for r in v1_results if r.should_alert}
    v2_alerts = {(r.business_date, r.country_code): r for r in v2_results if r.should_alert}

    v1_sends = set(v1_alerts.keys())
    v2_sends = set(v2_alerts.keys())

    # 新增告警（V1 不发 V2 发）
    new_alerts = v2_sends - v1_sends
    # 消除告警（V1 发 V2 不发）
    removed_alerts = v1_sends - v2_sends
    # 持续告警（都发）
    common_alerts = v1_sends & v2_sends

    for bd, cc in removed_alerts:
        r = v1_alerts[(bd, cc)]
        diffs.append({
            "date": bd, "country": cc,
            "change": "🔇 V1→V2 消除",
            "v1_reason": r.alert_reason,
            "v2_reason": "整体趋势未达标",
        })

    for bd, cc in new_alerts:
        r = v2_alerts[(bd, cc)]
        diffs.append({
            "date": bd, "country": cc,
            "change": "📢 V1→V2 新增",
            "v1_reason": "",
            "v2_reason": r.alert_reason,
        })

    comparison["diffs"] = diffs
    comparison["removed_count"] = len(removed_alerts)
    comparison["new_count"] = len(new_alerts)
    comparison["common_count"] = len(common_alerts)
    comparison["v1_send_count"] = len(v1_sends)
    comparison["v2_send_count"] = len(v2_sends)
    comparison["reduction"] = len(v1_sends) - len(v2_sends)

    # 打印对比摘要
    _print_comparison_summary(comparison)

    return comparison


def _print_comparison_summary(c: dict):
    """打印 V1 vs V2 对比摘要"""
    v1 = c["v1"]
    v2 = c["v2"]

    print(f"\n  {'='*55}")
    print(f"  📊 V1 vs V2 告警决策对比")
    print(f"  {'='*55}")

    print(f"\n  {'指标':<20} {'V1 (旧)':<15} {'V2 (新)':<15} {'变化':<15}")
    print(f"  {'─'*55}")
    print(f"  {'总任务数':<20} {v1['total_tasks']:<15} {v2['total_tasks']:<15} {'':<15}")
    print(f"  {'发送飞书':<20} {v1['total_alerts']:<15} {v2['total_alerts']:<15} "
          f"{'🔽 ' + str(c['reduction']) if c['reduction'] > 0 else ''}")
    print(f"  {'告警率':<20} {v1['alert_rate']:<15.1%} {v2['alert_rate']:<15.1%} "
          f"{'🔽 ' + str(round(v1['alert_rate'] - v2['alert_rate'], 3))}")

    print(f"\n  {'告警变化':<20}")
    print(f"  {'─'*55}")
    print(f"  V1+V2 都发: {c['common_count']}")
    print(f"  V1 发 V2 不发 (消除): {c['removed_count']}")
    print(f"  V1 不发 V2 发 (新增): {c['new_count']}")

    if c["diffs"]:
        print(f"\n  差异明细 (前 20 条):")
        for d in c["diffs"][:20]:
            print(f"    {d['date']} {d['country']}: {d['change']}")
            if d['v1_reason']:
                print(f"      V1: {d['v1_reason'][:80]}")
            if d['v2_reason']:
                print(f"      V2: {d['v2_reason'][:80]}")

    # 评估
    v2_rate = v2['alert_rate']
    if 0.25 <= v2_rate <= 0.50:
        assessment = f"✅ V2 告警率 {v2_rate:.1%} 在 30%-50% 目标区间内"
    elif v2_rate < 0.25:
        assessment = f"⚠️ V2 告警率 {v2_rate:.1%} 偏低 (<25%)，建议放宽阈值"
    else:
        assessment = f"⚠️ V2 告警率 {v2_rate:.1%} 偏高 (>50%)，建议收紧阈值"

    print(f"\n  💡 评估: {assessment}")
    print(f"{'='*55}\n")

    c["assessment"] = assessment


def generate_comparison_excel(comparison: dict, output_path: str = None):
    """生成 V1 vs V2 对比 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if output_path is None:
        output_path = os.path.join(REPORT_DIR, "Alert_V1_vs_V2_Comparison.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
    removed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.merge_cells("A1:E1")
    ws1.cell(row=1, column=1, value="Alert Decision V1 vs V2 对比").font = title_font

    v1, v2 = comparison["v1"], comparison["v2"]
    summary_rows = [
        ("指标", "V1 (旧规则)", "V2 (新规则)", "变化"),
        ("总任务数", v1["total_tasks"], v2["total_tasks"], ""),
        ("发送飞书", v1["total_alerts"], v2["total_alerts"],
         f"🔽 {comparison['reduction']}"),
        ("告警率", f"{v1['alert_rate']:.1%}", f"{v2['alert_rate']:.1%}",
         f"{v1['alert_rate'] - v2['alert_rate']:+.1%}"),
        ("P1 总计", v1["p1_total"], v2["p1_total"], ""),
        ("P2 总计", v1["p2_total"], v2["p2_total"], ""),
        ("P3 总计", v1["p3_total"], v2["p3_total"], ""),
        ("", "", "", ""),
        ("都发送", comparison["common_count"], "", ""),
        ("V1→V2 消除", comparison["removed_count"], "", ""),
        ("V1→V2 新增", comparison["new_count"], "", ""),
        ("", "", "", ""),
        ("评估", comparison.get("assessment", ""), "", ""),
    ]

    for i, (a, b, c, d) in enumerate(summary_rows):
        row = 3 + i
        for col, val in enumerate([a, b, c, d], 1):
            cell = ws1.cell(row=row, column=col, value=val)
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border

    _auto_column_width(ws1)

    # Sheet 2: Daily Comparison
    ws2 = wb.create_sheet("Daily Comparison")
    ws2.merge_cells("A1:G1")
    ws2.cell(row=1, column=1, value="逐日 V1 vs V2 对比").font = title_font

    headers2 = ["日期", "国家", "整体变化(pp)", "告警等级",
                 "V1决策", "V2决策", "变化", "V2置信度", "到期本金", "到期笔数"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 4
    for v1r, v2r in zip(comparison["v1_results"], comparison["v2_results"]):
        if not v1r.has_data:
            continue
        v1_decision = "📢" if v1r.should_alert else "🔇"
        v2_decision = "📢" if v2r.should_alert else "🔇"
        if v1r.should_alert and not v2r.should_alert:
            change = "🔇 消除"
        elif not v1r.should_alert and v2r.should_alert:
            change = "📢 新增"
        elif v1r.should_alert and v2r.should_alert:
            change = "—"
        else:
            change = "—"

        vals = [v1r.business_date, v1r.country_name,
                f"{v1r.overall_change_pp * 100:+.2f}pp", v1r.overall_alert,
                v1_decision, v2_decision, change,
                f"{v2r.confidence_score:.0f}%",
                f"{v2r.due_amount:,.0f}" if v2r.due_amount > 0 else "",
                str(v2r.case_count) if v2r.case_count > 0 else ""]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if "消除" in str(change):
                cell.fill = removed_fill
        row += 1

    _auto_column_width(ws2)
    ws2.freeze_panes = "A4"

    # Sheet 3: Per-Country Stats
    ws3 = wb.create_sheet("Per Country")
    ws3.merge_cells("A1:F1")
    ws3.cell(row=1, column=1, value="分国家统计").font = title_font

    headers3 = ["国家", "V1发送", "V2发送", "消除", "V1告警率", "V2告警率"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for cc in ["MX", "AR"]:
        v1_cc = [r for r in comparison["v1_results"] if r.country_code == cc and r.has_data]
        v2_cc = [r for r in comparison["v2_results"] if r.country_code == cc and r.has_data]
        v1_sends = sum(1 for r in v1_cc if r.should_alert)
        v2_sends = sum(1 for r in v2_cc if r.should_alert)
        total = len(v1_cc)

        vals = [v1_config.COUNTRIES[cc]["name"], v1_sends, v2_sends,
                v1_sends - v2_sends,
                f"{v1_sends/total:.1%}" if total else "",
                f"{v2_sends/total:.1%}" if total else ""]
        row = 4 + (["MX", "AR"].index(cc))
        for col, val in enumerate(vals, 1):
            ws3.cell(row=row, column=col, value=val).border = thin_border

    _auto_column_width(ws3)

    wb.save(output_path)
    print(f"  📊 对比 Excel 已生成: {output_path}")
    return output_path

def _compute_statistics(
    results: List[DayResult],
    business_dates: List[str],
    countries: List[str],
    stages: List[str],
) -> dict:
    """计算所有统计指标"""

    valid = [r for r in results if r.has_data]
    alerts = [r for r in valid if r.should_alert]
    silent_days = [r for r in valid if not r.should_alert]

    # 1. 30天共发送多少次飞书
    total_alerts = len(alerts)

    # 2. 每天哪些国家报警
    alerts_by_date = defaultdict(list)
    for r in alerts:
        alerts_by_date[r.business_date].append(r.country_code)

    # 3. 每天哪些阶段报警
    alerts_by_date_stage = defaultdict(lambda: defaultdict(list))
    for r in alerts:
        alerts_by_date_stage[r.business_date][r.stage].append(r.country_code)

    # 4. P1/P2/P3 各出现多少次
    p1_total = sum(r.p1_count for r in valid)
    p2_total = sum(r.p2_count for r in valid)
    p3_total = sum(r.p3_count for r in valid)

    # 5. Top 10 最常见 Root Cause（提取 grade 级别）
    grade_counter = Counter()
    path_counter = Counter()
    for r in alerts:
        for p in r.worsening_paths:
            grade = _extract_grade(p)
            if grade:
                grade_counter[f"📈 {grade}"] += 1
            path_counter[f"📈 {p}"] += 1
        for p in r.improving_paths:
            grade = _extract_grade(p)
            if grade:
                grade_counter[f"📉 {grade}"] += 1
            path_counter[f"📉 {p}"] += 1

    # 6. Top 10 最常见 Action
    action_counter = Counter()
    for r in alerts:
        for a in r.top_actions:
            # 提取纯动作名（去掉 P 标记和路径）
            action_counter[a] += 1

    # 7. 哪些天没有报警
    days_no_alert = []
    for bd in business_dates:
        day_results = [r for r in valid if r.business_date == bd]
        if day_results and not any(r.should_alert for r in day_results):
            days_no_alert.append(bd)

    # 8. 哪些天报警最多
    alert_count_by_date = Counter()
    for r in alerts:
        alert_count_by_date[r.business_date] += 1
    top_alert_days = alert_count_by_date.most_common(10)

    # 无数据的天
    days_no_data = []
    for bd in business_dates:
        day_results = [r for r in results if r.business_date == bd]
        if day_results and not any(r.has_data for r in day_results):
            days_no_data.append(bd)

    # 告警率
    alert_rate = len(alerts) / len(valid) if valid else 0

    return {
        "total_tasks": len(results),
        "total_days_with_data": len(valid),
        "total_alerts": total_alerts,
        "alert_rate": alert_rate,
        "total_silent": len(silent_days),
        "alerts_by_date": dict(alerts_by_date),
        "alerts_by_date_stage": {k: dict(v) for k, v in alerts_by_date_stage.items()},
        "p1_total": p1_total,
        "p2_total": p2_total,
        "p3_total": p3_total,
        "top_root_causes": path_counter.most_common(10),
        "top_grade_directions": grade_counter.most_common(10),
        "top_actions": action_counter.most_common(10),
        "days_no_alert": days_no_alert,
        "days_no_data": days_no_data,
        "top_alert_days": top_alert_days,
        "threshold_assessment": _assess_thresholds(alert_rate, total_alerts, len(valid)),
    }


def _assess_thresholds(alert_rate: float, total_alerts: int, total_days: int) -> str:
    """评估阈值是否合理"""
    if total_days == 0:
        return "无有效数据，无法评估"

    if alert_rate > 0.80:
        return (
            f"⚠️ 告警率 {alert_rate:.0%}（{total_alerts}/{total_days}），过高。"
            f"建议提高阈值（ALERT_MIN_IMPACT ↑ / ALERT_TARGET_PP ↑ / ALERT_CONSECUTIVE_DAYS ↑）"
        )
    elif alert_rate > 0.50:
        return (
            f"⚡ 告警率 {alert_rate:.0%}（{total_alerts}/{total_days}），偏高。"
            f"可考虑适当收紧阈值。"
        )
    elif alert_rate < 0.10:
        return (
            f"⚠️ 告警率 {alert_rate:.0%}（{total_alerts}/{total_days}），过低。"
            f"建议降低阈值（ALERT_MIN_IMPACT ↓ / ALERT_TARGET_PP ↓ / ALERT_CONSECUTIVE_DAYS ↓）"
        )
    elif alert_rate < 0.20:
        return (
            f"⚡ 告警率 {alert_rate:.0%}（{total_alerts}/{total_days}），偏低。"
            f"可考虑适当放宽阈值。"
        )
    else:
        return (
            f"✅ 告警率 {alert_rate:.0%}（{total_alerts}/{total_days}），"
            f"在合理范围内（20%~50%）。当前阈值适中。"
        )


# ============================================================
#  Excel 报告生成
# ============================================================

def generate_excel(results: List[DayResult], stats: dict, output_path: str = None):
    """
    生成 Simulation_Report.xlsx。

    Sheets:
      1. Daily Log — 每日详情
      2. Summary — 统计摘要
      3. Top Root Causes — 最常见根因
      4. Top Actions — 最常见行动建议
      5. P1/P2/P3 — 优先级分布
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    if output_path is None:
        output_path = os.path.join(REPORT_DIR, REPORT_FILE)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()

    # ---- 样式 ----
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    alert_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色告警
    silent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色静默
    nodata_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 灰色无数据
    title_font = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
    subtitle_font = Font(name="Microsoft YaHei", bold=True, size=12, color="2F5496")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="center")

    # ============================================
    #  Sheet 1: Daily Log
    # ============================================
    ws1 = wb.active
    ws1.title = "Daily Log"

    headers = [
        "业务日期", "国家", "阶段", "有数据", "整体首逾率",
        "整体变化(pp)", "告警等级", "有根因",
        "P1", "P2", "P3", "飞书决策",
        "告警原因", "触发来源", "最大影响金额",
        "高于目标(pp)", "连续恶化(天)",
        "置信度", "置信度分解",
        "到期本金", "到期笔数", "规模跳过",
        "恶化路径 Top3", "改善路径 Top3",
        "行动建议 Top3", "错误",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        row_data = r.to_row()
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = wrap_align

            # 条件着色
            if col_idx == 12:  # 飞书决策
                if "SEND" in str(val):
                    cell.fill = alert_fill
                elif "SILENT" in str(val):
                    cell.fill = silent_fill
            elif col_idx == 4 and val == "❌":
                cell.fill = nodata_fill

    # 冻结首行 + 自动列宽
    ws1.freeze_panes = "A2"
    _auto_column_width(ws1)

    # ============================================
    #  Sheet 2: Summary
    # ============================================
    ws2 = wb.create_sheet("Summary")

    # 标题
    ws2.merge_cells("A1:F1")
    ws2.cell(row=1, column=1, value="V2 Simulation Summary").font = title_font

    # 基本信息
    info_rows = [
        ("模拟天数", f"{SIMULATION_DAYS} 天"),
        ("任务总数", str(stats["total_tasks"])),
        ("有数据天数", str(stats["total_days_with_data"])),
        ("发送飞书", f"{stats['total_alerts']} 次"),
        ("静默跳过", f"{stats['total_silent']} 次"),
        ("告警率", f"{stats['alert_rate']:.1%}"),
        ("", ""),
        ("P1 总计", str(stats["p1_total"])),
        ("P2 总计", str(stats["p2_total"])),
        ("P3 总计", str(stats["p3_total"])),
        ("", ""),
        ("阈值评估", stats["threshold_assessment"]),
    ]

    row = 3
    for label, value in info_rows:
        ws2.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=value)
        if label == "阈值评估":
            ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # 告警最多的天
    row += 1
    ws2.cell(row=row, column=1, value="告警最多的天 (Top 10)").font = subtitle_font
    row += 1
    for date_str, count in stats["top_alert_days"]:
        ws2.cell(row=row, column=1, value=date_str)
        ws2.cell(row=row, column=2, value=f"{count} 次")
        row += 1

    # 无告警的天
    row += 1
    ws2.cell(row=row, column=1, value="无告警的天").font = subtitle_font
    row += 1
    if stats["days_no_alert"]:
        ws2.cell(row=row, column=1, value=", ".join(stats["days_no_alert"]))
    else:
        ws2.cell(row=row, column=1, value="无（每天都至少有一次告警）")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    # 无数据的天
    row += 1
    ws2.cell(row=row, column=1, value="无数据的天").font = subtitle_font
    row += 1
    if stats["days_no_data"]:
        ws2.cell(row=row, column=1, value=", ".join(stats["days_no_data"]))
    else:
        ws2.cell(row=row, column=1, value="无（每天都有数据）")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    # 每天告警详情
    row += 2
    ws2.cell(row=row, column=1, value="每天告警详情").font = subtitle_font
    row += 1
    ws2.cell(row=row, column=1, value="日期").font = header_font
    ws2.cell(row=row, column=1).fill = header_fill
    ws2.cell(row=row, column=2, value="告警国家").font = header_font
    ws2.cell(row=row, column=2).fill = header_fill
    ws2.cell(row=row, column=3, value="告警数").font = header_font
    ws2.cell(row=row, column=3).fill = header_fill
    row += 1
    for date_str, country_list in sorted(stats["alerts_by_date"].items()):
        ws2.cell(row=row, column=1, value=date_str)
        ws2.cell(row=row, column=2, value=", ".join(country_list))
        ws2.cell(row=row, column=3, value=len(country_list))
        row += 1

    _auto_column_width(ws2)

    # ============================================
    #  Sheet 3: Top Root Causes
    # ============================================
    ws3 = wb.create_sheet("Top Root Causes")
    ws3.merge_cells("A1:D1")
    ws3.cell(row=1, column=1, value="Top 10 最常见根因路径").font = title_font

    headers3 = ["排名", "根因路径", "出现次数", "方向"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, (path, count) in enumerate(stats["top_root_causes"], 1):
        direction = "📈恶化" if "📈" in path else "📉改善"
        ws3.cell(row=i + 3, column=1, value=i).border = thin_border
        ws3.cell(row=i + 3, column=2, value=path.replace("📈", "").replace("📉", "")).border = thin_border
        ws3.cell(row=i + 3, column=3, value=count).border = thin_border
        ws3.cell(row=i + 3, column=4, value=direction).border = thin_border

    _auto_column_width(ws3)

    # ============================================
    #  Sheet 4: Top Actions
    # ============================================
    ws4 = wb.create_sheet("Top Actions")
    ws4.merge_cells("A1:D1")
    ws4.cell(row=1, column=1, value="Top 10 最常见行动建议").font = title_font

    headers4 = ["排名", "行动建议", "出现次数", "优先级分布"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, (action, count) in enumerate(stats["top_actions"], 1):
        # 提取优先级
        priority = action[1:3] if action.startswith("[") else "?"
        ws4.cell(row=i + 3, column=1, value=i).border = thin_border
        ws4.cell(row=i + 3, column=2, value=action).border = thin_border
        ws4.cell(row=i + 3, column=3, value=count).border = thin_border
        ws4.cell(row=i + 3, column=4, value=priority).border = thin_border

    _auto_column_width(ws4)

    # ============================================
    #  Sheet 5: P1/P2/P3 Distribution
    # ============================================
    ws5 = wb.create_sheet("P1 P2 P3")
    ws5.merge_cells("A1:D1")
    ws5.cell(row=1, column=1, value="P1/P2/P3 优先级分布").font = title_font

    # 汇总
    ws5.cell(row=3, column=1, value="优先级").font = header_font
    ws5.cell(row=3, column=1).fill = header_fill
    ws5.cell(row=3, column=2, value="总次数").font = header_font
    ws5.cell(row=3, column=2).fill = header_fill
    ws5.cell(row=3, column=3, value="占比").font = header_font
    ws5.cell(row=3, column=3).fill = header_fill

    total_priority = stats["p1_total"] + stats["p2_total"] + stats["p3_total"] or 1
    for i, (label, count) in enumerate([
        ("P1 紧急", stats["p1_total"]),
        ("P2 重要", stats["p2_total"]),
        ("P3 监控", stats["p3_total"]),
    ]):
        row = 4 + i
        ws5.cell(row=row, column=1, value=label).border = thin_border
        ws5.cell(row=row, column=2, value=count).border = thin_border
        ws5.cell(row=row, column=3, value=f"{count / total_priority:.1%}").border = thin_border

    # 每天 P1/P2/P3
    row = 8
    ws5.cell(row=row, column=1, value="每日 P1/P2/P3 明细").font = subtitle_font
    row += 1
    for col, h in enumerate(["日期", "国家", "阶段", "P1", "P2", "P3", "告警?"], 1):
        cell = ws5.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row += 1
    for r in results:
        if not r.has_data:
            continue
        ws5.cell(row=row, column=1, value=r.business_date).border = thin_border
        ws5.cell(row=row, column=2, value=r.country_name).border = thin_border
        ws5.cell(row=row, column=3, value=r.stage).border = thin_border
        ws5.cell(row=row, column=4, value=r.p1_count).border = thin_border
        ws5.cell(row=row, column=5, value=r.p2_count).border = thin_border
        ws5.cell(row=row, column=6, value=r.p3_count).border = thin_border
        alert_cell = ws5.cell(row=row, column=7, value="📢" if r.should_alert else "🔇")
        alert_cell.border = thin_border
        if r.should_alert:
            alert_cell.fill = alert_fill
        row += 1

    _auto_column_width(ws5)

    # ---- 保存 ----
    wb.save(output_path)
    print(f"\n  📊 Excel 报告已生成: {output_path}")
    return output_path


def _auto_column_width(ws):
    """自动调整列宽"""
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # 中文字符算 2 个宽度
                val_str = str(cell.value)
                char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, min(char_len + 2, 60))
        ws.column_dimensions[col_letter].width = max(max_len, 8)


# ============================================================
#  控制台 Summary
# ============================================================

def print_simulation_summary(stats: dict):
    """打印控制台总结"""
    print(f"\n{'='*60}")
    print(f"  📊 模拟运行总结")
    print(f"{'='*60}")

    print(f"\n  📈 告警统计")
    print(f"  {'─'*50}")
    print(f"  总任务数:     {stats['total_tasks']}")
    print(f"  有数据天数:   {stats['total_days_with_data']}")
    print(f"  发送飞书:     {stats['total_alerts']} 次")
    print(f"  静默跳过:     {stats['total_silent']} 次")
    print(f"  告警率:       {stats['alert_rate']:.1%}")

    print(f"\n  📊 优先级分布")
    print(f"  {'─'*50}")
    total_p = stats['p1_total'] + stats['p2_total'] + stats['p3_total'] or 1
    print(f"  P1 紧急:      {stats['p1_total']:>5}  ({stats['p1_total']/total_p:.1%})")
    print(f"  P2 重要:      {stats['p2_total']:>5}  ({stats['p2_total']/total_p:.1%})")
    print(f"  P3 监控:      {stats['p3_total']:>5}  ({stats['p3_total']/total_p:.1%})")

    print(f"\n  🔝 Top 5 根因路径")
    print(f"  {'─'*50}")
    for i, (path, count) in enumerate(stats['top_root_causes'][:5], 1):
        direction = "📈" if "📈" in path else "📉"
        clean = path.replace("📈", "").replace("📉", "")
        print(f"  {i}. {direction} {clean} — {count} 次")

    print(f"\n  🔝 Top 5 行动建议")
    print(f"  {'─'*50}")
    for i, (action, count) in enumerate(stats['top_actions'][:5], 1):
        print(f"  {i}. {action} — {count} 次")

    print(f"\n  📅 告警最多的天")
    print(f"  {'─'*50}")
    for date_str, count in stats['top_alert_days'][:5]:
        countries_str = ", ".join(stats['alerts_by_date'].get(date_str, []))
        print(f"  {date_str}: {count} 次 ({countries_str})")

    if stats['days_no_alert']:
        print(f"\n  🔇 无告警的天 ({len(stats['days_no_alert'])} 天)")
        print(f"  {'─'*50}")
        print(f"  {', '.join(stats['days_no_alert'][:10])}"
              f"{'...' if len(stats['days_no_alert']) > 10 else ''}")

    if stats['days_no_data']:
        print(f"\n  ❌ 无数据的天 ({len(stats['days_no_data'])} 天)")
        print(f"  {'─'*50}")
        print(f"  {', '.join(stats['days_no_data'])}")

    print(f"\n  💡 阈值评估")
    print(f"  {'─'*50}")
    print(f"  {stats['threshold_assessment']}")

    print(f"\n{'='*60}\n")
