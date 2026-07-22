#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
催收日报自动填充脚本
 - 从 Redash 催收员成绩统计报表获取数据并写入 Excel
 - --sync-google: 将 Excel 数据同步到 Google Sheet

用法:
  python auto_fill_report.py [YYYY-MM-DD]               # 从 Redash 取数填 Excel
  python auto_fill_report.py --sync-google [YYYY-MM-DD]  # Excel → Google Sheet 同步
"""

import requests, re, openpyxl, sys, io, time
from datetime import datetime, date, timedelta
from typing import Dict, Optional
from pathlib import Path

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# 定位仓库根目录，加载基础设施
_repo = Path(__file__).resolve()
while not (_repo / ".git").exists() and _repo != _repo.parent:
    _repo = _repo.parent
sys.path.insert(0, str(_repo))
from 基础设施.配置.env import require

# ======================== 配置 ========================
REDASH_URL = "http://redash.oneprestamos.com"
LOGIN_EMAIL = require("REDASH_EMAIL")
LOGIN_PASSWORD = require("REDASH_PASSWORD")
QUERY_ID = 79  # 催收员成绩统计报表
EXCEL_PATH = "D:/新建文件夹/xby/日报7.6-7.11.xlsx"
GOOGLE_CREDENTIALS = "D:/新建文件夹/xby/google_credentials.json"
GOOGLE_SHEET_KEY = "1-0sWINYtKx7Sfq_zJLijuRfGO098bJ057VTyvu5XElo"

# (阶段, 账户, 新客行号, 老客行号)
ACCOUNTS = [
    ("D-1", "YNT101", 3, 4), ("D-1", "YNT102", 6, 7),
    ("D-1", "ZYT103", 9, 10), ("D-1", "YNT103", 12, 13),
    ("D0",  "YND001", 15, 16), ("D0",  "YND002", 18, 19),
    ("D0",  "YND003", 21, 22), ("D0",  "ZYD004", 24, 25),
    ("S1",  "ZYS101", 27, 28), ("S1",  "ZYS102", 30, 31),
    ("S1",  "YNS101", 33, 34), ("S1",  "RECS101", 36, 37),
    ("S2",  "RECS201", 39, 40), ("S2",  "S2beiyong", 42, 43),
]

# 日期 -> (案件数列, 回收数列) in Excel
DATE_COLS = {
    "2026-07-11": ("G", "H"), "2026-07-10": ("K", "L"),
    "2026-07-09": ("O", "P"), "2026-07-08": ("S", "T"),
    "2026-07-07": ("W", "X"), "2026-07-06": ("AA", "AB"),
}

# 日期 -> 目标列 (金额周度) - 用于读取公式，现改用Python计算
# 目标比率表 (阶段, 客群) -> 比率（金额用）
TARGET_RATES_AMT = {
    ("D-1", "新客"): 0.28, ("D-1", "老客"): 0.40,
    ("D0", "新客"): 0.35, ("D0", "老客"): 0.25,
    ("S1", "新客"): 0.22, ("S1", "老客"): 0.25,
    ("S2", "新客"): 0.06, ("S2", "老客"): 0.06,
}
# 目标比率表（案件数用）
TARGET_RATES_CASE = {
    ("D-1", "新客"): 0.35, ("D-1", "老客"): 0.40,
    ("D0", "新客"): 0.30, ("D0", "老客"): 0.30,
    ("S1", "新客"): 0.22, ("S1", "老客"): 0.25,
    ("S2", "新客"): 0.06, ("S2", "老客"): 0.06,
}

# 非账户的sheet名称，不参与同步
GS_SKIP_SHEETS = {"Asignacion", "Temporary adjustment"}

# 账户-客群映射（None表示新老客都查）
ACCT_CUST_MAP = {
    "YNT101": "新客", "YND002": "新客",   # 只负责新客
    "YNT102": "老客", "YND001": "老客",                      # 只负责老客
    "ZYT103": "老客", "ZYD004": "老客",                      # 只负责老客
}

# 阶段汇总（用于每日文字简报）
STAGE_GROUPS = {
    "-2阶段": ["YNT301", "YNT302"],
    "-1阶段": ["YNT101", "YNT102", "YNT103", "ZYT103"],
    "D0阶段": ["YND001", "YND002", "ZYD003", "ZYD004"],
    "S1阶段": ["ZYS101", "ZYS102"],  # S1用新增绑定案件，非在手
}
# ======================================================

class RedashClient:
    def __init__(self):
        self.session = requests.Session()
        # 走系统代理（VPN提供DNS解析）
        self._login()

    def _login(self):
        r = self.session.get(f"{REDASH_URL}/login", timeout=15)
        csrf = re.search(r'name="csrf_token".*?value="([^"]+)"', r.text, re.I)
        token = csrf.group(1) if csrf else ""
        r = self.session.post(f"{REDASH_URL}/login", data={
            "csrf_token": token, "email": LOGIN_EMAIL,
            "password": LOGIN_PASSWORD, "remember": "on"
        }, allow_redirects=True, timeout=15)
        if "login" in r.url.lower():
            raise Exception("登录失败")

    def _run_query(self, acct: str, start_date: str, end_date: str,
                    cust_type: str) -> Optional[Dict]:
        """查询单个账户的数据"""
        payload = {"parameters": {
            "STARTTIME": start_date, "ENDTIME": end_date,
            "账户": [acct], "逾期天数": ["all"], "首复借客户": cust_type,
        }}
        r = self.session.post(f"{REDASH_URL}/api/queries/{QUERY_ID}/results",
                              json=payload, timeout=30)
        if r.status_code != 200:
            return None
        result = r.json()

        # 处理异步 job
        if "job" in result:
            jid = result["job"]["id"]
            for _ in range(90):
                time.sleep(1)
                r2 = self.session.get(f"{REDASH_URL}/api/jobs/{jid}", timeout=15)
                if r2.status_code == 200:
                    job = r2.json()["job"]
                    if job["status"] == 4 and job.get("query_result_id"):
                        r3 = self.session.get(
                            f"{REDASH_URL}/api/query_results/{job['query_result_id']}.json",
                            timeout=15)
                        if r3.status_code == 200:
                            rows = r3.json()["query_result"]["data"]["rows"]
                            return rows[0] if rows else None
                        break
                    elif job["status"] == 5:
                        return None
            raise Exception(f"[ERR] {cust_type} {start_date}~{end_date} 查询超时")

        rows = result.get("query_result", {}).get("data", {}).get("rows", [])
        return rows[0] if rows else None

    def _run_query_all(self, start_date: str, end_date: str,
                        cust_type: str, extra_accts: list = None) -> list:
        """一次性查询所有账户（按客群筛选账户列表，避免跨客群查询超时）"""
        # 只传对应客群的账户，减少Redash负担
        all_accts = []
        for _, acct, _, _ in ACCOUNTS:
            allowed = ACCT_CUST_MAP.get(acct)
            if not allowed or allowed == cust_type:
                all_accts.append(acct)
        # 追加额外账户（如简报用的YNT103等）
        if extra_accts:
            for a in extra_accts:
                if a not in all_accts:
                    all_accts.append(a)
        payload = {"parameters": {
            "STARTTIME": start_date, "ENDTIME": end_date,
            "账户": all_accts,  "逾期天数": ["all"], "首复借客户": cust_type,
        }}
        r = self.session.post(f"{REDASH_URL}/api/queries/{QUERY_ID}/results",
                              json=payload, timeout=30)
        if r.status_code != 200:
            return []

        result = r.json()
        # 处理异步 job
        if "job" in result:
            jid = result["job"]["id"]
            for _ in range(90):
                time.sleep(1)
                r2 = self.session.get(f"{REDASH_URL}/api/jobs/{jid}", timeout=15)
                if r2.status_code == 200:
                    job = r2.json()["job"]
                    if job["status"] == 4 and job.get("query_result_id"):
                        r3 = self.session.get(
                            f"{REDASH_URL}/api/query_results/{job['query_result_id']}.json",
                            timeout=15)
                        if r3.status_code == 200:
                            return r3.json()["query_result"]["data"]["rows"]
                        break
                    elif job["status"] == 5:
                        return []
            raise Exception(f"[ERR] {cust_type} {start_date}~{end_date} 查询超时")

        return result.get("query_result", {}).get("data", {}).get("rows", [])

    def _run_query_total(self, start_date: str, end_date: str,
                          cust_type: str) -> list:
        """用 total 方式查询，每次清除cookie缓存确保最新数据"""
        self.session.cookies.clear()
        self._login()
        payload = {"parameters": {
            "STARTTIME": start_date, "ENDTIME": end_date,
            "账户": ["total"], "逾期天数": ["all"], "首复借客户": cust_type,
        }}
        r = self.session.post(f"{REDASH_URL}/api/queries/{QUERY_ID}/results",
                              json=payload, timeout=30)
        if r.status_code != 200:
            return []
        result = r.json()
        if "job" in result:
            jid = result["job"].get("id")
            if not jid:
                return []
            # 最多等90秒，每1秒轮询
            for _ in range(90):
                time.sleep(1)
                try:
                    r2 = self.session.get(f"{REDASH_URL}/api/jobs/{jid}", timeout=10)
                    if r2.status_code == 200:
                        job = r2.json()["job"]
                        if job["status"] == 4 and job.get("query_result_id"):
                            qid = job["query_result_id"]
                            r3 = self.session.get(
                                f"{REDASH_URL}/api/query_results/{qid}.json", timeout=10)
                            if r3.status_code == 200:
                                return r3.json()["query_result"]["data"]["rows"]
                            break
                        elif job["status"] == 5:
                            return []
                except:
                    return []
            raise Exception(f"[ERR] {cust_type} {start_date}~{end_date} 查询超时")
        return result.get("query_result", {}).get("data", {}).get("rows", [])

    def fetch_rows(self, start_date: str, end_date: str) -> dict:
        """查询新客+老客，返回 {账户: {新客/老客: row}}  — 2次API调用，带重试"""
        extra = set()
        for accounts in STAGE_GROUPS.values():
            for a in accounts:
                if a not in [acct for _, acct, _, _ in ACCOUNTS]:
                    extra.add(a)
        rows_map = {}
        for ct in ["新客", "老客"]:
            for attempt in range(3):
                try:
                    print(f"[*] 查询{ct} {start_date}~{end_date}...")
                    start_time = time.time()
                    rows = self._run_query_all(start_date, end_date, ct, extra_accts=list(extra))
                    cost = round(time.time() - start_time, 1)
                    if not rows:
                        print(f"  查询完成 耗时:{cost}秒 返回:0行")
                    else:
                        print(f"  查询完成 耗时:{cost}秒 返回:{len(rows)}行")
                    for r in (rows or []):
                        a = r.get("账户", "")
                        if a not in rows_map:
                            rows_map[a] = {}
                        rows_map[a][ct] = dict(r)
                    break
                except Exception as e:
                    print(f"  {ct} 第{attempt+1}次失败: {e}")
                    if attempt < 2:
                        self.session.cookies.clear()
                        self._login()
                        time.sleep(2)
        return rows_map

    def parse_account_data(self, rows_map: dict) -> Dict:
        """从fetch_rows的结果解析账户数据（不查询API）"""
        data = {a: {"新客": {}, "老客": {}} for _, a, _, _ in ACCOUNTS}
        filtered_warnings = []

        for acct, ct_data in rows_map.items():
            if acct not in data:
                continue
            for ct, row in ct_data.items():
                allowed_ct = ACCT_CUST_MAP.get(acct)
                if allowed_ct and allowed_ct != ct:
                    stage = next((s for s, a, _, _ in ACCOUNTS if a == acct), None)
                    use_onhand = stage in ("D-1", "D0")
                    case_val = int(row.get("在手案件" if use_onhand else "新增绑定案件", 0) or 0)
                    if case_val > 0:
                        filtered_warnings.append(
                            f"⚠ {acct} 有 {ct} {case_val}件（只统计{allowed_ct}，已过滤）")
                    continue
                stage = next((s for s, a, _, _ in ACCOUNTS if a == acct), None)
                use_onhand = stage in ("D-1", "D0")
                if use_onhand:
                    case_val = int(row.get("在手案件", 0) or 0)
                    amt_val = float(row.get("在手金额", 0) or 0)
                else:
                    case_val = int(row.get("新增绑定案件", 0) or 0)
                    amt_val = float(row.get("新增绑定金额", 0) or 0)
                data[acct][ct] = {
                    "案件数": case_val,
                    "回收数": int(row.get("总回收件数", 0) or 0),
                    "入催金额": amt_val,
                    "回收金额": float(row.get("总回收金额", 0) or 0),
                }

        if filtered_warnings:
            print(f"\n{'='*50}")
            for w in filtered_warnings:
                print(w)

        total_cases = sum(data[a][c].get("案件数", 0) for a in data for c in ["新客", "老客"])
        if total_cases == 0:
            raise Exception("查询结果异常，全部案件数为0，已阻止写入Excel")

        return data


def read_excel_summary(target_date: str) -> Dict:
    """从本地 Excel 读取指定日期的汇总数据（新客+老客），目标优先读用户手动填的值"""
    from openpyxl.utils import column_index_from_string, get_column_letter
    case_col, recv_col = DATE_COLS[target_date]
    # 目标列 = 回收数列+1
    target_col = get_column_letter(column_index_from_string(recv_col) + 1)
    case_target_col = get_column_letter(column_index_from_string(case_col) + 2)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    data = {}
    ws_case = wb["案件周度"]
    ws_amt = wb["金额周度"]

    for stage, acct, nr, or_ in ACCOUNTS:
        def to_num(v):
            return v if isinstance(v, (int, float)) else 0

        # 案件数 & 回收数
        new_cases = to_num(ws_case[f"{case_col}{nr}"].value)
        old_cases = to_num(ws_case[f"{case_col}{or_}"].value)
        new_recvs = to_num(ws_case[f"{recv_col}{nr}"].value)
        old_recvs = to_num(ws_case[f"{recv_col}{or_}"].value)

        # 回收金额
        new_amt = to_num(ws_amt[f"{recv_col}{nr}"].value)
        old_amt = to_num(ws_amt[f"{recv_col}{or_}"].value)

        # 入催金额
        new_in = to_num(ws_amt[f"{case_col}{nr}"].value)
        old_in = to_num(ws_amt[f"{case_col}{or_}"].value)

        cases_total = new_cases + old_cases
        recvs_total = new_recvs + old_recvs
        amt_total = new_amt + old_amt

        # 目标：优先读用户手动填的值（非公式），否则用比率计算
        def read_target(ws, col, row, default_func):
            v = ws[f"{col}{row}"].value
            if isinstance(v, (int, float)):
                return v
            return default_func()

        new_target_amt = read_target(ws_amt, target_col, nr,
            lambda: max(1, round(new_in * TARGET_RATES_AMT.get((stage, "新客"), 0))) if new_in > 0 else 0)
        old_target_amt = read_target(ws_amt, target_col, or_,
            lambda: max(1, round(old_in * TARGET_RATES_AMT.get((stage, "老客"), 0))) if old_in > 0 else 0)
        new_target_case = read_target(ws_case, case_target_col, nr,
            lambda: max(1, round(new_cases * TARGET_RATES_CASE.get((stage, "新客"), 0))) if new_cases > 0 else 0)
        old_target_case = read_target(ws_case, case_target_col, or_,
            lambda: max(1, round(old_cases * TARGET_RATES_CASE.get((stage, "老客"), 0))) if old_cases > 0 else 0)

        data[acct] = {
            "案件数": cases_total,
            "回收数": recvs_total,
            "回收金额": amt_total,
            "目标金额": new_target_amt + old_target_amt,
            "案件目标": new_target_case + old_target_case,
        }

    wb.close()
    return data


def sync_to_google(target_date: str, force: bool = False):
    """将本地 Excel 的数据同步到 Google Sheet"""
    dt = datetime.strptime(target_date, "%Y-%m-%d")
    today_str = date.today().strftime("%Y-%m-%d")
    is_today = (target_date == today_str)
    gs_date = dt.strftime("%d/%m/%Y")  # Google Sheet 日期格式: DD/MM/YYYY

    # 1. 读取本地 Excel
    print("[*] 读取本地 Excel 数据...")
    local_data = read_excel_summary(target_date)
    if not local_data:
        print("[!] Excel 中没有数据")
        return

    print(f"\n[*] 本地数据 ({target_date}):")
    print(f"{'账户':<10} {'案件数':>6} {'回收数':>6} {'回收金额':>10} {'目标金额':>10} {'案件目标':>8}")
    print("-" * 60)
    for stage, acct, _, _ in ACCOUNTS:
        d = local_data.get(acct, {})
        if d:
            print(f"{acct:<10} {d['案件数']:>6} {d['回收数']:>6} {d['回收金额']:>10.2f} {d['目标金额']:>10.2f} {d.get('案件目标',0):>8}")

    # 2. 连接 Google Sheet
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(GOOGLE_CREDENTIALS, scopes=scope)
        client = gspread.authorize(creds)
        sh = client.open_by_key(GOOGLE_SHEET_KEY)
        print("[OK] Google Sheet 连接成功")
    except Exception as e:
        print(f"[ERR] Google Sheet 连接失败: {e}")
        return

    # 3. 获取所有 sheet 名称，匹配账户
    all_sheets = {s.title for s in sh.worksheets() if s.title not in GS_SKIP_SHEETS}
    matched = [a for _, a, _, _ in ACCOUNTS if a in all_sheets]

    has_conflict = False
    print(f"\n[*] 同步到 Google Sheet ({gs_date}):")

    for acct in matched:
        d = local_data.get(acct, {"案件数":0,"回收数":0,"回收金额":0,"目标金额":0,"案件目标":0})

        try:
            ws = sh.worksheet(acct)
        except:
            print(f"  {acct}: Sheet 不存在，跳过")
            continue

        # 查找日期行（column A）
        all_dates = ws.col_values(1)
        try:
            row_idx = all_dates.index(gs_date) + 1  # 1-based row index
        except ValueError:
            print(f"  {acct}: 未找到日期 {gs_date}，跳过")
            continue

        updates = []
        # 全部日期写入全字段：入催数(C) + 案件目标(D) + 回收数(E) + 目标金额(G) + 回收金额(H)
        new_in = d["案件数"]
        new_ct = d.get("案件目标", 0)
        new_rv = d["回收数"]
        new_ta = d["目标金额"]
        new_ra = round(d["回收金额"], 2)

        updates.append({"range": f"C{row_idx}", "values": [[new_in]]})
        updates.append({"range": f"D{row_idx}", "values": [[new_ct]]})
        updates.append({"range": f"E{row_idx}", "values": [[new_rv]]})
        updates.append({"range": f"G{row_idx}", "values": [[new_ta]]})
        updates.append({"range": f"H{row_idx}", "values": [[new_ra]]})

        # 批量写入
        if updates:
            ws.batch_update(updates, value_input_option='USER_ENTERED')
            print(f"  {acct}: 入催={new_in}, 案件目标={new_ct}, 回收={new_rv}, 金额目标={new_ta}, 回收金额={new_ra}")

    if has_conflict:
        print(f"\n[⚠] 存在数据冲突，请确认后再用 --sync-google --force 覆盖")
    else:
        print(f"\n[OK] 同步完成")


def print_stage_summary(rows_map: dict, target_date: str):
    """按阶段汇总数据（使用已查询的rows_map，不再查API）"""
    print(f"\n{'='*50}")
    print(f"  📊 回收简报 ({target_date})")
    print(f"{'='*50}")

    for stage, accounts in STAGE_GROUPS.items():
        total_hand = 0
        total_recv = 0
        total_amt = 0.0
        acct_info = []

        for a in accounts:
            if a in rows_map:
                # merge 新客+老客（只合并数字字段）
                merged = {}
                NUMERIC_FIELDS = ["在手案件","在手金额","新增绑定案件","新增绑定金额",
                                  "总回收件数","总回收金额"]
                for ct_data in rows_map[a].values():
                    for k in NUMERIC_FIELDS:
                        if k in ct_data:
                            merged[k] = (merged.get(k, 0) or 0) + (ct_data[k] or 0)
                use_new_bind = (stage == "S1阶段")
                h = int(merged.get("新增绑定案件" if use_new_bind else "在手案件", 0) or 0)
                rv = int(merged.get("总回收件数", 0) or 0)
                ra = float(merged.get("总回收金额", 0) or 0)
                total_hand += h
                total_recv += rv
                total_amt += ra
                field_name = "新增" if use_new_bind else "在手"
                acct_info.append(f"{a}({field_name}{h}/回收{rv})")

        if acct_info:
            label = "新增绑定" if stage == "S1阶段" else "在手"
            print(f"\n  {stage}:")
            print(f"    {label}: {total_hand} 件 | 回收: {total_recv} 件, {total_amt:.2f}")
            print(f"    {' '.join(acct_info)}")


def main():
    # 参数解析
    is_sync = "--sync-google" in sys.argv
    force = "--force" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]

    target_date = date.today().strftime("%Y-%m-%d")
    if args:
        target_date = args[0]

    if target_date not in DATE_COLS:
        print(f"[!] 不支持的日期: {target_date}，支持: {list(DATE_COLS.keys())}")
        return

    if is_sync:
        sync_to_google(target_date, force=force)
        # 同步昨天回收
        dt = datetime.strptime(target_date, "%Y-%m-%d")
        yesterday = (dt - timedelta(days=1)).strftime("%Y-%m-%d")
        if yesterday in DATE_COLS:
            print(f"\n[*] 自动补推昨日({yesterday})...")
            sync_to_google(yesterday, force=force)
        return

    # ====== 正常取数流程 ======
    dt = datetime.strptime(target_date, "%Y-%m-%d")
    next_date = (dt + timedelta(days=1)).strftime("%Y-%m-%d")

    print(f"[*] 目标日期: {target_date}")

    # 1. 登录
    try:
        redash = RedashClient()
        print("[OK] Redash 登录成功")
    except Exception as e:
        print(f"[ERR] 登录失败: {e}")
        return

    # 2. 查询：今天(2次) + 昨天(2次) = 4次API，数据全在这
    print("[*] 正在从催收员成绩统计报表获取数据...")
    today_rows = redash.fetch_rows(target_date, next_date)
    account_data = redash.parse_account_data(today_rows)

    # 3. 打印摘要
    print(f"\n[*] 数据摘要:")
    print(f"{'账户':<10} {'客群':<4} {'案件数':<6} {'回收数':<6} {'入催金额':<12} {'回收金额':<12}")
    print("-" * 60)
    for _, acct, nr, or_ in ACCOUNTS:
        for ct, _ in [("新客", nr), ("老客", or_)]:
            d = account_data.get(acct, {}).get(ct, {})
            print(f"{acct:<10} {ct:<4} {d.get('案件数',0):<6} {d.get('回收数',0):<6} "
                  f"{d.get('入催金额',0):<12.2f} {d.get('回收金额',0):<12.2f}")

    # 4. 简报 = 复用 today_rows（当天）或查昨天
    is_today_run = (target_date == date.today().strftime("%Y-%m-%d"))
    if is_today_run:
        briefing_date = (dt - timedelta(days=1)).strftime("%Y-%m-%d")
        briefing_next = dt.strftime("%Y-%m-%d")
    else:
        briefing_date = target_date
        briefing_next = next_date

    if is_today_run:
        yesterday_rows = redash.fetch_rows(briefing_date, briefing_next)
        print_stage_summary(yesterday_rows, briefing_date)
    else:
        print_stage_summary(today_rows, briefing_date)

    # 5. 写入Excel
    is_today_fill = (target_date == date.today().strftime("%Y-%m-%d"))
    print(f"\n[*] 写入 Excel{'（当天不填回收数）' if is_today_fill else ''}...")
    try:
        case_col, recv_col = DATE_COLS[target_date]
        wb = openpyxl.load_workbook(EXCEL_PATH)

        for sheet_name in ["案件周度", "金额周度"]:
            ws = wb[sheet_name]
            for _, acct, nr, or_ in ACCOUNTS:
                d = account_data.get(acct, {})
                is_case = sheet_name == "案件周度"
                for ct, rn in [("新客", nr), ("老客", or_)]:
                    dd = d.get(ct, {})
                    if is_case:
                        ws[f"{case_col}{rn}"] = dd.get("案件数", 0)
                        if not is_today_fill:
                            ws[f"{recv_col}{rn}"] = dd.get("回收数", 0)
                    else:
                        ws[f"{case_col}{rn}"] = round(dd.get("入催金额", 0), 2)
                        if not is_today_fill:
                            ws[f"{recv_col}{rn}"] = round(dd.get("回收金额", 0), 2)

                # ===== 自动写全周汇总公式 =====
                sum_row = or_ + 1
                from openpyxl.utils import column_index_from_string, get_column_letter
                for dc, dv in DATE_COLS.values():
                    target_col = get_column_letter(column_index_from_string(dc) + 2)
                    rate_col = get_column_letter(column_index_from_string(dc) + 3)
                    ws[f"{dc}{sum_row}"] = f"={dc}{nr}+{dc}{or_}"
                    ws[f"{dv}{sum_row}"] = f"={dv}{nr}+{dv}{or_}"
                    ws[f"{target_col}{sum_row}"] = f"={target_col}{nr}+{target_col}{or_}"
                    ws[f"{rate_col}{sum_row}"] = f'=IF({target_col}{sum_row}=0,0,{dv}{sum_row}/{target_col}{sum_row})'

        wb.save(EXCEL_PATH)
        wb.close()
        print(f"[OK] 数据已写入 {EXCEL_PATH}")
    except Exception as e:
        print(f"[ERR] Excel 写入失败: {e}")
        import traceback; traceback.print_exc()

    # 6. 补填昨天回收 = 复用 yesterday_rows（已是4次查询后，不再重复查）
    if is_today_fill:
        yesterday = (dt - timedelta(days=1)).strftime("%Y-%m-%d")
        if yesterday in DATE_COLS:
            print(f"\n[*] 补填昨日({yesterday})回收数...")
            try:
                recv_col_y = DATE_COLS[yesterday][1]
                wb = openpyxl.load_workbook(EXCEL_PATH)
                for sheet_name in ["案件周度", "金额周度"]:
                    ws = wb[sheet_name]
                    is_case = sheet_name == "案件周度"
                    for _, acct, nr, or_ in ACCOUNTS:
                        nd = yesterday_rows.get(acct, {}).get("新客", {})
                        od = yesterday_rows.get(acct, {}).get("老客", {})
                        cust = ACCT_CUST_MAP.get(acct)
                        if cust == "新客":
                            ws[f"{recv_col_y}{nr}"] = int(nd.get("总回收件数", 0) or 0)
                            if not is_case:
                                ws[f"{recv_col_y}{nr}"] = round(float(nd.get("总回收金额", 0) or 0), 2)
                        elif cust == "老客":
                            ws[f"{recv_col_y}{or_}"] = int(od.get("总回收件数", 0) or 0)
                            if not is_case:
                                ws[f"{recv_col_y}{or_}"] = round(float(od.get("总回收金额", 0) or 0), 2)
                        else:
                            ws[f"{recv_col_y}{nr}"] = int(nd.get("总回收件数", 0) or 0)
                            ws[f"{recv_col_y}{or_}"] = int(od.get("总回收件数", 0) or 0)
                            if not is_case:
                                ws[f"{recv_col_y}{nr}"] = round(float(nd.get("总回收金额", 0) or 0), 2)
                                ws[f"{recv_col_y}{or_}"] = round(float(od.get("总回收金额", 0) or 0), 2)

                        sum_row = or_ + 1
                        for dc, dv in DATE_COLS.values():
                            ws[f"{dv}{sum_row}"] = f"={dv}{nr}+{dv}{or_}"
                wb.save(EXCEL_PATH)
                wb.close()
                print(f"[OK] 昨日回收数已补填")
            except Exception as e:
                print(f"[ERR] 补填失败: {e}")


if __name__ == "__main__":
    main()
