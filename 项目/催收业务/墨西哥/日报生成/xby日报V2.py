#!/usr/bin/env python3
# -*- coding: utf-8 -*-
print("Python启动", flush=True)
"""
xby日报 V2  —  催收日报自动填充脚本

流程:
  Redash 查询 (最多4次/周一2次) → 更新Excel → 生成简报 → 用户确认 → 同步Google

设计原则:
  - Redash 查询结束后，所有模块禁止再次访问 Redash
  - Excel 只写入数据值，不触碰公式 / 汇总 / 目标 / 达成率
  - Google 同步从 Excel 读取（用户可能手动修改过数据）
  - 每次启动全新 Session，查询结束立即释放

用法:
  python xby日报V2.py                       # 当天运行
  python xby日报V2.py 2026-07-09            # 指定日期
  python xby日报V2.py --sync-only           # 仅同步 Google（读Excel）
  python xby日报V2.py --dry-run             # 仅查询+简报，不写Excel/Google
  python xby日报V2.py --query-only          # 仅查询+保存原始JSON
"""

VERSION = "2.2.0"
BUILD   = "2026-07-07"

import requests, re, openpyxl, sys, io, time, os, shutil
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field
from functools import wraps
from openpyxl.utils import get_column_letter
from pathlib import Path

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# 定位仓库根目录，加载基础设施
_repo = Path(__file__).resolve()
while not (_repo / ".git").exists() and _repo != _repo.parent:
    _repo = _repo.parent
sys.path.insert(0, str(_repo))
from 基础设施.配置.env import require

# ╔══════════════════════════════════════════════════════════╗
# ║              配  置  区  (所有可修改参数)                  ║
# ╚══════════════════════════════════════════════════════════╝

# ── 运行模式 ──
AUTO_SYNC_GOOGLE   = False   # True = 跳过确认，直接同步 Google
ENABLE_BACKUP      = True    # True = 写 Excel 前自动备份
SHOW_ZERO_RECOVERY = False   # True = 简报显示回收为 0 的账户

# ── Redash 缓存 ──
USE_REDASH_CACHE = True      # True = 使用 Redash 默认缓存 (最快)
                              # False = max_age=0 强制实时查询

# ── Redash ──
REDASH_URL       = "http://redash.oneprestamos.com"
LOGIN_EMAIL      = require("REDASH_EMAIL")
LOGIN_PASSWORD   = require("REDASH_PASSWORD")
QUERY_ID         = 79

# ── 路径 ──
EXCEL_DIR         = "D:/新建文件夹/xby"
EXCEL_PATH        = f"{EXCEL_DIR}/日报7.6-7.11.xlsx"
GOOGLE_CREDENTIALS = f"{EXCEL_DIR}/google_credentials.json"
GOOGLE_SHEET_KEY   = "1-0sWINYtKx7Sfq_zJLijuRfGO098bJ057VTyvu5XElo"
LOG_DIR           = f"{EXCEL_DIR}/logs"

# (阶段, 账户, 新客行号, 老客行号)
ACCOUNTS = [
    ("D-1", "YNT101", 3, 4),  ("D-1", "YNT102", 6, 7),
    ("D-1", "ZYT103", 9, 10),
    ("D0",  "YND001", 12, 13), ("D0",  "YND002", 15, 16),
    ("D0",  "ZYD003", 18, 19), ("D0",  "ZYD004", 21, 22),
    ("S1",  "ZYS101", 24, 25), ("S1",  "ZYS102", 27, 28),
    ("S1",  "YNS101", 30, 31), ("S1",  "RECS101", 33, 34),
    ("S2",  "RECS201", 36, 37), ("S2",  "S2beiyong", 39, 40),
]

# 账户 → 客群映射 (None = 两种客群都统计)
ACCT_CUST_MAP = {
    "YNT101": "新客", "YND002": "新客", "ZYD003": "新客",
    "YNT102": "老客", "YND001": "老客",
    "ZYT103": "老客", "ZYD004": "老客",
}

# 阶段分组（用于文字简报）
STAGE_GROUPS = {
    "-2阶段": ["YNT301", "YNT302"],
    "-1阶段": ["YNT101", "YNT102", "YNT103", "ZYT103"],
    "D0阶段": ["YND001", "YND002", "ZYD003", "ZYD004"],
    "S1阶段": ["ZYS101", "ZYS102"],
}

# 目标比率表（金额）
TARGET_RATES_AMT = {
    ("D-1", "新客"): 0.28, ("D-1", "老客"): 0.40,
    ("D0",  "新客"): 0.35, ("D0",  "老客"): 0.25,
    ("S1",  "新客"): 0.22, ("S1",  "老客"): 0.25,
    ("S2",  "新客"): 0.06, ("S2",  "老客"): 0.06,
}

# 目标比率表（案件数）
TARGET_RATES_CASE = {
    ("D-1", "新客"): 0.35, ("D-1", "老客"): 0.40,
    ("D0",  "新客"): 0.30, ("D0",  "老客"): 0.30,
    ("S1",  "新客"): 0.22, ("S1",  "老客"): 0.25,
    ("S2",  "新客"): 0.06, ("S2",  "老客"): 0.06,
}

# Google Sheet 非账户工作表
GS_SKIP_SHEETS = {"Asignacion", "Temporary adjustment"}

# 重试配置
MAX_RETRIES = 3
RETRY_DELAY = 2


# ╔══════════════════════════════════════════════════════════╗
# ║            通  用  工  具                                  ║
# ╚══════════════════════════════════════════════════════════╝

def make_weekly_filename(week_start: date, week_end: date) -> str:
    """生成周报文件名，无前导零: 日报7.6-7.11.xlsx"""
    return f"日报{week_start.month}.{week_start.day}-{week_end.month}.{week_end.day}.xlsx"


def calc_targets(stage: str, new_cases: int, old_cases: int,
                 new_amt: float, old_amt: float) -> Tuple[int, float]:
    """
    计算账户汇总目标（新客+老客合并）。
    返回: (案件目标, 金额目标)

    规则: 每个客群独立应用目标比率，然后相加。
    """
    def _case_target(cases, cust):
        rate = TARGET_RATES_CASE.get((stage, cust), 0)
        return max(1, round(cases * rate)) if cases > 0 else 0

    def _amt_target(amt, cust):
        rate = TARGET_RATES_AMT.get((stage, cust), 0)
        return max(1, round(amt * rate)) if amt > 0 else 0

    ct = _case_target(new_cases, "新客") + _case_target(old_cases, "老客")
    at = _amt_target(new_amt, "新客") + _amt_target(old_amt, "老客")
    return ct, float(at)


def retry(max_attempts=MAX_RETRIES, delay=RETRY_DELAY, label=""):
    """统一重试装饰器：失败自动重试，最多 max_attempts 次"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_error = None
            for attempt in range(1, max_attempts + 1):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    last_error = e
                    if attempt < max_attempts:
                        tag = f"[{label}] " if label else ""
                        print(f"  {tag}第 {attempt}/{max_attempts} 次失败: {e}，"
                              f"{delay}s 后重试...")
                        time.sleep(delay)
            raise last_error
        return wrapper
    return decorator


# ╔══════════════════════════════════════════════════════════╗
# ║              日  志  &  计  时                            ║
# ╚══════════════════════════════════════════════════════════╝

class Logger:
    """计时 + 事件记录 + 日志文件"""

    def __init__(self):
        self._times: Dict[str, float] = {}
        self._order: List[str] = []
        self._started: Dict[str, float] = {}
        # ── 事件记录 ──
        self.query_counts: Dict[str, int] = {}    # label → 返回行数
        self.query_cache_info: Dict[str, dict] = {} # label → {"query_result_id": ..., "source": ...}
        self.new_accounts: List[str] = []
        self.errors: List[str] = []
        self.sync_info: str = ""
        self.excel_status: str = ""     # 日报状态
        self.google_status: str = ""    # Google 同步状态

    # ── 计时 ──
    def start(self, name: str):
        self._started[name] = time.time()

    def stop(self, name: str):
        if name in self._started:
            elapsed = time.time() - self._started.pop(name)
            self._times[name] = elapsed
            if name not in self._order:
                self._order.append(name)

    def elapsed(self, name: str) -> float:
        return self._times.get(name, 0)

    def total(self) -> float:
        return sum(self._times.values())

    # ── 事件 ──
    def log_query(self, label: str, row_count: int):
        self.query_counts[label] = row_count

    def log_query_cache(self, label: str, cache_info: dict):
        self.query_cache_info[label] = cache_info

    def log_new_accounts(self, accounts: List[str]):
        self.new_accounts = accounts

    def log_error(self, msg: str):
        self.errors.append(msg)

    def log_sync(self, info: str):
        self.sync_info = info

    # ── 输出 ──
    def summary(self) -> str:
        lines = []
        for name in self._order:
            if name in self._times:
                lines.append(f"  {name}: {self._times[name]:.1f}s")
        lines.append(f"  总耗时: {self.total():.1f}s")
        return "\n".join(lines)

    def print_summary(self):
        print(f"\n{'─'*40}")
        print("⏱ 耗时统计:")
        print(self.summary())
        print(f"{'─'*40}")

    def save(self, target_date: str):
        """保存日志到 logs/ 目录"""
        os.makedirs(LOG_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = f"{LOG_DIR}/daily_{target_date}_{timestamp}.log"

        # 判定结果
        has_errors = len(self.errors) > 0
        result = "FAIL" if has_errors else "PASS"

        with open(log_path, "w", encoding="utf-8") as f:
            f.write(f"xby日报 V2  v{VERSION}  Build {BUILD}\n")
            f.write(f"日期: {target_date}\n")
            f.write(f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            if self.query_counts:
                f.write("Query 返回:\n")
                for label in ["今天新客", "今天老客", "昨天新客", "昨天老客"]:
                    if label in self.query_counts:
                        cache = self.query_cache_info.get(label, {})
                        qid = cache.get("query_result_id", "N/A")
                        src = cache.get("source", "N/A")
                        f.write(f"  {label}: {self.query_counts[label]} 行  "
                                f"query_result_id={qid}  来源={src}\n")
                f.write("\n")

            f.write("耗时统计:\n")
            f.write(self.summary())
            f.write("\n\n")

            if self.new_accounts:
                f.write(f"发现新账户: {', '.join(self.new_accounts)}\n\n")

            if self.errors:
                f.write("异常:\n")
                for e in self.errors:
                    f.write(f"  {e}\n")
                f.write("\n")

            if self.sync_info:
                f.write(f"Google 同步: {self.sync_info}\n\n")

            f.write(f"RESULT: {result}\n")

        print(f"\n📝 日志已保存: {log_path}")


# ╔══════════════════════════════════════════════════════════╗
# ║         Redash  客  户  端  (全新 Session / 带重试)         ║
# ╚══════════════════════════════════════════════════════════╝

class RedashClient:
    """
    每次实例化 = 全新 Session = 清除所有 Cookie。
    生命周期: 创建 → query(最多4次) → close()
    登录和查询均支持自动重试。
    """

    def __init__(self, logger: Logger):
        self.logger = logger
        self.session = requests.Session()
        self.session.trust_env = False
        self._login_with_retry()

    def _login_with_retry(self):
        """带重试的登录"""
        last_error = None
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                self._do_login()
                return
            except Exception as e:
                last_error = e
                if attempt < MAX_RETRIES:
                    print(f"  [登录] 第 {attempt}/{MAX_RETRIES} 次失败: {e}，"
                          f"{RETRY_DELAY}s 后重试...")
                    self.session = requests.Session()
                    self.session.trust_env = False
                    time.sleep(RETRY_DELAY)
        raise last_error

    def _do_login(self):
        """执行登录：获取 CSRF token → POST 登录"""
        print("[登录] 正在连接 Redash ...", end=" ", flush=True)
        t0 = time.time()
        r = self.session.get(f"{REDASH_URL}/login", timeout=15)
        csrf = re.search(r'name="csrf_token".*?value="([^"]+)"', r.text, re.I)
        token = csrf.group(1) if csrf else ""
        r = self.session.post(f"{REDASH_URL}/login", data={
            "csrf_token": token, "email": LOGIN_EMAIL,
            "password": LOGIN_PASSWORD, "remember": "on",
        }, allow_redirects=True, timeout=15)
        if "login" in r.url.lower():
            raise Exception("Redash 登录失败，请检查账号密码")
        print(f"成功 ({time.time()-t0:.1f}s)", flush=True)

    @retry(max_attempts=MAX_RETRIES, delay=RETRY_DELAY, label="Redash查询")
    def query(self, start_date: str, end_date: str,
              cust_type: str,
              accounts: Optional[List[str]] = None,
              ) -> Tuple[List[dict], dict]:
        """
        统一查询接口 — 整个程序唯一调用 Redash 的地方。
        支持自动重试（装饰器）。
        返回: (rows, cache_info)  cache_info = {"query_result_id": ..., "source": "Cache"|"Fresh"}
        """
        if accounts is None:
            accounts = []
            for _, acct, _, _ in ACCOUNTS:
                allowed = ACCT_CUST_MAP.get(acct)
                if not allowed or allowed == cust_type:
                    accounts.append(acct)
            for group_accts in STAGE_GROUPS.values():
                for a in group_accts:
                    if a not in accounts:
                        accounts.append(a)

        payload = {"parameters": {
            "STARTTIME": start_date, "ENDTIME": end_date,
            "账户": accounts, "逾期天数": ["all"], "首复借客户": cust_type,
        }}
        if not USE_REDASH_CACHE:
            payload["max_age"] = 0  # 强制绕过缓存

        r = self.session.post(
            f"{REDASH_URL}/api/queries/{QUERY_ID}/results",
            json=payload, timeout=30)
        if r.status_code != 200:
            raise Exception(f"Redash 返回 HTTP {r.status_code}")

        result = r.json()

        # ── 直接返回缓存 ──
        if "query_result" in result:
            qr = result["query_result"]
            qr_id = qr.get("id", "N/A")
            rows = qr.get("data", {}).get("rows", [])
            return rows, {"query_result_id": str(qr_id), "source": "Cache"}

        # ── 异步 Job ──
        if "job" in result:
            jid = result["job"].get("id")
            if not jid:
                return [], {"query_result_id": "N/A", "source": "Fresh"}
            wait_start = time.time()
            for i in range(1, 31):  # 单次最长等待 30 秒
                time.sleep(1)
                r2 = self.session.get(f"{REDASH_URL}/api/jobs/{jid}", timeout=15)
                if r2.status_code != 200:
                    continue
                job = r2.json()["job"]
                if job["status"] == 4 and job.get("query_result_id"):
                    qid = job["query_result_id"]
                    r3 = self.session.get(
                        f"{REDASH_URL}/api/query_results/{qid}.json", timeout=15)
                    if r3.status_code == 200:
                        rows = r3.json()["query_result"]["data"]["rows"]
                        return rows, {"query_result_id": str(qid), "source": "Fresh"}
                    break
                elif job["status"] == 5:
                    return [], {"query_result_id": "N/A", "source": "Fresh"}
                if i % 10 == 0:
                    print(f"  [查询] 等待中... {i}s / 30s", flush=True)
            waited = time.time() - wait_start
            raise Exception(f"查询超时 ({waited:.0f}s): {cust_type} {start_date}~{end_date}")

        return [], {"query_result_id": "N/A", "source": "Cache"}

    def close(self):
        self.session.close()


# ╔══════════════════════════════════════════════════════════╗
# ║          查  询  结  果  (不可变数据中心)                   ║
# ╚══════════════════════════════════════════════════════════╝

@dataclass(frozen=True)
class QueryResult:
    """Redash 全部查询结果 — 创建后只读"""
    target_date:     str
    today_start:     str
    today_end:       str
    yesterday_start: str
    yesterday_end:   str
    today_rows:      Dict[str, Dict[str, dict]] = field(default_factory=dict)
    yesterday_rows:  Dict[str, Dict[str, dict]] = field(default_factory=dict)

    @classmethod
    def from_redash(cls, redash: RedashClient, target_date: str, logger: Logger):
        """执行最多 4 次查询（周一仅2次），构建 QueryResult"""
        dt = datetime.strptime(target_date, "%Y-%m-%d")
        is_monday = (dt.weekday() == 0)
        today_start     = target_date
        today_end       = (dt + timedelta(days=1)).strftime("%Y-%m-%d")
        yesterday_start = (dt - timedelta(days=1)).strftime("%Y-%m-%d")
        yesterday_end   = target_date

        today_rows     = cls._fetch_both(redash, today_start, today_end,
                                          "今天", logger)
        if is_monday:
            print("[查询] 周一，跳过昨天查询（周日无回收数据）", flush=True)
            yesterday_rows = {}
        else:
            yesterday_rows = cls._fetch_both(redash, yesterday_start, yesterday_end,
                                              "昨天", logger)

        return cls(
            target_date=target_date,
            today_start=today_start, today_end=today_end,
            yesterday_start=yesterday_start, yesterday_end=yesterday_end,
            today_rows=today_rows, yesterday_rows=yesterday_rows,
        )

    @staticmethod
    def _fetch_both(redash: RedashClient, start: str, end: str,
                    label: str, logger: Logger) -> Dict[str, Dict[str, dict]]:
        """查询新客 + 老客 = 2次 API 调用"""
        rows_map: Dict[str, Dict[str, dict]] = {}
        for ct in ["新客", "老客"]:
            tag = f"{label}{ct}"
            print(f"[查询] 开始查询 {tag} ({start} ~ {end}) ...", flush=True)
            logger.start(tag)
            rows, cache_info = redash.query(start, end, ct)
            logger.stop(tag)
            logger.log_query(tag, len(rows))
            logger.log_query_cache(tag, cache_info)
            source_icon = "📦" if cache_info["source"] == "Cache" else "🔄"
            print(f"[查询] {tag} 完成 → {len(rows)}行 {logger.elapsed(tag):.1f}s  "
                  f"{source_icon} {cache_info['source']}  "
                  f"query_result_id={cache_info['query_result_id']}", flush=True)
            for r in (rows or []):
                a = r.get("账户", "")
                if a not in rows_map:
                    rows_map[a] = {}
                rows_map[a][ct] = dict(r)
        return rows_map

    def get_cell(self, acct: str, cust: str) -> dict:
        return self.today_rows.get(acct, {}).get(cust, {})

    def get_value(self, acct: str, cust: str, field: str, default=0):
        return self.get_cell(acct, cust).get(field, default) or default

    def is_empty(self) -> bool:
        total = 0
        for _, acct, _, _ in ACCOUNTS:
            for ct in ("新客", "老客"):
                stage = next((s for s, a, _, _ in ACCOUNTS if a == acct), "")
                field = "在手案件" if stage in ("D-1", "D0") else "新增绑定案件"
                total += int(self.get_value(acct, ct, field))
        return total == 0

    def find_new_accounts(self) -> List[str]:
        # 已知账户 = ACCOUNTS 中的 + STAGE_GROUPS 中的（简报用，不写Excel）
        known = {acct for _, acct, _, _ in ACCOUNTS}
        known.update(a for lst in STAGE_GROUPS.values() for a in lst)
        new_accounts = []
        for acct in self.today_rows:
            if acct not in known and acct != "total":
                has_data = any(
                    v and v != 0
                    for ct_data in self.today_rows[acct].values()
                    for v in ct_data.values()
                )
                if has_data:
                    new_accounts.append(acct)
        return new_accounts


# ╔══════════════════════════════════════════════════════════╗
# ║         Excel  操  作  (只写数据，不动公式)                ║
# ╚══════════════════════════════════════════════════════════╝

class ExcelHelper:
    """Excel 操作：日期识别 / 周一复制 / 写入数据 / 备份"""

    def __init__(self, excel_path: str = EXCEL_PATH):
        self.excel_path = excel_path

    # ── 日期识别 ──────────────────────────────────────────

    @staticmethod
    def _parse_date_value(value) -> Optional[date]:
        """
        解析单元格值为 date 对象。
        支持: datetime / date / 常见字符串格式
        """
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            value = value.strip()
            formats = ["%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y",
                       "%Y.%m.%d", "%m.%d.%Y"]
            for fmt in formats:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def detect_dates(self) -> Dict[str, Tuple[str, str]]:
        """
        扫描 Excel 案件周度 Row 1，找到所有日期列。
        返回: {"YYYY-MM-DD": (案件数列, 回收数列), ...}  按日期降序
        """
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb["案件周度"]
        dates = {}

        for col_idx in range(1, ws.max_column + 1):
            parsed = self._parse_date_value(ws.cell(row=1, column=col_idx).value)
            if parsed:
                ds = parsed.strftime("%Y-%m-%d")
                cl = get_column_letter(col_idx)
                rl = get_column_letter(col_idx + 1)
                dates[ds] = (cl, rl)

        wb.close()
        return dict(sorted(dates.items(), key=lambda x: x[0], reverse=True))

    def has_date(self, target_date: str) -> bool:
        return target_date in self.detect_dates()

    # ── 周一复制 ──────────────────────────────────────────

    def handle_monday(self, target_date: str, logger: Logger) -> str:
        """
        周一自动复制上周日报 → 生成本周日报。

        1. 非周一直接返回
        2. 如果本周日报已存在，直接使用
        3. 找到上周日报文件 → 复制 → 所有日期 +7 天

        返回: 最终使用的 Excel 路径
        """
        dt = datetime.strptime(target_date, "%Y-%m-%d")
        if dt.weekday() != 0:
            return self.excel_path

        week_start = dt
        week_end = dt + timedelta(days=5)

        new_name = make_weekly_filename(week_start, week_end)
        new_path = f"{EXCEL_DIR}/{new_name}"

        if os.path.exists(new_path):
            print(f"  [*] 本周日报已存在: {new_name}")
            self.excel_path = new_path
            return new_path

        last_week_start = week_start - timedelta(days=7)
        last_week_end   = week_end - timedelta(days=7)
        old_name = make_weekly_filename(last_week_start, last_week_end)
        old_path = f"{EXCEL_DIR}/{old_name}"

        if not os.path.exists(old_path):
            print(f"  [!] 上周日报不存在: {old_name}，跳过复制")
            return self.excel_path

        print(f"\n  📅 今天是周一，自动生成本周日报...")
        print(f"     复制: {old_name}")
        print(f"     生成: {new_name}")

        logger.start("周一复制日报")
        shutil.copy2(old_path, new_path)

        wb = openpyxl.load_workbook(new_path)
        for sheet_name in ["案件周度", "金额周度"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                if isinstance(cell.value, datetime):
                    cell.value = cell.value + timedelta(days=7)

        wb.save(new_path)
        wb.close()

        self.excel_path = new_path
        logger.stop("周一复制日报")
        print(f"     [OK] 日期已自动更新 (+7天)")
        return new_path

    # ── 备份 ──────────────────────────────────────────────

    def backup(self):
        """写入前生成带时间戳的 _backup 副本"""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.excel_path.replace(".xlsx", f"_backup_{ts}.xlsx")
        shutil.copy2(self.excel_path, backup_path)
        print(f"  [*] 已备份: {os.path.basename(backup_path)}")

    # ── 写入数据 ──────────────────────────────────────────

    def write_today(self, qr: QueryResult, target_date: str):
        """
        写入今天数据到 Excel — 只写数据值，不碰公式。

        案件周度: 案件数 / 回收数
        金额周度: 入催金额 / 回收金额

        当天运行: 只写案件数/入催金额，不写回收数/回收金额。
        """
        dates = self.detect_dates()
        if target_date not in dates:
            print(f"  [!] Excel 中未找到日期 {target_date}")
            return

        case_col, recv_col = dates[target_date]
        is_today = (target_date == date.today().strftime("%Y-%m-%d"))
        wb = openpyxl.load_workbook(self.excel_path)

        for sheet_name in ["案件周度", "金额周度"]:
            ws = wb[sheet_name]
            is_case = (sheet_name == "案件周度")

            for stage, acct, nr, or_ in ACCOUNTS:
                use_onhand = stage in ("D-1", "D0")
                case_field = "在手案件" if use_onhand else "新增绑定案件"
                amt_field  = "在手金额" if use_onhand else "新增绑定金额"

                nd = qr.get_cell(acct, "新客")
                od = qr.get_cell(acct, "老客")

                if is_case:
                    ws[f"{case_col}{nr}"] = int(nd.get(case_field, 0) or 0)
                    ws[f"{case_col}{or_}"] = int(od.get(case_field, 0) or 0)
                    if not is_today:
                        ws[f"{recv_col}{nr}"] = int(nd.get("总回收件数", 0) or 0)
                        ws[f"{recv_col}{or_}"] = int(od.get("总回收件数", 0) or 0)
                else:
                    ws[f"{case_col}{nr}"] = round(float(nd.get(amt_field, 0) or 0), 2)
                    ws[f"{case_col}{or_}"] = round(float(od.get(amt_field, 0) or 0), 2)
                    if not is_today:
                        ws[f"{recv_col}{nr}"] = round(float(nd.get("总回收金额", 0) or 0), 2)
                        ws[f"{recv_col}{or_}"] = round(float(od.get("总回收金额", 0) or 0), 2)

        wb.save(self.excel_path)
        wb.close()

        fill_type = "仅案件数" if is_today else "案件+回收"
        print(f"  [OK] 已写入 {target_date} ({fill_type}) → "
              f"{os.path.basename(self.excel_path)}")


# ╔══════════════════════════════════════════════════════════╗
# ║         简  报  (纯内存数据，不读 Excel)                   ║
# ╚══════════════════════════════════════════════════════════╝

def print_briefing(qr: QueryResult, target_date: str):
    """从 QueryResult 生成昨日简报（仅阶段汇总，不显示账号明细）"""
    dt = datetime.strptime(target_date, "%Y-%m-%d")
    yesterday = (dt - timedelta(days=1)).strftime("%Y-%m-%d")
    is_monday = (dt.weekday() == 0)

    rows_map = qr.yesterday_rows if qr.yesterday_rows else qr.today_rows

    print(f"\n{'='*50}")
    print(f"  📊 昨日回收简报 ({yesterday})")
    print(f"{'='*50}")

    if is_monday:
        print(f"\n  周一，无昨日回收数据。")
        return

    FIELDS = ["在手案件", "在手金额", "新增绑定案件", "新增绑定金额",
              "总回收件数", "总回收金额"]

    for stage, accounts in STAGE_GROUPS.items():
        total_cases = 0
        total_recv  = 0
        total_amt   = 0.0

        for a in accounts:
            if a not in rows_map:
                continue
            merged = {}
            for ct_data in rows_map[a].values():
                for k in FIELDS:
                    if k in ct_data:
                        merged[k] = (merged.get(k, 0) or 0) + (ct_data.get(k, 0) or 0)

            use_new = (stage == "S1阶段")
            cases = int(merged.get("新增绑定案件" if use_new else "在手案件", 0) or 0)
            recv  = int(merged.get("总回收件数", 0) or 0)
            amt   = float(merged.get("总回收金额", 0) or 0)
            total_cases += cases
            total_recv  += recv
            total_amt   += amt

        label = "新增绑定" if stage == "S1阶段" else "在手"
        print(f"\n  {stage}")
        print(f"    {label}: {total_cases} 件")
        print(f"    回收: {total_recv} 件")
        print(f"    回收金额: {total_amt:,.2f}")


# ╔══════════════════════════════════════════════════════════╗
# ║       Google Sheet  同  步  (从 Excel 读取)               ║
# ╚══════════════════════════════════════════════════════════╝

def sync_google(target_date: str, excel_path: str, logger: Logger):
    """
    从 Excel 读取数据 → 同步到 Google Sheet。
    写入 5 个字段: C(入催数) D(案件目标) E(回收数) G(金额目标) H(回收金额)
    与 V1 完全兼容。
    """
    import gspread
    from google.oauth2.service_account import Credentials

    dt = datetime.strptime(target_date, "%Y-%m-%d")
    gs_date = dt.strftime("%d/%m/%Y")

    # ── 读 Excel ──
    helper = ExcelHelper(excel_path)
    dates = helper.detect_dates()
    if target_date not in dates:
        logger.log_error(f"Google同步: Excel中未找到日期 {target_date}")
        return

    case_col, recv_col = dates[target_date]
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws_case = wb["案件周度"]
    ws_amt  = wb["金额周度"]

    def num(v, default=0):
        return v if isinstance(v, (int, float)) else default

    data = {}
    for stage, acct, nr, or_ in ACCOUNTS:
        # 案件周度
        new_cases = num(ws_case[f"{case_col}{nr}"].value)
        old_cases = num(ws_case[f"{case_col}{or_}"].value)
        new_recv  = num(ws_case[f"{recv_col}{nr}"].value)
        old_recv  = num(ws_case[f"{recv_col}{or_}"].value)
        # 金额周度
        new_amt   = num(ws_amt[f"{recv_col}{nr}"].value)
        old_amt   = num(ws_amt[f"{recv_col}{or_}"].value)
        new_in    = num(ws_amt[f"{case_col}{nr}"].value)
        old_in    = num(ws_amt[f"{case_col}{or_}"].value)

        # 目标 = 从 Excel 数据 + 目标比率计算
        case_target, amt_target = calc_targets(
            stage, new_cases, old_cases, new_in, old_in)

        data[acct] = {
            "入催数":   new_cases + old_cases,
            "回收数":   new_recv  + old_recv,
            "回收金额": new_amt   + old_amt,
            "入催金额": new_in    + old_in,
            "案件目标": case_target,
            "金额目标": amt_target,
        }

    wb.close()

    # ── 打印 ──
    print(f"\n  本地 Excel 数据 ({target_date}):")
    print(f"  {'账户':<10} {'入催':>6} {'案件目标':>8} {'回收':>6} "
          f"{'金额目标':>10} {'回收金额':>10}")
    print(f"  {'─'*58}")
    for _, acct, _, _ in ACCOUNTS:
        d = data.get(acct, {})
        if d:
            print(f"  {acct:<10} {d['入催数']:>6} {d['案件目标']:>8} "
                  f"{d['回收数']:>6} {d['金额目标']:>10} {d['回收金额']:>10.2f}")

    # ── 连接 Google ──
    try:
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(GOOGLE_CREDENTIALS, scopes=scope)
        client = gspread.authorize(creds)
        sh = client.open_by_key(GOOGLE_SHEET_KEY)
        print(f"  [OK] Google Sheet 连接成功")
    except Exception as e:
        msg = f"Google 连接失败: {e}"
        print(f"  [ERR] {msg}")
        logger.log_error(msg)
        return

    all_sheets = {s.title for s in sh.worksheets()
                  if s.title not in GS_SKIP_SHEETS}
    matched = [acct for _, acct, _, _ in ACCOUNTS if acct in all_sheets]

    synced_count = 0
    for acct in matched:
        d = data.get(acct, {})
        if not d:
            continue
        try:
            ws = sh.worksheet(acct)
        except Exception:
            print(f"  {acct}: Sheet 不存在，跳过")
            continue

        all_dates = ws.col_values(1)
        try:
            row_idx = all_dates.index(gs_date) + 1
        except ValueError:
            print(f"  {acct}: 未找到日期 {gs_date}，跳过")
            continue

        updates = [
            {"range": f"C{row_idx}", "values": [[d["入催数"]]]},
            {"range": f"D{row_idx}", "values": [[d["案件目标"]]]},
            {"range": f"E{row_idx}", "values": [[d["回收数"]]]},
            {"range": f"G{row_idx}", "values": [[int(d["金额目标"])]]},
            {"range": f"H{row_idx}", "values": [[round(d["回收金额"], 2)]]},
        ]
        ws.batch_update(updates, value_input_option='USER_ENTERED')
        synced_count += 1
        print(f"  {acct}: 入催={d['入催数']} 案件目标={d['案件目标']} "
              f"回收={d['回收数']} 金额目标={int(d['金额目标'])} "
              f"回收金额={d['回收金额']:.2f}")

    result_msg = f"完成 ({synced_count}/{len(matched)} 账户)"
    print(f"  [OK] Google Sheet 同步{result_msg}")
    logger.log_sync(result_msg)


# ╔══════════════════════════════════════════════════════════╗
# ║         主  流  程                                        ║
# ╚══════════════════════════════════════════════════════════╝

def _print_end_banner(log: Logger):
    """打印程序结束横幅"""
    total = log.total()
    print(f"\n{'='*40}")
    print(f"  日报: {log.excel_status or 'N/A'}")
    print(f"  Google: {log.google_status or 'N/A'}")
    if log.query_cache_info:
        print(f"  Redash:")
        for label in ["今天新客", "今天老客", "昨天新客", "昨天老客"]:
            cache = log.query_cache_info.get(label)
            if cache:
                source_icon = "📦" if cache["source"] == "Cache" else "🔄"
                print(f"    {label}: {source_icon} {cache['source']}  "
                      f"({cache['query_result_id']})")
    print(f"  总耗时: {total:.1f} 秒")
    print(f"  Version: {VERSION}")
    print(f"{'='*40}", flush=True)


def main():
    print("main开始", flush=True)
    log = Logger()
    sync_only  = "--sync-only"  in sys.argv
    dry_run    = "--dry-run"    in sys.argv
    query_only = "--query-only" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]

    target_date = date.today().strftime("%Y-%m-%d")
    if args:
        target_date = args[0]

    dt = datetime.strptime(target_date, "%Y-%m-%d")
    weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    day_name = weekdays[dt.weekday()]

    print(f"\n{'='*50}")
    print(f"  xby日报 V2  v{VERSION}  Build {BUILD}")
    print(f"  日期: {target_date} ({day_name})")
    if query_only:
        print(f"  模式: QUERY-ONLY (仅查询+保存JSON)")
    elif dry_run:
        print(f"  模式: DRY-RUN (仅查询+简报，不写Excel/Google)")
    elif sync_only:
        print(f"  模式: 仅同步 Google")
    print(f"{'='*50}", flush=True)

    print(f"\n[程序启动] {datetime.now().strftime('%H:%M:%S')}", flush=True)

    # ── 仅同步模式 ──
    if sync_only:
        log.excel_status = "跳过 (sync-only)"
        log.google_status = "已同步"
        print("[步骤] 开始 Google 同步 ...", flush=True)
        log.start("Google同步")
        try:
            sync_google(target_date, EXCEL_PATH, log)
        except Exception as e:
            log.google_status = "同步失败"
            log.log_error(f"Google 同步失败: {e}")
        log.stop("Google同步")
        _print_end_banner(log)
        log.print_summary()
        return

    # ── 仅查询模式 ──
    if query_only:
        print("\n[步骤] 开始登录 Redash ...", flush=True)
        log.start("Redash登录")
        try:
            redash = RedashClient(log)
        except Exception as e:
            print(f"\n[ERR] 登录失败: {e}", flush=True)
            log.log_error(f"登录失败: {e}")
            log.print_summary()
            log.save(target_date)
            return
        log.stop("Redash登录")
        print(f"[OK] Redash 登录成功 ({log.elapsed('Redash登录'):.1f}s)\n", flush=True)

        print("[步骤] 开始查询 Redash ...", flush=True)
        log.start("Redash查询总计")
        try:
            qr = QueryResult.from_redash(redash, target_date, log)
        except Exception as e:
            print(f"\n[ERR] 查询失败: {e}", flush=True)
            log.log_error(f"查询失败: {e}")
            redash.close()
            log.print_summary()
            log.save(target_date)
            return
        redash.close()
        log.stop("Redash查询总计")
        print(f"\n[OK] 查询完成 ({log.elapsed('Redash查询总计'):.1f}s)", flush=True)

        # 打印统计
        print(f"\n{'─'*40}")
        print("Query 统计:")
        for label in ["今天新客", "今天老客", "昨天新客", "昨天老客"]:
            count = log.query_counts.get(label, "跳过")
            print(f"  {label}: {count} 行")
        print(f"  总耗时: {log.total():.1f}s")
        print(f"{'─'*40}")

        # 保存原始 JSON
        json_path = EXCEL_PATH.replace(".xlsx", f"_query_{target_date}.json")
        import json
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({
                "target_date": target_date,
                "version": VERSION,
                "today_rows": {a: {c: r for c, r in d.items()}
                               for a, d in qr.today_rows.items()},
                "yesterday_rows": {a: {c: r for c, r in d.items()}
                                   for a, d in qr.yesterday_rows.items()},
            }, f, ensure_ascii=False, indent=2, default=str)
        print(f"\n📝 原始数据已保存: {os.path.basename(json_path)}", flush=True)

        log.excel_status = "跳过 (query-only)"
        log.google_status = "跳过 (query-only)"
        _print_end_banner(log)
        log.print_summary()
        log.save(target_date)
        return

    # ── 1. 周一处理 ──
    print("[步骤] 检查周一 / Excel 日期 ...", flush=True)
    excel = ExcelHelper()
    log.start("周一检查")
    excel_path = excel.handle_monday(target_date, log)
    log.stop("周一检查")

    # ── 2. 检查日期 ──
    if not excel.has_date(target_date):
        print(f"\n[!] Excel 中未找到日期 {target_date}")
        print(f"    可用日期: {list(excel.detect_dates().keys())}")
        return

    # ── 3. Redash 查询 ──
    print(f"\n[步骤] 开始登录 Redash ...", flush=True)
    log.start("Redash登录")
    try:
        redash = RedashClient(log)
    except Exception as e:
        msg = f"Redash 登录失败: {e}"
        print(f"\n[ERR] {msg}", flush=True)
        log.log_error(msg)
        log.print_summary()
        return
    log.stop("Redash登录")
    print(f"[OK] Redash 登录成功 ({log.elapsed('Redash登录'):.1f}s)", flush=True)

    print(f"\n[步骤] 开始查询 Redash (Query #{QUERY_ID}) ...", flush=True)
    log.start("Redash查询总计")
    try:
        qr = QueryResult.from_redash(redash, target_date, log)
    except Exception as e:
        msg = f"Redash 查询失败: {e}"
        print(f"\n[ERR] {msg}", flush=True)
        log.log_error(msg)
        redash.close()
        log.print_summary()
        log.save(target_date)
        return
    redash.close()
    log.stop("Redash查询总计")
    print(f"\n[OK] Redash 查询完成 ({log.elapsed('Redash查询总计'):.1f}s)", flush=True)

    # ── 4. 数据校验 ──
    print(f"[步骤] 数据校验 ...", flush=True)
    if qr.is_empty():
        msg = "所有账户案件数均为 0，已阻止写入 Excel 和同步 Google"
        print(f"\n{'!'*50}", flush=True)
        print(f"  ⚠ 数据异常：{msg}", flush=True)
        print(f"  请检查 Redash 查询 #{QUERY_ID}", flush=True)
        print(f"{'!'*50}", flush=True)
        log.log_error(msg)
        log.print_summary()
        log.save(target_date)
        return

    # ── 5. 新账户检测 ──
    new_accts = qr.find_new_accounts()
    if new_accts:
        log.log_new_accounts(new_accts)
        print(f"\n{'─'*40}")
        print(f"  🔔 发现新账户: {', '.join(new_accts)}")
        try:
            print(f"  是否加入配置？(Y/N): ", end="", flush=True)
            answer = input().strip().upper()
        except (EOFError, OSError):
            answer = "N"
            print("(自动跳过)", flush=True)
        if answer == "Y":
            print(f"  请在 ACCOUNTS 列表中添加上述账户后重新运行。")
            print(f"  格式: (\"阶段\", \"账户名\", 新客行, 老客行)")
            log.save(target_date)
            return
        else:
            print(f"  已忽略新账户，继续执行...", flush=True)

    # ── 6. 简报 ──
    print(f"\n[步骤] 开始生成简报 ...", flush=True)
    print_briefing(qr, target_date)

    # ── DRY-RUN 模式在此结束 ──
    if dry_run:
        log.excel_status = "跳过 (dry-run)"
        log.google_status = "跳过 (dry-run)"
        print(f"\n{'─'*40}", flush=True)
        print(f"  DRY-RUN 完成（未写入 Excel / 未同步 Google）", flush=True)
        _print_end_banner(log)
        log.print_summary()
        log.save(target_date)
        return

    # ── 7. 写 Excel ──
    log.excel_status = "✅ 成功"
    log.start("Excel写入")
    try:
        excel.excel_path = excel_path
        if ENABLE_BACKUP:
            excel.backup()
        excel.write_today(qr, target_date)
    except Exception as e:
        msg = f"Excel 写入失败: {e}"
        print(f"[ERR] {msg}")
        log.log_error(msg)
        log.excel_status = "❌ 失败"
    log.stop("Excel写入")

    # ── 8. 用户确认 → Google 同步 ──
    print(f"\n{'─'*40}")
    if AUTO_SYNC_GOOGLE:
        answer = "Y"
        print(f"  AUTO_SYNC_GOOGLE=True，自动同步...")
    else:
        print(f"  请检查 Excel 数据。")
        try:
            answer = input(f"  是否同步 Google Sheet？(Y/N): ").strip().upper()
        except (EOFError, OSError):
            answer = "N"
            print("(自动跳过)", flush=True)

    if answer == "Y":
        log.google_status = "已同步"
        log.start("Google同步")
        try:
            sync_google(target_date, excel_path, log)
        except Exception as e:
            msg = f"Google 同步异常: {e}"
            print(f"[ERR] {msg}")
            log.log_error(msg)
            log.log_sync("异常终止")
            log.google_status = "同步失败"
        log.stop("Google同步")
    else:
        print(f"  已跳过 Google 同步。")
        log.log_sync("用户跳过")
        log.google_status = "未同步"

    # ── 9. 结束 ──
    _print_end_banner(log)
    log.print_summary()
    log.save(target_date)


# ╔══════════════════════════════════════════════════════════╗
# ║         测  试  函  数                                    ║
# ╚══════════════════════════════════════════════════════════╝

def test_login():
    """测试 Redash 登录"""
    print("开始登录", flush=True)
    log = Logger()
    client = RedashClient(log)
    print("登录成功", flush=True)
    client.close()


def test_query():
    """测试 Redash 单次查询"""
    print("开始登录", flush=True)
    log = Logger()
    client = RedashClient(log)
    print("登录成功", flush=True)

    from datetime import date, timedelta
    today = date.today().strftime("%Y-%m-%d")
    tomorrow = (date.today() + timedelta(days=1)).strftime("%Y-%m-%d")

    print(f"开始查询今天新客 ({today} ~ {tomorrow}) ...", flush=True)
    rows, cache_info = client.query(today, tomorrow, "新客")
    print(f"查询完成 → {len(rows)} 行  "
          f"query_result_id={cache_info['query_result_id']}  "
          f"来源={cache_info['source']}", flush=True)

    if rows:
        print(f"第一行 keys: {list(rows[0].keys())}", flush=True)
        print(f"第一行: {rows[0]}", flush=True)

    client.close()
    print("测试结束", flush=True)


print("进入main前", flush=True)
if __name__ == "__main__":
    main()
