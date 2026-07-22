# -*- coding: utf-8 -*-
"""
WW 智能绩效与人员调配系统 — Excel 输出

每日 sheet 写入 + 绩效汇总重建 + 输出生成。
"""

import os

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import config


# ============================================================
# 样式常量
# ============================================================

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# 日 Sheet 表头
DAY_HEADERS = ["姓名", "队列", "阶段", "金额催回率", "系数", "绩效"]


# ============================================================
# 样式辅助
# ============================================================

def _style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_data_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws, col_count):
    for c in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            length = 0
            for ch in val:
                if "一" <= ch <= "鿿" or "　" <= ch <= "〿":
                    length += 2
                else:
                    length += 1
            max_len = max(max_len, length)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 30)


# ============================================================
# 单日 Sheet
# ============================================================

def write_day_sheet(wb, day_str, day_results):
    """写入单日 Sheet（休息排最后，其余按催回率降序）"""
    ws = wb.create_sheet(title=day_str)

    for c, h in enumerate(DAY_HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(DAY_HEADERS))

    # 休息排最后，其余按催回率降序
    active = [r for r in day_results if not r.is_rest]
    resting = [r for r in day_results if r.is_rest]
    sorted_results = sorted(active, key=lambda r: r.rate, reverse=True) + resting

    for i, r in enumerate(sorted_results):
        row = i + 2
        ws.cell(row=row, column=1, value=r.name)
        ws.cell(row=row, column=2, value=r.queue)
        ws.cell(row=row, column=3, value=r.stage)
        ws.cell(row=row, column=4, value=round(r.rate, config.RATE_DISPLAY_DECIMALS) if not r.is_rest else "-")
        ws.cell(row=row, column=5, value=r.coefficient if not r.is_rest else "-")
        ws.cell(row=row, column=6, value=r.amount)
        _style_data_row(ws, row, len(DAY_HEADERS))

    _auto_width(ws, len(DAY_HEADERS))


# ============================================================
# 绩效汇总
# ============================================================

def rebuild_summary(wb):
    """扫描所有日 Sheet，完全重建「绩效汇总」Sheet"""
    all_day_strs = sorted(
        [s for s in wb.sheetnames if s.isdigit()],
        key=int
    )

    # 从每个日 Sheet 读取 {name: amount}
    day_perf = {}
    for day_str in all_day_strs:
        ws = wb[day_str]
        day_perf[day_str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            amount = row[5]
            if name:
                if amount == "休":
                    day_perf[day_str][name] = "休"
                else:
                    day_perf[day_str][name] = amount if amount else 0

    # 收集所有姓名（保持首次出现顺序）
    all_names = []
    seen = set()
    for day_str in all_day_strs:
        for name in day_perf[day_str]:
            if name not in seen:
                seen.add(name)
                all_names.append(name)

    # 删除旧"绩效汇总"
    if "绩效汇总" in wb.sheetnames:
        del wb["绩效汇总"]

    # 创建新"绩效汇总"（放到最前面）
    day_labels = [f"{d}号" for d in all_day_strs]
    headers = ["姓名"] + day_labels + ["累计"]
    ws = wb.create_sheet(title="绩效汇总", index=0)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for i, name in enumerate(all_names):
        row = i + 2
        ws.cell(row=row, column=1, value=name)
        total = 0
        for j, day_str in enumerate(all_day_strs):
            amount = day_perf[day_str].get(name, 0)
            if amount == "休":
                ws.cell(row=row, column=2 + j, value="休")
            else:
                val = amount if amount else 0
                ws.cell(row=row, column=2 + j, value=val)
                total += val
        ws.cell(row=row, column=2 + len(all_day_strs), value=total)
        _style_data_row(ws, row, len(headers))

    _auto_width(ws, len(headers))


# ============================================================
# 输出生成（入口）
# ============================================================

def generate_output(all_day_results, output_path):
    """生成或增量更新绩效汇总 Excel。

    日 Sheet 已存在则跳过不覆盖，最后统一重建绩效汇总。
    """
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        if "绩效汇总" in wb.sheetnames:
            del wb["绩效汇总"]
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    skipped = []
    for data_date, day_results, _scheme_name in all_day_results:
        day_str = str(data_date.day)
        if day_str in wb.sheetnames:
            skipped.append(day_str)
            continue
        write_day_sheet(wb, day_str, day_results)

    if skipped:
        print(f"  跳过已有日期（未覆盖）: {', '.join(f'{d}号' for d in skipped)}")

    rebuild_summary(wb)
    wb.save(output_path)
