# -*- coding: utf-8 -*-
"""
WW 智能绩效与人员调配系统 — 日报读取

读取桌面日报 Excel 文件，输出结构化记录。
"""

import os
import openpyxl

from utils import clean_name, parse_value, infer_date
from models import StaffRecord


def read_daily_report(filepath):
    """读取单个日报文件。

    返回：
        (data_date, records, warnings)
        - data_date: date
        - records: list[StaffRecord]
        - warnings: list[str]
    """
    filename = os.path.basename(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sheet_name = ws.title

    data_date = infer_date(filename, sheet_name)

    records = []
    warnings = []

    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        # 跳过汇总行
        if row[0] and "汇总" in str(row[0]):
            continue

        name_raw = row[0]       # A列：姓名
        staff_id = row[1]       # B列：工号
        dept = row[2]           # C列：部门
        queue = row[3]          # D列：队列
        coverage_raw = row[4]   # E列：覆盖率

        if not name_raw or not str(name_raw).strip():
            continue

        name = clean_name(name_raw)
        coverage = parse_value(coverage_raw)

        # N列（index 13）：债务汇总维度·金额催回率
        amount_rate = parse_value(row[13])

        # L列（index 11）：债务汇总维度·回款总额
        repaid_amount = parse_value(row[11])

        # T列（index 19）：队列新案维度·应催总额
        due_amount = parse_value(row[19])

        # O列（index 14）：队列新案维度·分案户数，为 0 = 休息
        rest_cases = parse_value(row[14])
        is_rest = (rest_cases == 0)

        if is_rest:
            warnings.append(f"  [!] {name}({staff_id}): O列分案户数为0，当天休息")
        elif coverage == 0:
            warnings.append(f"  [!] {name}({staff_id}): 覆盖率 0%，当天未出勤")

        records.append(StaffRecord(
            name=name,
            staff_id=str(staff_id).strip() if staff_id else "",
            dept=str(dept).strip() if dept else "",
            queue=str(queue).strip() if queue else "",
            coverage=coverage,
            amount_rate=amount_rate,
            repaid_amount=repaid_amount,
            due_amount=due_amount,
            is_rest=is_rest,
        ))

    wb.close()
    return data_date, records, warnings
