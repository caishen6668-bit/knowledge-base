# -*- coding: utf-8 -*-
"""
WW 智能绩效与人员调配系统 — 管理驾驶舱

在输出 Excel 中新增「管理驾驶舱」Sheet，提供：
  1. 今日运营概况
  2. 人员能力排行榜 Top10
  3. 今日调岗建议
  4. 能力预警
  5. 调岗统计

纯展示层，不参与任何计算。
"""

from datetime import date

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import config
from utils import get_recommendation_level, RECOMMENDATION_LEVELS


# ============================================================
# 样式常量
# ============================================================

_TITLE_FONT = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
_SUBTITLE_FONT = Font(name="微软雅黑", size=11, color="333333")
_SECTION_FONT = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
_SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
_DATA_FONT = Font(name="微软雅黑", size=10)
_LABEL_FONT = Font(name="微软雅黑", bold=True, size=10, color="333333")
_VALUE_FONT = Font(name="微软雅黑", size=10, color="1F4E79")
_WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_RED_FONT = Font(name="微软雅黑", size=10, color="C00000")
_GREEN_FONT = Font(name="微软雅黑", size=10, color="375623")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")

# Dashboard 总列数
_COL_COUNT = 7

# 等级颜色映射
_LEVEL_COLORS = {
    "S": "375623",   # 深绿
    "A": "548235",   # 绿
    "B": "1F4E79",   # 蓝
    "C": "BF8F00",   # 橙
    "D": "C00000",   # 红
}

# 等级背景色
_LEVEL_FILLS = {
    "S": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # 绿底
    "A": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),  # 浅绿底
    "B": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # 蓝底
    "C": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # 橙底
    "D": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),  # 红底
}


# ============================================================
# 辅助函数
# ============================================================

def _apply_cell(ws, row, col, value, font=_DATA_FONT, fill=None,
                alignment=_CENTER, number_format=None):
    """写单元格并应用样式。"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = _THIN_BORDER
    if number_format:
        cell.number_format = number_format
    return cell


def _write_section_header(ws, row, title):
    """写 section 标题行（蓝底白字，跨全部列）。"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_COL_COUNT)
    _apply_cell(ws, row, 1, title, font=_SECTION_FONT, fill=_SECTION_FILL)
    for c in range(2, _COL_COUNT + 1):
        _apply_cell(ws, row, c, None, font=_SECTION_FONT, fill=_SECTION_FILL)


def _write_table_header(ws, row, headers):
    """写表格头行。"""
    for c, h in enumerate(headers, 1):
        _apply_cell(ws, row, c, h, font=_HEADER_FONT, fill=_HEADER_FILL)


def _write_data_row(ws, row, values, warn=False):
    """写一行数据，可选预警底色。"""
    fill = _WARN_FILL if warn else None
    for c, v in enumerate(values, 1):
        _apply_cell(ws, row, c, v, fill=fill)


def _auto_width(ws):
    """自动调整列宽。"""
    for c in range(1, _COL_COUNT + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            length = 0
            for ch in val:
                if "一" <= ch <= "鿿" or "　" <= ch <= "〿":
                    length += 2
                else:
                    length += 1
            max_len = max(max_len, length)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 32)


def _build_ranking(ability_ranking):
    """从多阶段能力数据中，每人取 OverallScore 最高的阶段，按 OverallScore 降序排列。

    返回：list[dict]，每个 dict 有 staff_id, name, stage, overall_score, level。
    """
    best = {}
    for a in ability_ranking:
        sid = a["staff_id"]
        score = a.get("overall_score", 0)
        if sid not in best or score > best[sid]["overall_score"]:
            best[sid] = {
                "staff_id": sid,
                "name": a["name"],
                "stage": a["stage"],
                "overall_score": score,
                "level": get_recommendation_level(score),
            }
    # 按 OverallScore 降序
    return sorted(best.values(), key=lambda x: x["overall_score"], reverse=True)


# ============================================================
# 各 Section 写入函数
# ============================================================

def _write_overview(ws, row, data_date, last_records, demand, opt_result):
    """Section 1: 今日运营概况"""
    _write_section_header(ws, row, "  [1] 今日运营概况")
    row += 1

    # 统计出勤/休息
    attendance = sum(1 for r in last_records if not r.is_rest)
    rest_count = sum(1 for r in last_records if r.is_rest)

    # 调岗人数
    transfer_count = len(opt_result.transfers) if opt_result else 0
    # 预警人数
    warning_count = len(opt_result.warnings) if opt_result else 0
    # 未满足需求
    if opt_result and opt_result.unmet_demand:
        unmet_text = "；".join(
            f"{u.queue}:缺{u.shortfall}人" for u in opt_result.unmet_demand
        )
    else:
        unmet_text = "全部满足"

    # 需求文本
    demand_items = []
    for q in sorted(demand.keys()):
        demand_items.append(f"{q}:{demand[q]}人")
    demand_text = "  ".join(demand_items)

    # 4 对标签-值，每行 4 对（8 列），但我们只用 7 列
    # 第一行：4 个指标
    pairs_row1 = [
        ("出勤人数", str(attendance)),
        ("休息人数", str(rest_count)),
        ("今日需求", demand_text),
    ]
    for j, (label, val) in enumerate(pairs_row1):
        col_label = j * 2 + 1
        col_val = j * 2 + 2
        _apply_cell(ws, row, col_label, label, font=_LABEL_FONT)
        _apply_cell(ws, row, col_val, val, font=_VALUE_FONT)
        if col_val < _COL_COUNT:
            # 清除可能遗留的合并单元格边框
            pass

    row += 1

    # 第二行：3 个指标
    pairs_row2 = [
        ("推荐调岗", f"{transfer_count}人"),
        ("能力预警", f"{warning_count}人"),
        ("未满足需求", unmet_text),
    ]
    for j, (label, val) in enumerate(pairs_row2):
        col_label = j * 2 + 1
        col_val = j * 2 + 2
        _apply_cell(ws, row, col_label, label, font=_LABEL_FONT)
        font = _RED_FONT if (j == 2 and opt_result and opt_result.unmet_demand) else _VALUE_FONT
        _apply_cell(ws, row, col_val, val, font=font)

    return row + 1  # 返回下一可用行（留一个空行）


def _write_ranking(ws, row, ability_ranking):
    """Section 2: 人员能力排行榜 Top10"""
    _write_section_header(ws, row, "  [2] 人员能力排行榜 Top10")
    row += 1

    headers = ["排名", "姓名", "当前阶段", "OverallScore", "推荐等级"]
    _write_table_header(ws, row, headers)
    row += 1

    # 每人取最佳阶段
    ranking = _build_ranking(ability_ranking)
    top10 = ranking[:10]

    if not top10:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_COL_COUNT)
        _apply_cell(ws, row, 1, "暂无能力数据", font=_DATA_FONT, alignment=_CENTER)
        return row + 2

    for i, person in enumerate(top10):
        rank = i + 1
        level = person["level"]
        level_text = f"{level} ({RECOMMENDATION_LEVELS.get(level, '')})"
        values = [
            rank,
            person["name"],
            person["stage"],
            round(person["overall_score"], 1),
            level_text,
        ]
        # 整行使用等级对应的背景色
        row_fill = _LEVEL_FILLS.get(level)
        for c, v in enumerate(values, 1):
            font = _DATA_FONT
            if c == 5:  # 推荐等级列用对应颜色
                font = Font(name="微软雅黑", bold=True, size=10,
                           color=_LEVEL_COLORS.get(level, "000000"))
            _apply_cell(ws, row, c, v, font=font, fill=row_fill)
        row += 1

    return row + 1


def _write_transfers(ws, row, opt_result):
    """Section 3: 今日调岗建议"""
    _write_section_header(ws, row, "  [3] 今日调岗建议")
    row += 1

    transfers = opt_result.transfers if opt_result else []

    if not transfers:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_COL_COUNT)
        _apply_cell(ws, row, 1, "今日无需调岗，建议保持昨日配置。",
                    font=_GREEN_FONT, fill=_GREEN_FILL, alignment=_CENTER)
        for c in range(2, _COL_COUNT + 1):
            _apply_cell(ws, row, c, None, fill=_GREEN_FILL)
        return row + 2

    headers = ["姓名", "昨日阶段", "今日阶段", "综合收益", "调岗原因"]
    _write_table_header(ws, row, headers)
    row += 1

    for t in transfers:
        gain_str = f"+{t.composite_gain:.1f}" if t.composite_gain >= 0 else f"{t.composite_gain:.1f}"
        font = _GREEN_FONT if t.composite_gain >= config.OPTIMIZER_TRANSFER_THRESHOLD else _DATA_FONT
        values = [t.name, t.from_queue, t.to_queue, gain_str, t.reason]
        for c, v in enumerate(values, 1):
            f = font if c == 4 else _DATA_FONT
            _apply_cell(ws, row, c, v, font=f)
        row += 1

    return row + 1


def _write_warnings(ws, row, opt_result):
    """Section 4: 能力预警"""
    _write_section_header(ws, row, "  [4] 能力预警")
    row += 1

    warnings = opt_result.warnings if opt_result else []

    if not warnings:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_COL_COUNT)
        _apply_cell(ws, row, 1, "无能力预警",
                    font=_GREEN_FONT, fill=_GREEN_FILL, alignment=_CENTER)
        for c in range(2, _COL_COUNT + 1):
            _apply_cell(ws, row, c, None, fill=_GREEN_FILL)
        return row + 2

    headers = ["姓名", "阶段", "OverallScore", "预警原因"]
    _write_table_header(ws, row, headers)
    row += 1

    for w in warnings:
        reason = f"综合能力 {w.composite_score:.1f} < {config.OPTIMIZER_WARNING_THRESHOLD}，建议关注"
        values = [w.name, w.queue, round(w.composite_score, 1), reason]
        _write_data_row(ws, row, values, warn=True)
        row += 1

    return row + 1


def _write_stats(ws, row, opt_result):
    """Section 5: 调岗统计"""
    _write_section_header(ws, row, "  [5] 调岗统计")
    row += 1

    if not opt_result:
        _apply_cell(ws, row, 1, "暂无数据", font=_DATA_FONT)
        return row + 2

    s = opt_result.summary
    transfers = opt_result.transfers if opt_result.transfers else []

    avg_gain = 0.0
    if transfers:
        avg_gain = sum(t.composite_gain for t in transfers) / len(transfers)

    stats = [
        ("今日调岗人数", f"{s.get('total_transfers', 0)}人"),
        ("保持岗位人数", f"{s.get('staying_count', 0)}人"),
        ("岗位保持率", f"{s.get('stability_pct', 0):.0f}%"),
        ("平均调岗收益", f"{avg_gain:+.1f}"),
        ("综合总分", f"{s.get('total_composite', 0):.1f}"),
        ("调岗阈值", f">= {s.get('transfer_threshold', 8)}"),
    ]

    # 分为两行，每行 3 对
    for i, (label, val) in enumerate(stats):
        r = row + (i // 3)
        col_label = (i % 3) * 2 + 1
        col_val = col_label + 1
        _apply_cell(ws, r, col_label, label, font=_LABEL_FONT)
        _apply_cell(ws, r, col_val, val, font=_VALUE_FONT)

    return row + 2 + 1


# ============================================================
# 入口：添加管理驾驶舱 Sheet
# ============================================================

def add_dashboard_sheet(wb, data_date, last_records, ability_ranking, opt_result, demand):
    """在 workbook 中新增「管理驾驶舱」Sheet（插入到最前面）。

    参数：
        wb:               openpyxl Workbook（已加载，含现有 sheet）
        data_date:        date — 报表日期
        last_records:     list[StaffRecord] — 最后一日的原始员工记录
        ability_ranking:  list[dict] — ability.get_all_abilities() 结果，按 overall_score 降序
        opt_result:       OptimizationResult — optimizer.optimize() 结果
        demand:           dict[str, int] — 各队列需求人数
    """
    # 如果已存在，先删除再重建
    if "管理驾驶舱" in wb.sheetnames:
        del wb["管理驾驶舱"]

    ws = wb.create_sheet(title="管理驾驶舱", index=0)

    # --- 标题 ---
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_COL_COUNT)
    _apply_cell(ws, row, 1, "WW 智能绩效与人员调配系统 — 管理驾驶舱",
                font=_TITLE_FONT, alignment=_CENTER)

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_COL_COUNT)
    date_str = data_date.strftime("%Y-%m-%d") if isinstance(data_date, date) else str(data_date)
    _apply_cell(ws, row, 1, f"报表日期：{date_str}",
                font=_SUBTITLE_FONT, alignment=_CENTER)

    # --- 各 Section（每个 section 留一空行） ---
    row = 4
    row = _write_overview(ws, row, data_date, last_records, demand, opt_result)
    row = _write_ranking(ws, row, ability_ranking)
    row = _write_transfers(ws, row, opt_result)
    row = _write_warnings(ws, row, opt_result)
    _write_stats(ws, row, opt_result)

    # --- 列宽 ---
    _auto_width(ws)
