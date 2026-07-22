# -*- coding: utf-8 -*-
"""
WW 智能绩效 — 排班 API Debug 模式

纯数据导出，不参与任何业务计算。
  1. 调用 staffSchedulePage 获取 2026-07 整月排班
  2. 将 response.json() 原样保存 → output/schedule_raw.json
  3. 展开每天每人一行 → output/schedule_parsed.xlsx

用法:
    python debug_schedule.py
"""

import json
import os
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ============================================================
# 配置
# ============================================================

_BASE_URL = os.environ.get(
    "COLLECT_BASE_URL",
    "https://loan-collect.maxicredito.loan/vitech-collect-gateway",
)
_SCHEDULE_URL = f"{_BASE_URL}/collect/staff/staffSchedulePage"

_TOKEN = os.environ.get(
    "MEX_COLLECT_TOKEN",
    "",
)

_REQUEST_TIMEOUT = 30
_API_TZ = timezone(timedelta(hours=-6))  # UTC-6 墨西哥中部

# 输出目录
_OUTPUT_DIR = Path(__file__).parent / "output"

# scheduleStatus → 显示文本
_STATUS_MAP = {
    1:  "上班",
    0:  "休息",
    -1: "请假",
    2:  "未知(2)",
    3:  "未知(3)",
}


# ============================================================
# API 请求
# ============================================================

def fetch_raw(year: int, month: int) -> dict:
    """调用 staffSchedulePage，返回完整 response.json()。"""

    start_dt = datetime(year, month, 1, tzinfo=_API_TZ)
    if month == 12:
        end_dt = datetime(year + 1, 1, 1, tzinfo=_API_TZ)
    else:
        end_dt = datetime(year, month + 1, 1, tzinfo=_API_TZ)

    now = datetime.now(_API_TZ)

    body = {
        "region": "",
        "date": now.strftime("%Y-%m-%dT%H:%M:%S"),
        "scheduleStart": int(start_dt.timestamp() * 1000),
        "scheduleEnd": int(end_dt.timestamp() * 1000),
        "page": 1,
        "rows": 500,
        "staffStatus": "1",
        "secondLevelDeptId": "",
        "thirdLevelDeptId": "",
        "fourthLevelDeptId": "",
        "staffName": "",
        "overdueLevelIdList": [],
        "scheduleStatusList": [],
    }

    headers = {
        "Content-Type": "application/json",
        "X-CHOICE-TAG": "ms",
        "X-END-LANGUAGE": "zh_cn",
        "X-Fixed-Token": _TOKEN,
    }

    print(f"[1/4] 请求 API ...")
    print(f"      URL: {_SCHEDULE_URL}")
    print(f"      月份: {year}-{month:02d}")
    print(f"      时间范围: {start_dt.strftime('%Y-%m-%d')} ~ {end_dt.strftime('%Y-%m-%d')}")

    resp = requests.post(_SCHEDULE_URL, headers=headers, json=body, timeout=_REQUEST_TIMEOUT)
    resp.raise_for_status()

    data = resp.json()

    if data.get("resultCode") != 1:
        print(f"      [!] API 返回异常: resultCode={data.get('resultCode')}, "
              f"message={data.get('message')}")
        return data

    rows = data.get("data", {}).get("rows", [])
    total_page = data.get("data", {}).get("totalPage", 1)
    total_record = data.get("data", {}).get("totalRecord", len(rows))

    print(f"      员工数: {len(rows)}")
    print(f"      总页数: {total_page}")

    # 多页
    if total_page > 1:
        all_rows = list(rows)
        for p in range(2, total_page + 1):
            body["page"] = p
            resp2 = requests.post(_SCHEDULE_URL, headers=headers, json=body, timeout=_REQUEST_TIMEOUT)
            page_data = resp2.json()
            all_rows.extend(page_data.get("data", {}).get("rows", []))
            print(f"      第 {p}/{total_page} 页: {len(page_data.get('data', {}).get('rows', []))} 人")
        # 回写合并后的 rows
        data["data"]["rows"] = all_rows
        data["data"]["totalRecord"] = len(all_rows)

    return data


# ============================================================
# 展开解析
# ============================================================

def expand_rows(raw_data: dict, year: int, month: int) -> list[dict]:
    """将 API 原始 rows 展开为每天一行的列表。"""
    rows = raw_data.get("data", {}).get("rows", [])

    entries = []
    for row in rows:
        staff_id = row.get("staffId", "")
        staff_no = row.get("staffNo", "")
        staff_name = row.get("staffName", "")
        staff_status = row.get("staffStatus", "")
        staff_position = row.get("staffPosition", "")
        overdue_level_id = row.get("overdueLevelId", "")
        third_dept_name = row.get("thirdLevelDeptName", "")

        schedules = row.get("scheduleRespDtos", [])
        if not schedules:
            # 无排班数据也保留一条
            entries.append({
                "staffId": staff_id,
                "staffNo": staff_no,
                "staffName": staff_name,
                "staffStatus": staff_status,
                "staffPosition": staff_position,
                "overdueLevelId": overdue_level_id,
                "thirdLevelDeptName": third_dept_name,
                "scheduleTime": "",
                "scheduleTimeDate": "",
                "scheduleStatus": "",
                "scheduleStatusText": "无排班数据",
            })
            continue

        for sched in schedules:
            ts_raw = sched.get("scheduleTime", 0)
            # scheduleTime 可能是 int/float 毫秒时间戳，也可能是字符串
            try:
                ts_ms = int(ts_raw)
            except (ValueError, TypeError):
                ts_ms = 0

            if ts_ms and ts_ms > 0:
                try:
                    sched_dt = datetime.fromtimestamp(ts_ms / 1000, tz=_API_TZ)
                    date_str = sched_dt.strftime("%Y-%m-%d")
                except (ValueError, OSError):
                    date_str = str(ts_raw)
            else:
                date_str = str(ts_raw) if ts_raw else ""

            status_code = sched.get("scheduleStatus", "")
            status_text = _STATUS_MAP.get(status_code, str(status_code))

            entries.append({
                "staffId": staff_id,
                "staffNo": staff_no,
                "staffName": staff_name,
                "staffStatus": staff_status,
                "staffPosition": staff_position,
                "overdueLevelId": overdue_level_id,
                "thirdLevelDeptName": third_dept_name,
                "scheduleTime": ts_ms,
                "scheduleTimeDate": date_str,
                "scheduleStatus": status_code,
                "scheduleStatusText": status_text,
            })

    return entries


# ============================================================
# Excel 输出
# ============================================================

def write_excel(entries: list[dict], output_path: Path):
    """将展开后的排班数据写入 Excel。"""
    print(f"[3/4] 生成 Excel ...")
    print(f"      总行数: {len(entries)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "排班明细"

    # --- 样式 ---
    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(name="微软雅黑", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    work_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    rest_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    leave_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # --- 表头 ---
    headers = [
        "staffId", "staffNo", "staffName", "staffStatus",
        "staffPosition", "overdueLevelId", "thirdLevelDeptName",
        "scheduleTime", "scheduleTimeDate", "scheduleStatus", "scheduleStatusText",
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # --- 数据行 ---
    for i, entry in enumerate(entries):
        row = i + 2
        for c, h in enumerate(headers, 1):
            val = entry.get(h, "")
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border

        # 按状态着色
        status_text = entry.get("scheduleStatusText", "")
        if status_text == "上班":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = work_fill
        elif status_text == "休息":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = rest_fill
        elif status_text == "请假":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = leave_fill

    # --- 冻结首行 + 自动筛选 ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- 列宽 ---
    col_widths = {
        1: 12, 2: 18, 3: 22, 4: 12,
        5: 14, 6: 16, 7: 20,
        8: 18, 9: 16, 10: 15, 11: 16,
    }
    for c, w in col_widths.items():
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(str(output_path))
    print(f"      已保存: {output_path}")


# ============================================================
# 统计分析
# ============================================================

def print_summary(raw_data: dict, entries: list[dict]):
    """打印数据摘要。"""
    print()
    print("=" * 60)
    print("  数据摘要")
    print("=" * 60)

    rows = raw_data.get("data", {}).get("rows", [])
    print(f"  员工总数: {len(rows)}")

    # scheduleStatus 分布
    status_counts = {}
    for e in entries:
        s = e["scheduleStatusText"]
        status_counts[s] = status_counts.get(s, 0) + 1

    print(f"  排班条目总数: {len(entries)}")
    print(f"  scheduleStatus 分布:")
    for s, c in sorted(status_counts.items(), key=lambda x: -x[1]):
        print(f"    {s}: {c} 条")

    # 每人的排班天数
    staff_days = {}
    for e in entries:
        name = e["staffName"]
        if name not in staff_days:
            staff_days[name] = {"total": 0, "work": 0, "rest": 0, "leave": 0, "other": 0}
        staff_days[name]["total"] += 1
        st = e["scheduleStatusText"]
        if st == "上班":
            staff_days[name]["work"] += 1
        elif st == "休息":
            staff_days[name]["rest"] += 1
        elif st == "请假":
            staff_days[name]["leave"] += 1
        else:
            staff_days[name]["other"] += 1

    # WW 团队成员
    ww_names = [
        "Saul Lopez", "Karime Resendiz", "Adriana Pacheco", "Yelitza Ruiz",
        "Owen Angeles", "Victor Lerma", "Víctor Lerma", "Brandom Molina",
        "Perla Sanchez", "Guadalupe Cruz", "Juan Ortega",
    ]
    print()
    print(f"  WW 团队排班明细:")
    for name in ww_names:
        if name in staff_days:
            d = staff_days[name]
            print(f"    {name}: 上班{d['work']}天  休息{d['rest']}天  "
                  f"请假{d['leave']}天  其他{d['other']}天  (共{d['total']}天)")
        else:
            # 模糊匹配
            matched = None
            for sn in staff_days:
                if name.lower().replace(" ", "") in sn.lower().replace(" ", ""):
                    matched = sn
                    break
            if matched:
                d = staff_days[matched]
                print(f"    {name} → API「{matched}」: 上班{d['work']}天  休息{d['rest']}天  "
                      f"请假{d['leave']}天  其他{d['other']}天  (共{d['total']}天)")
            else:
                print(f"    {name}: [X] 未在排班中找到")

    # 异常状态
    print()
    unknown = [e for e in entries if e["scheduleStatusText"] not in ("上班", "休息", "请假", "无排班数据")]
    if unknown:
        print(f"  [!] 异常状态条目: {len(unknown)} 条")
        for e in unknown[:10]:
            print(f"    {e['staffName']} | {e['scheduleTimeDate']} | "
                  f"scheduleStatus={e['scheduleStatus']} | text={e['scheduleStatusText']}")
    else:
        print(f"  [OK] 无异常状态")

    # JSON 顶层字段
    print()
    print(f"  JSON 顶层字段: {list(raw_data.keys())}")
    if "data" in raw_data and isinstance(raw_data["data"], dict):
        data_keys = [k for k in raw_data["data"].keys() if k != "rows"]
        print(f"  data 字段 (不含 rows): {data_keys}")

    # 第一条 row 的所有字段
    if rows:
        first_row = rows[0]
        row_keys = [k for k in first_row.keys() if k != "scheduleRespDtos"]
        print(f"  row 字段 (不含 scheduleRespDtos): {row_keys}")

        # scheduleRespDtos 第一条
        scheds = first_row.get("scheduleRespDtos", [])
        if scheds:
            print(f"  scheduleRespDtos[0] 字段: {list(scheds[0].keys())}")
            print(f"  scheduleRespDtos[0] 示例: {scheds[0]}")


# ============================================================
# 主流程
# ============================================================

def main():
    print("=" * 60)
    print("  WW 排班 API Debug 模式")
    print("=" * 60)
    print()

    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    year, month = 2026, 7

    # [1] 请求 API
    raw_data = fetch_raw(year, month)

    # [2] 保存原始 JSON
    raw_path = _OUTPUT_DIR / "schedule_raw.json"
    print(f"\n[2/4] 保存原始 JSON ...")
    with open(raw_path, "w", encoding="utf-8") as f:
        json.dump(raw_data, f, ensure_ascii=False, indent=2, default=str)
    file_size = raw_path.stat().st_size
    print(f"      已保存: {raw_path} ({file_size:,} bytes)")

    # [3] 展开 → Excel
    entries = expand_rows(raw_data, year, month)
    xlsx_path = _OUTPUT_DIR / "schedule_parsed.xlsx"
    write_excel(entries, xlsx_path)

    # [4] 摘要
    print(f"\n[4/4] 分析 ...")
    print_summary(raw_data, entries)

    print()
    print("=" * 60)
    print("  Debug 导出完成")
    print(f"  原始 JSON : {raw_path}")
    print(f"  展开 Excel: {xlsx_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
