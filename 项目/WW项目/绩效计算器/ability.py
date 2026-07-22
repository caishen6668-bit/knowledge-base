# -*- coding: utf-8 -*-
"""
WW 智能绩效与人员调配系统 — Ability Engine（人员能力计算）

每个人、每个阶段独立计算三项指标：
  - Ability（能力）：最近 N 天内在该阶段的平均金额催回率
  - Confidence（稳定性）：基于最近 N 天内在该阶段工作天数的置信度评分
  - Experience（经验）：历史累计在该阶段工作的总天数

数据存储：ability_database.xlsx（不存在则自动创建，不覆盖历史经验）
"""

import os
from datetime import date, timedelta
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import config
from utils import experience_score


# ============================================================
# 稳定性评分
# ============================================================

def stability_score(working_days: int) -> int:
    """根据最近窗口内工作天数计算稳定性评分（分段线性插值）。

    评分参考点（config.STABILITY_POINTS）：
        2天→30, 5天→60, 10天→85, 14天→100
    """
    points = config.STABILITY_POINTS

    if working_days <= points[0][0]:
        return points[0][1]
    if working_days >= points[-1][0]:
        return points[-1][1]

    for i in range(len(points) - 1):
        d1, s1 = points[i]
        d2, s2 = points[i + 1]
        if d1 <= working_days <= d2:
            if d2 == d1:
                return s1
            ratio = (working_days - d1) / (d2 - d1)
            return round(s1 + ratio * (s2 - s1))

    return 0


# ============================================================
# Excel 样式（ability_database 专用）
# ============================================================

_FONT_HEADER = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FONT_DATA = Font(name="微软雅黑", size=10)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

# 每日明细表头
_DETAIL_HEADERS = ["数据日期", "工号", "姓名", "阶段", "金额催回率", "是否休息"]

# 能力总览表头
_SUMMARY_HEADERS = [
    "工号", "姓名", "阶段",
    "RawRate", "AbilityScore", "OverallScore",
    "Confidence", "Experience",
    "LastWorkDate", "Trend",
    "更新时间",
]


def _style_sheet(ws, col_count, data_rows):
    """给 sheet 加统一样式"""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER
    for r in range(2, data_rows + 2):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _FONT_DATA
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER
    for c in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            length = sum(2 if "一" <= ch <= "鿿" else 1 for ch in val)
            max_len = max(max_len, length)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 24)


# ============================================================
# 数据库加载 / 创建
# ============================================================

def _load_or_create_db():
    """加载或创建能力数据库。返回 (workbook, detail_rows)。

    detail_rows: list[dict]，每项为每日明细中的一条记录。
    """
    path = str(config.ABILITY_DB_PATH)
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        detail_rows = []
        if "每日明细" in wb.sheetnames:
            ws = wb["每日明细"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                detail_rows.append({
                    "date": _as_date(row[0]),
                    "staff_id": str(row[1]).strip() if row[1] else "",
                    "name": str(row[2]).strip() if row[2] else "",
                    "stage": str(row[3]).strip() if row[3] else "",
                    "rate": float(row[4]) if row[4] else 0.0,
                    "is_rest": str(row[5]) == "是" if row[5] else False,
                })
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        detail_rows = []

    return wb, detail_rows


def _as_date(val):
    """将单元格值转为 date 对象（兼容 openpyxl 返回的 datetime）"""
    if val is None:
        return None
    if isinstance(val, date):
        # date 也可能是 datetime（datetime 是 date 的子类）
        from datetime import datetime as dt
        if isinstance(val, dt):
            return val.date()
        return val
    if isinstance(val, str):
        from datetime import datetime as dt
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return dt.strptime(val, fmt).date()
            except ValueError:
                pass
    return None


# ============================================================
# 百分位 AbilityScore 计算
# ============================================================

def _compute_percentile(sid_rates):
    """按阶段内 RawRate 排名计算百分位 AbilityScore (0~100)。

    参数：
        sid_rates: {staff_id: raw_rate}  — 同一阶段内所有人的 RawRate

    返回：
        {staff_id: ability_score}  — 0~100 分
    """
    if not sid_rates:
        return {}

    # 按 RawRate 升序
    sorted_items = sorted(sid_rates.items(), key=lambda x: x[1])
    n = len(sorted_items)

    if n == 1:
        return {sorted_items[0][0]: 50.0}

    result = {}
    i = 0
    while i < n:
        # 找到所有相同 RawRate 的人（处理并列）
        j = i
        while j < n and abs(sorted_items[j][1] - sorted_items[i][1]) < 1e-8:
            j += 1
        # 平均排名（0-indexed）
        avg_rank = (i + j - 1) / 2.0
        score = (avg_rank / (n - 1)) * 100.0
        for k in range(i, j):
            result[sorted_items[k][0]] = round(score, 1)
        i = j

    return result


# ============================================================
# 更新能力数据库
# ============================================================

def update(records, data_date):
    """用一天的数据更新能力数据库。

    参数：
        records: list[StaffRecord] — 从 excel_reader 读取的当天员工记录
        data_date: date — 数据日期

    该方法由 main.py 在处理每个日报后调用。
    幂等：同一 (data_date, staff_id) 不会重复写入。
    """
    wb, detail_rows = _load_or_create_db()

    # --- 1. 建立已有记录的索引，避免重复写入 ---
    existing_keys = set()
    for d in detail_rows:
        existing_keys.add((d["date"], d["staff_id"]))

    # --- 2. 将本日新记录追加到每日明细 ---
    new_count = 0
    for rec in records:
        staff_id = rec.staff_id
        if not staff_id:
            continue
        key = (data_date, staff_id)
        if key in existing_keys:
            continue

        # 队列 → 阶段映射
        stage = config.QUEUE_STAGE_MAP.get(rec.queue, rec.queue)

        detail_rows.append({
            "date": data_date,
            "staff_id": staff_id,
            "name": rec.name,
            "stage": stage,
            "rate": rec.amount_rate,
            "is_rest": rec.is_rest,
        })
        existing_keys.add(key)
        new_count += 1

    if new_count == 0:
        wb.close()
        return  # 没有新数据，不重建

    # --- 3. 重建每日明细 sheet ---
    if "每日明细" in wb.sheetnames:
        del wb["每日明细"]

    ws_detail = wb.create_sheet(title="每日明细")
    for c, h in enumerate(_DETAIL_HEADERS, 1):
        ws_detail.cell(row=1, column=c, value=h)

    detail_rows.sort(key=lambda d: (d["date"], d["staff_id"]))
    for i, d in enumerate(detail_rows):
        row = i + 2
        ws_detail.cell(row=row, column=1, value=d["date"])
        ws_detail.cell(row=row, column=2, value=d["staff_id"])
        ws_detail.cell(row=row, column=3, value=d["name"])
        ws_detail.cell(row=row, column=4, value=d["stage"])
        ws_detail.cell(row=row, column=5, value=round(d["rate"], 4))
        ws_detail.cell(row=row, column=6, value="是" if d["is_rest"] else "否")
    _style_sheet(ws_detail, len(_DETAIL_HEADERS), len(detail_rows))

    # --- 4. 计算能力总览 ---
    # 最近 N 天的窗口
    window_start = data_date - timedelta(days=config.ABILITY_WINDOW_DAYS - 1)
    latest_date = data_date  # 用当前数据日期作为窗口结束

    # 按 (staff_id, stage) 分组
    # 能力窗口内的记录（用于 Ability 和 Confidence）
    # 全部历史记录（用于 Experience）
    stage_window_days = defaultdict(set)     # (staff_id, stage) → set of dates (窗口内，非休息)
    stage_window_rates = defaultdict(list)    # (staff_id, stage) → list of rates (窗口内，非休息)
    stage_all_days = defaultdict(set)         # (staff_id, stage) → set of dates (全部历史，非休息)
    staff_name = {}                           # staff_id → 最新 name

    for d in detail_rows:
        sid = d["staff_id"]
        stage = d["stage"]
        key = (sid, stage)
        staff_name[sid] = d["name"]

        if d["is_rest"]:
            continue  # 休息日不算工作

        # Experience：累计所有历史
        stage_all_days[key].add(d["date"])

        # Ability / Confidence：仅窗口内
        if window_start <= d["date"] <= latest_date:
            stage_window_days[key].add(d["date"])
            stage_window_rates[key].append(d["rate"])

    # --- 5. 重建能力总览 sheet ---
    # 跟踪每个 (staff_id, stage) 的最后工作日期
    stage_last_date = {}  # (staff_id, stage) → date
    for d in detail_rows:
        if d["is_rest"]:
            continue
        key = (d["staff_id"], d["stage"])
        cur = stage_last_date.get(key)
        if cur is None or d["date"] > cur:
            stage_last_date[key] = d["date"]

    if "能力总览" in wb.sheetnames:
        del wb["能力总览"]

    ws_summary = wb.create_sheet(title="能力总览")
    for c, h in enumerate(_SUMMARY_HEADERS, 1):
        ws_summary.cell(row=1, column=c, value=h)

    # 排序：工号 → 阶段
    summary_keys = sorted(set(
        list(stage_window_days.keys()) + list(stage_all_days.keys())
    ), key=lambda k: (k[0], k[1]))

    # --- 第一遍：收集每人的 RawRate / Confidence / Experience ---
    summary_data = {}  # (sid, stage) → {raw_rate, confidence, experience, last_date, name}
    stage_raw_rates = defaultdict(dict)  # stage → {sid: raw_rate}

    for sid, stage in summary_keys:
        name = staff_name.get(sid, "")

        rates = stage_window_rates.get((sid, stage), [])
        raw_rate = round(sum(rates) / len(rates), 4) if rates else 0.0

        working_days = len(stage_window_days.get((sid, stage), set()))
        confidence = stability_score(working_days)

        experience = len(stage_all_days.get((sid, stage), set()))

        last_date = stage_last_date.get((sid, stage))

        summary_data[(sid, stage)] = {
            "name": name,
            "raw_rate": raw_rate,
            "confidence": confidence,
            "experience": experience,
            "last_date": last_date,
        }
        stage_raw_rates[stage][sid] = raw_rate

    # --- 第二遍：按阶段计算百分位 AbilityScore (0~100) ---
    stage_ability = {}  # stage → {sid: ability_score}
    for stage, sid_rates in stage_raw_rates.items():
        stage_ability[stage] = _compute_percentile(sid_rates)

    # --- 第三遍：计算 OverallScore 并写入 sheet ---
    row_idx = 2
    for sid, stage in summary_keys:
        data = summary_data[(sid, stage)]
        name = data["name"]
        raw_rate = data["raw_rate"]
        confidence = data["confidence"]
        experience = data["experience"]
        last_date = data["last_date"]

        # AbilityScore：同阶段百分位排名 → 0~100
        ability_score = stage_ability.get(stage, {}).get(sid, 50.0)

        # OverallScore = AbilityScore×0.5 + Confidence×0.3 + ExperienceScore×0.2
        exp_s = experience_score(experience)
        overall = (ability_score * config.OPTIMIZER_W_ABILITY
                   + confidence * config.OPTIMIZER_W_CONFIDENCE
                   + exp_s * config.OPTIMIZER_W_EXPERIENCE)

        trend = ""

        ws_summary.cell(row=row_idx, column=1, value=sid)
        ws_summary.cell(row=row_idx, column=2, value=name)
        ws_summary.cell(row=row_idx, column=3, value=stage)
        ws_summary.cell(row=row_idx, column=4, value=raw_rate)
        ws_summary.cell(row=row_idx, column=5, value=round(ability_score, 1))
        ws_summary.cell(row=row_idx, column=6, value=round(overall, 1))
        ws_summary.cell(row=row_idx, column=7, value=confidence)
        ws_summary.cell(row=row_idx, column=8, value=experience)
        ws_summary.cell(row=row_idx, column=9, value=last_date if last_date else "")
        ws_summary.cell(row=row_idx, column=10, value=trend)
        ws_summary.cell(row=row_idx, column=11, value=latest_date)
        row_idx += 1

    _style_sheet(ws_summary, len(_SUMMARY_HEADERS), row_idx - 2)

    # --- 6. 保存 ---
    wb.save(str(config.ABILITY_DB_PATH))
    wb.close()


# ============================================================
# 查询接口（供后续 recommendation 等模块调用）
# ============================================================

def get_ability(staff_id: str, stage: str) -> dict:
    """查询某人某阶段的能力数据。

    返回：
        {"raw_rate": float, "ability_score": float, "confidence": int,
         "experience": int, "last_work_date": date|None, "trend": str}
        不存在时返回全 0 / None。
    """
    path = str(config.ABILITY_DB_PATH)
    if not os.path.exists(path):
        return {"raw_rate": 0, "ability_score": 0, "confidence": 0,
                "experience": 0, "last_work_date": None, "trend": ""}

    wb = openpyxl.load_workbook(path, data_only=True)
    result = {"raw_rate": 0, "ability_score": 0, "overall_score": 0,
              "confidence": 0, "experience": 0,
              "last_work_date": None, "trend": ""}
    if "能力总览" in wb.sheetnames:
        ws = wb["能力总览"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == staff_id and str(row[2]).strip() == stage:
                result = {
                    "raw_rate": float(row[3]) if row[3] else 0,
                    "ability_score": float(row[4]) if row[4] else 0,
                    "overall_score": float(row[5]) if row[5] else 0,
                    "confidence": int(row[6]) if row[6] else 0,
                    "experience": int(row[7]) if row[7] else 0,
                    "last_work_date": _as_date(row[8]),
                    "trend": str(row[9]) if row[9] else "",
                }
                break
    wb.close()
    return result


def get_all_abilities(staff_id: str = None) -> list:
    """查询全部或某人的能力总览。

    返回：list[dict]
    """
    path = str(config.ABILITY_DB_PATH)
    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    results = []
    if "能力总览" in wb.sheetnames:
        ws = wb["能力总览"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid = str(row[0]).strip() if row[0] else ""
            if staff_id and sid != staff_id:
                continue
            results.append({
                "staff_id": sid,
                "name": str(row[1]).strip() if row[1] else "",
                "stage": str(row[2]).strip() if row[2] else "",
                "raw_rate": float(row[3]) if row[3] else 0,
                "ability_score": float(row[4]) if row[4] else 0,
                "overall_score": float(row[5]) if row[5] else 0,
                "confidence": int(row[6]) if row[6] else 0,
                "experience": int(row[7]) if row[7] else 0,
                "last_work_date": _as_date(row[8]),
                "trend": str(row[9]) if row[9] else "",
            })
    wb.close()
    return results


def get_detail_history(staff_id: str = None, days: int = None) -> list:
    """查询每日明细历史。

    返回：list[dict]
    """
    path = str(config.ABILITY_DB_PATH)
    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    results = []
    if "每日明细" in wb.sheetnames:
        ws = wb["每日明细"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid = str(row[1]).strip() if row[1] else ""
            if staff_id and sid != staff_id:
                continue
            results.append({
                "date": _as_date(row[0]),
                "staff_id": sid,
                "name": str(row[2]).strip() if row[2] else "",
                "stage": str(row[3]).strip() if row[3] else "",
                "rate": float(row[4]) if row[4] else 0.0,
                "is_rest": str(row[5]) == "是" if row[5] else False,
            })
    wb.close()
    return results
