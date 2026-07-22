# -*- coding: utf-8 -*-
"""
WW项目绩效计算器 v1.0
控制台运行，选择日报Excel → 自动计算绩效 → 生成 WW项目绩效汇总.xlsx

用法：
    python WW项目绩效计算器.py

然后按提示选择日报文件即可。
"""

import os
import re
import sys
from pathlib import Path
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

# 确保控制台输出不会因特殊字符报错
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ============================================================
# 绩效方案（硬编码）
# ============================================================
# 格式：{阶段: [(最低值, 最高值, 系数, 绩效金额), ...]}
# 匹配规则：min <= rate < max（最后一档：min <= rate）

# 7月9日方案（2026-07-13 以前使用）
SCHEME_0709 = {
    "D-1": [
        (0,    0.17, 0,   0),
        (0.17, 0.22, 1,   0),
        (0.22, 0.27, 1.2, 0),
        (0.27, 0.32, 1.4, 0),
        (0.32, 0.37, 1.7, 0),
        (0.37, 1,    2,   0),
    ],
    "D0": [
        (0,    0.32, 0,   0),
        (0.32, 0.35, 1,   160),
        (0.35, 0.38, 1.2, 192),
        (0.38, 0.41, 1.4, 224),
        (0.41, 0.44, 1.7, 272),
        (0.44, 1,    2,   320),
    ],
    "S1": [
        (0,    0.08, 0,   0),
        (0.08, 0.10, 1,   220),
        (0.10, 0.12, 1.2, 264),
        (0.12, 0.14, 1.4, 308),
        (0.14, 0.16, 1.7, 374),
        (0.16, 1,    2,   440),
    ],
    "S2": [
        (0,    0.03,  0,   0),
        (0.03, 0.05,  1,   220),
        (0.05, 0.065, 1.2, 264),
        (0.065,0.08,  1.4, 308),
        (0.085,0.10,  1.7, 374),
        (0.10, 1,     2,   440),
    ],
}

# 7月13日方案（2026-07-13 起使用）
SCHEME_0713 = {
    "D-1": [
        (0,    0.17, 0,   0),
        (0.17, 0.22, 1,   0),
        (0.22, 0.27, 1.2, 0),
        (0.27, 0.32, 1.4, 0),
        (0.32, 0.37, 1.7, 0),
        (0.37, 1,    2,   0),
    ],
    "D0": [
        (0,    0.22, 0,   0),
        (0.22, 0.26, 1,   160),
        (0.26, 0.30, 1.2, 192),
        (0.30, 0.34, 1.4, 224),
        (0.34, 0.38, 1.7, 272),
        (0.38, 1,    2,   320),
    ],
    "D1": [
        (0,    0.08, 0,   0),
        (0.08, 0.10, 1,   220),
        (0.10, 0.12, 1.2, 264),
        (0.12, 0.14, 1.4, 308),
        (0.14, 0.16, 1.7, 374),
        (0.16, 1,    2,   440),
    ],
    "S1": [
        (0,    0.05, 0,   0),
        (0.05, 0.08, 1,   220),
        (0.08, 0.11, 1.2, 264),
        (0.11, 0.14, 1.4, 308),
        (0.14, 0.17, 1.7, 374),
        (0.17, 1,    2,   440),
    ],
    "S2": [
        (0,    0.03,  0,   0),
        (0.03, 0.05,  1,   220),
        (0.05, 0.065, 1.2, 264),
        (0.065,0.08,  1.4, 308),
        (0.085,0.10,  1.7, 374),
        (0.10, 1,     2,   440),
    ],
}

# 绩效方案版本列表（按生效日期排序）。
# 每个元素: (生效日期, 方案字典, 方案名称)
# 新增方案只需追加一行，无需修改任何判断逻辑。
SCHEME_VERSIONS = [
    (date(2026, 7, 9),  SCHEME_0709, "7月9日方案"),
    (date(2026, 7, 13), SCHEME_0713, "7月13日方案"),
]

# 队列 → 阶段映射（未列出的队列直接当作阶段名）
QUEUE_STAGE_MAP = {
    "S0": "D0",    # 旧命名 → 新阶段名
    "D2-3": "S1",  # D2-3 即 S1 阶段
}

# 姓名前缀清洗（按顺序尝试，命中即止）
NAME_PREFIXES = ["Feimi-", "Fm-"]

# 日报存放目录（桌面/WW项目日报存档）
DAILY_DIR = Path.home() / "Desktop" / "WW项目日报存档"

# 输出文件（桌面，固定文件名，长期维护同一份）
OUTPUT_PATH = Path.home() / "Desktop" / "WW项目绩效汇总.xlsx"


# ============================================================
# 工具函数
# ============================================================

def get_scheme(data_date):
    """根据数据日期，从 SCHEME_VERSIONS 中自动选择生效日期 <= data_date 的最新一版方案。

    新增方案时只需在 SCHEME_VERSIONS 列表中追加一行，本函数无需修改。
    """
    best = None
    for eff_date, scheme_dict, scheme_name in SCHEME_VERSIONS:
        if eff_date <= data_date:
            best = (scheme_dict, scheme_name)
        else:
            break  # 列表已按生效日期排序，后面的更大，无需继续
    if best is None:
        raise ValueError(
            f"{data_date} 没有可用绩效方案（所有方案的生效日期均晚于该日期）。\n"
            f"请先在 SCHEME_VERSIONS 中新增对应版本的绩效方案。"
        )
    return best


def clean_name(raw_name):
    """清洗姓名：去前后空格、去前缀"""
    if raw_name is None:
        return ""
    name = str(raw_name).strip()
    for prefix in NAME_PREFIXES:
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    return name


def parse_value(raw):
    """将日报文本单元格转为数值。
    '37.25%' -> 0.3725
    '20400.00' -> 20400.0
    '51' -> 51
    """
    if raw is None:
        return 0.0
    s = str(raw).strip()
    if not s or s == "-":
        return 0.0
    if s.endswith("%"):
        try:
            return float(s.rstrip("%")) / 100.0
        except ValueError:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def infer_date(filename, sheet_name):
    """从文件名和 sheet 名推断数据日期。

    文件名如 "10号.xlsx" → day=10
    sheet 名如 "data2026-07-12" → 导出日期 2026-07-12
    规则：day ≤ 导出日 → 同月；否则上一个月
    """
    # 从文件名提取"日"
    m = re.match(r"^(\d{1,2})号\.xlsx$", filename)
    if not m:
        raise ValueError(f"无法从文件名推断日期: {filename}")
    day = int(m.group(1))

    # 从 sheet 名提取导出日期
    m2 = re.search(r"data(\d{4}-\d{2}-\d{2})", sheet_name)
    if not m2:
        raise ValueError(f"无法从 sheet 名推断年月: {sheet_name}")
    export_date = datetime.strptime(m2.group(1), "%Y-%m-%d").date()

    if day <= export_date.day:
        data_date = date(export_date.year, export_date.month, day)
    else:
        # 上一个月
        if export_date.month == 1:
            data_date = date(export_date.year - 1, 12, day)
        else:
            data_date = date(export_date.year, export_date.month - 1, day)

    return data_date


def lookup_tier(rate, tiers):
    """根据催回率查找匹配的档位。

    tiers: [(min, max, coefficient, amount), ...]
    匹配规则：min <= rate < max（最后一档：min <= rate）
    返回：(coefficient, amount)
    """
    for i, (lo, hi, coeff, amount) in enumerate(tiers):
        if i == len(tiers) - 1:
            if rate >= lo:
                return coeff, amount
        else:
            if lo <= rate < hi:
                return coeff, amount
    # 不应该走到这里（第一档 lo=0）
    return 0, 0


# ============================================================
# 日报读取
# ============================================================

def read_daily_report(filepath):
    """读取单个日报文件，返回记录列表和元信息。

    返回：
        (data_date, records, warnings)
        records: [{"name": str, "staff_id": str, "dept": str, "queue": str,
                   "coverage": float, "amount_rate": float,
                   "repaid_amount": float, "due_amount": float}, ...]
    """
    filename = os.path.basename(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sheet_name = ws.title

    # 推断数据日期
    data_date = infer_date(filename, sheet_name)

    records = []
    warnings = []

    # 从第3行开始读数据，跳过表头（第1-2行）和末尾汇总行
    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        # 判断是否为汇总行
        if row[0] and "汇总" in str(row[0]):
            continue

        name_raw = row[0]   # A列：姓名
        staff_id = row[1]   # B列：工号
        dept = row[2]       # C列：部门
        queue = row[3]      # D列：队列
        coverage_raw = row[4]  # E列：覆盖率

        # 跳过空行
        if not name_raw or not str(name_raw).strip():
            continue

        name = clean_name(name_raw)
        coverage = parse_value(coverage_raw)

        # 债务汇总维度：金额催回率（N列，index 13）
        amount_rate_raw = row[13]
        amount_rate = parse_value(amount_rate_raw)

        # 债务汇总维度：回款总额（L列，index 11）
        repaid_amount = parse_value(row[11])

        # 队列新案维度：应催总额（T列，index 19）
        due_amount = parse_value(row[19])

        # O列（index 14）：队列新案维度·分案户数，为 0 表示休息
        rest_cases = parse_value(row[14])
        is_rest = (rest_cases == 0)

        # 数据校验警告
        if is_rest:
            warnings.append(f"  [!] {name}({staff_id}): O列分案户数为0，当天休息")
        elif coverage == 0:
            warnings.append(f"  [!] {name}({staff_id}): 覆盖率 0%，当天未出勤")

        records.append({
            "name": name,
            "staff_id": str(staff_id).strip() if staff_id else "",
            "dept": str(dept).strip() if dept else "",
            "queue": str(queue).strip() if queue else "",
            "coverage": coverage,
            "amount_rate": amount_rate,
            "repaid_amount": repaid_amount,
            "due_amount": due_amount,
            "is_rest": is_rest,
        })

    wb.close()
    return data_date, records, warnings


# ============================================================
# 绩效计算
# ============================================================

def calculate_day(data_date, records):
    """计算一天的绩效。

    返回：
        results: [{"name", "staff_id", "queue", "stage", "rate", "coefficient", "amount"}, ...]
    """
    scheme, scheme_name = get_scheme(data_date)
    results = []

    for rec in records:
        # 休息（O列分案户数为0）：不计绩效
        if rec.get("is_rest"):
            results.append({
                "name": rec["name"],
                "staff_id": rec["staff_id"],
                "queue": rec["queue"],
                "stage": "",
                "rate": 0,
                "coefficient": 0,
                "amount": "休",
                "is_rest": True,
            })
            continue

        queue = rec["queue"]
        # 队列 → 阶段映射
        stage = QUEUE_STAGE_MAP.get(queue, queue)

        # 根据阶段计算催回率
        if stage == "D0":
            # D0：直接使用金额催回率
            rate = rec["amount_rate"]
        elif stage == "S1":
            # S1：回款总额 ÷ 应催总额
            if rec["due_amount"] > 0:
                rate = rec["repaid_amount"] / rec["due_amount"]
            else:
                rate = 0.0
        else:
            # 其他阶段暂用金额催回率（后续可按阶段扩展）
            rate = rec["amount_rate"]

        # 查找档位
        tiers = scheme.get(stage)
        if tiers is None:
            # 阶段不在方案中 → 系数和绩效为 0
            coeff, amount = 0, 0
        else:
            coeff, amount = lookup_tier(rate, tiers)

        results.append({
            "name": rec["name"],
            "staff_id": rec["staff_id"],
            "queue": queue,
            "stage": stage,
            "rate": rate,
            "coefficient": coeff,
            "amount": amount,
            "is_rest": False,
        })

    return results, scheme_name


# ============================================================
# Excel 输出
# ============================================================

# 样式定义
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, col_count):
    """给表头行加样式"""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_data_row(ws, row, col_count):
    """给数据行加样式"""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws, col_count):
    """自动调整列宽"""
    for c in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            # 估算中文字符宽度
            length = 0
            for ch in val:
                if "一" <= ch <= "鿿" or "　" <= ch <= "〿":
                    length += 2
                else:
                    length += 1
            max_len = max(max_len, length)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 30)


def _write_day_sheet(wb, day_str, day_results):
    """写入单日 Sheet（按催回率降序）。"""
    headers = ["姓名", "队列", "阶段", "金额催回率", "系数", "绩效"]
    ws = wb.create_sheet(title=day_str)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    # 休息排最后，其余按催回率降序
    active = [r for r in day_results if not r.get("is_rest")]
    resting = [r for r in day_results if r.get("is_rest")]
    sorted_results = sorted(active, key=lambda r: r["rate"], reverse=True) + resting

    for i, r in enumerate(sorted_results):
        row = i + 2
        ws.cell(row=row, column=1, value=r["name"])
        ws.cell(row=row, column=2, value=r["queue"])
        ws.cell(row=row, column=3, value=r["stage"])
        ws.cell(row=row, column=4, value=round(r["rate"], 4) if not r.get("is_rest") else "-")
        ws.cell(row=row, column=5, value=r["coefficient"] if not r.get("is_rest") else "-")
        ws.cell(row=row, column=6, value=r["amount"])
        style_data_row(ws, row, len(headers))

    auto_width(ws, len(headers))


def _rebuild_summary(wb):
    """扫描所有日 Sheet，完全重建"绩效汇总"Sheet。"""
    # 收集所有日 Sheet 名称
    all_day_strs = sorted(
        [s for s in wb.sheetnames if s.isdigit()],
        key=int
    )

    # 从每个日 Sheet 读取 {name: amount}（"休" 保持为字符串）
    day_perf = {}  # {day_str: {name: amount}}
    for day_str in all_day_strs:
        ws = wb[day_str]
        day_perf[day_str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            amount = row[5]
            if name:
                if amount == "休":
                    day_perf[day_str][name] = "休"
                else:
                    day_perf[day_str][name] = amount if amount else 0

    # 收集所有姓名（保持首次出现的顺序）
    all_names = []
    seen = set()
    for day_str in all_day_strs:
        for name in day_perf[day_str]:
            if name not in seen:
                seen.add(name)
                all_names.append(name)

    # 删除旧"绩效汇总"
    if "绩效汇总" in wb.sheetnames:
        del wb["绩效汇总"]

    # 创建新"绩效汇总"（放到最前面）
    day_labels = [f"{d}号" for d in all_day_strs]
    headers = ["姓名"] + day_labels + ["累计"]
    ws = wb.create_sheet(title="绩效汇总", index=0)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for i, name in enumerate(all_names):
        row = i + 2
        ws.cell(row=row, column=1, value=name)
        total = 0
        for j, day_str in enumerate(all_day_strs):
            amount = day_perf[day_str].get(name, 0)
            if amount == "休":
                ws.cell(row=row, column=2 + j, value="休")
            else:
                val = amount if amount else 0
                ws.cell(row=row, column=2 + j, value=val)
                total += val
        ws.cell(row=row, column=2 + len(all_day_strs), value=total)
        style_data_row(ws, row, len(headers))

    auto_width(ws, len(headers))


def generate_output(all_day_results, output_path):
    """生成或增量更新 WW项目绩效汇总.xlsx。

    如果 output_path 已存在 → 打开，日 Sheet 已存在则跳过不覆盖。
    如果 output_path 不存在 → 新建。
    最后统一重建"绩效汇总"Sheet。
    """
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        if "绩效汇总" in wb.sheetnames:
            del wb["绩效汇总"]
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    # 写入当天 Sheet（已存在则跳过，保护历史数据）
    skipped = []
    for data_date, day_results, scheme_name in all_day_results:
        day_str = str(data_date.day)
        if day_str in wb.sheetnames:
            skipped.append(day_str)
            continue
        _write_day_sheet(wb, day_str, day_results)

    if skipped:
        print(f"  跳过已有日期（未覆盖）: {', '.join(f'{d}号' for d in skipped)}")

    # 扫描所有日 Sheet，完全重建绩效汇总
    _rebuild_summary(wb)

    wb.save(output_path)


# ============================================================
# 文件扫描
# ============================================================

def scan_daily_files(directory):
    """递归扫描目录及子目录下的日报文件（匹配 N号.xlsx 模式）。
    返回 (相对路径, 日数, 文件完整路径) 列表，按日数排序。
    """
    if not directory.exists():
        return []
    files = []
    for root, _dirs, filenames in os.walk(directory):
        for f in filenames:
            m = re.match(r"^(\d{1,2})号\.xlsx$", f)
            if m:
                full_path = Path(root) / f
                rel_path = str(full_path.relative_to(directory))
                files.append((rel_path, int(m.group(1)), str(full_path)))
    # 按日数排序
    files.sort(key=lambda x: x[1])
    return files


# ============================================================
# 主流程
# ============================================================

def main():
    print("=" * 40)
    print("  WW项目绩效计算器 V1.0")
    print("=" * 40)

    # 扫描日报目录
    daily_files = scan_daily_files(DAILY_DIR)
    if not daily_files:
        print(f"\n[!] 日报目录为空或不存在: {DAILY_DIR}")
        print(f"    请将日报文件（如 10号.xlsx）放入该目录后重试。")
        return

    print(f"\n发现日报：\n")
    for rel, day, full in daily_files:
        print(f"  {rel}")

    print(f"\n请选择：\n")
    print(f"  1、全部计算（默认，直接回车）")
    print(f"  2、选择部分日报")
    print()

    choice = input("> ").strip()

    if choice == "2":
        # 选择部分日报
        print()
        for i, (rel, day, full) in enumerate(daily_files, 1):
            print(f"  [{i}] {rel}")
        print(f"\n  输入编号，多个用逗号分隔（如 1,2,3）")

        choice2 = input("\n> ").strip()
        if not choice2:
            selected_files = daily_files  # 没输入就全选
        else:
            selected_files = []
            try:
                choice_normalized = choice2.replace("，", ",")
                indices = [int(x.strip()) for x in choice_normalized.split(",")]
                for idx in indices:
                    if 1 <= idx <= len(daily_files):
                        selected_files.append(daily_files[idx - 1])
                    else:
                        print(f"  [!] 编号 {idx} 超出范围，已跳过")
            except ValueError:
                print(f"  [!] 输入格式错误")
                return
            if not selected_files:
                print("  未选择任何文件，退出。")
                return
    else:
        # 默认：全部计算
        selected_files = daily_files

    print(f"\n共 {len(selected_files)} 个日报待处理。\n")

    # 逐个处理
    all_day_results = []

    for item in selected_files:
        rel_path, day_num, filepath = item

        print(f"正在处理：{rel_path}")
        print("-" * 30)

        try:
            data_date, records, warnings = read_daily_report(filepath)
            results, scheme_name = calculate_day(data_date, records)
            all_day_results.append((data_date, results, scheme_name))

            # 按阶段统计人数
            stage_counts = {}
            for r in results:
                if r.get("is_rest"):
                    stage = "休息"
                elif r["stage"]:
                    stage = r["stage"]
                else:
                    stage = "未分配"
                stage_counts[stage] = stage_counts.get(stage, 0) + 1

            total_perf = sum(r["amount"] for r in results if not r.get("is_rest"))

            print(f"  使用绩效方案：{scheme_name}")
            print(f"  读取员工：{len(records)}人")
            for stage in sorted(stage_counts.keys()):
                print(f"  {stage}：{stage_counts[stage]}人")
            print(f"  总绩效：{total_perf:.0f}")
            print(f"  处理完成。")
            print()

        except Exception as e:
            print(f"  [XX] 处理失败: {e}")
            import traceback
            traceback.print_exc()
            print()
            continue

    if not all_day_results:
        print("没有成功处理任何文件，退出。")
        return

    # 生成输出
    generate_output(all_day_results, str(OUTPUT_PATH))

    # 结束
    print("=" * 40)
    print("  全部处理完成！")
    print(f"  输出文件：{OUTPUT_PATH.name}")
    print("=" * 40)

    # 自动打开
    try:
        os.startfile(str(OUTPUT_PATH))
    except Exception:
        pass


if __name__ == "__main__":
    main()
