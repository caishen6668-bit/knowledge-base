import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Alignment, Font

wb = openpyxl.load_workbook(r'Desktop\周报_6.1-6.7.xlsx')

new_warnings = ['David Osorio']
old_warnings = ['Dante Hernandez', 'Adolfo Monzon']
all_warnings = new_warnings + old_warnings

ws_new = wb[wb.sheetnames[1]]  # 新客
ws_old = wb[wb.sheetnames[2]]  # 老客
ws_all = wb['整体']

# ---- 新客: data ends R26, warning at R29-30 ----
for mr in list(ws_new.merged_cells.ranges):
    if mr.min_row >= 27 and mr.max_row <= 35:
        ws_new.unmerge_cells(str(mr))
for r in range(27, 35):
    for c in range(1, 15):
        try: ws_new.cell(row=r, column=c).value = None
        except AttributeError: pass
ws_new.merge_cells('A29:I30')
c = ws_new.cell(row=29, column=1)
c.value = '警告信名单：' + ', '.join(new_warnings)
c.alignment = Alignment(horizontal='center', vertical='center')
c.font = Font(bold=True, size=11, color='CC0000', name='Microsoft YaHei')

# ---- 老客: data ends R27, warning at R29-30 ----
for mr in list(ws_old.merged_cells.ranges):
    if mr.min_row >= 28 and mr.max_row <= 35:
        ws_old.unmerge_cells(str(mr))
for r in range(28, 35):
    for c in range(1, 15):
        try: ws_old.cell(row=r, column=c).value = None
        except AttributeError: pass
ws_old.merge_cells('A29:I30')
c = ws_old.cell(row=29, column=1)
c.value = '警告信名单：' + ', '.join(old_warnings)
c.alignment = Alignment(horizontal='center', vertical='center')
c.font = Font(bold=True, size=11, color='CC0000', name='Microsoft YaHei')

# ---- 整体 新客区: ends R26, warning at R27-28 ----
for mr in list(ws_all.merged_cells.ranges):
    if mr.min_row >= 27 and mr.max_row <= 29:
        ws_all.unmerge_cells(str(mr))
for r in range(27, 31):
    for c in range(1, 15):
        try: ws_all.cell(row=r, column=c).value = None
        except AttributeError: pass
if new_warnings:
    ws_all.merge_cells('A27:I28')
    c = ws_all.cell(row=27, column=1)
    c.value = '警告信名单：' + ', '.join(new_warnings)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.font = Font(bold=True, size=11, color='CC0000', name='Microsoft YaHei')

# Ensure R30 has title
ws_all.cell(row=30, column=1).value = '6.1-6.7 PL组员回收情况汇总'
ws_all.cell(row=30, column=9).value = '老客组别'

# ---- 整体 老客区: ends R54, warning at R55-56 ----
for mr in list(ws_all.merged_cells.ranges):
    if mr.min_row >= 55 and mr.max_row <= 57:
        ws_all.unmerge_cells(str(mr))
for r in range(55, 60):
    for c in range(1, 15):
        try: ws_all.cell(row=r, column=c).value = None
        except AttributeError: pass
if old_warnings:
    ws_all.merge_cells('A55:I56')
    c = ws_all.cell(row=55, column=1)
    c.value = '警告信名单：' + ', '.join(old_warnings)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.font = Font(bold=True, size=11, color='CC0000', name='Microsoft YaHei')

wb.save(r'Desktop\周报_6.1-6.7.xlsx')
print('Done!')
print('New warnings:', new_warnings)
print('Old warnings:', old_warnings)
print('All warnings:', all_warnings)
