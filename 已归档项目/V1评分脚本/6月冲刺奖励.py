"""
6月冲刺奖励活动 - PL团队分析 一键脚本
用法: python june_reward.py [--send-user] [--send-group]
  --send-user   生成后发到黄悦华私聊审核
  --send-group  生成后直接发到群（需先审核通过）
"""
import sys, io, os, unicodedata, json, requests, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from collections import defaultdict
from datetime import datetime, date, timedelta
from calendar import monthrange
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from alibabacloud_quickbi_public20220101.client import Client
from alibabacloud_quickbi_public20220101 import models as qb_models
from alibabacloud_tea_openapi import models as open_api_models

# ============================================================
# CONFIG
# ============================================================
# ============================================================
# 活动配置
# ============================================================
CHAMPION_PRIZES = [1000, 800, 500]
PROGRESS_PRIZES = [800, 500]
NEWCOMER_RULES = {1.20: "1000 MXN礼包", 1.10: "800 MXN礼包", 1.00: "500 MXN礼包"}
MIN_CHAMPION_RATE = 1.00
MIN_PROGRESS_RATE = 0.90
MIN_ATTEND_RATE = 0.90
MIN_ATTEND_DAYS = 7

# ============================================================
# API & 飞书配置
# ============================================================
AK_ID = os.environ.get("QBI_ACCESS_KEY", "")
AK_SECRET = os.environ.get("QBI_SECRET_KEY", "")
FEISHU_APP_ID = os.environ.get("FEISHU_COLLECTION_APP_ID", "")
FEISHU_APP_SECRET = os.environ.get("FEISHU_COLLECTION_APP_SECRET", "")
USER_OPEN_ID = "ou_ffab3a07f1ff9fbca2a593c0d5e152ac"
CHAT_ID = "oc_8b5ef4aee4e93b29326cd8c0f3c24d90"
YANG_OPEN_ID = "ou_bedb769615878c83d5fbd09beb0108e1"
today_calc = date.today()
year = today_calc.year
month = today_calc.month
day = today_calc.day
last_day = monthrange(year, month)[1]

# Previous month info
if month == 1:
    py = year - 1; pm = 12
else:
    py = year; pm = month - 1
prev_last_day = monthrange(py, pm)[1]

if day <= 2:
    # 月初1-2号：默认回到上个月第二周期
    PERIOD_NAME = "第二周期"
    CYCLE_START = f"{py}{pm:02d}16"
    CYCLE_END = f"{py}{pm:02d}{prev_last_day:02d}"
    PREV_START = f"{py}{pm:02d}01"
    PREV_END = f"{py}{pm:02d}15"
elif day <= 15:
    PERIOD_NAME = "第一周期"
    CYCLE_START = f"{year}{month:02d}01"
    CYCLE_END = f"{year}{month:02d}15"
    PREV_START = f"{py}{pm:02d}16"
    PREV_END = f"{py}{pm:02d}{prev_last_day:02d}"
else:
    PERIOD_NAME = "第二周期"
    CYCLE_START = f"{year}{month:02d}16"
    CYCLE_END = f"{year}{month:02d}{last_day:02d}"
    PREV_START = f"{year}{month:02d}01"
    PREV_END = f"{year}{month:02d}15"

def norm(n):
    n = n.strip()
    n = unicodedata.normalize('NFKD', n).encode('ascii', 'ignore').decode('ascii')
    return n.lower()

# ============================================================
# STEP 1: DETECT PERIOD
# ============================================================
print("[1/6] Detecting period...")

config = open_api_models.Config(access_key_id=AK_ID, access_key_secret=AK_SECRET, region_id="cn-hangzhou")
config.endpoint = "quickbi-public.cn-hangzhou.aliyuncs.com"
config.read_timeout = 180000; config.connect_timeout = 30000
client = Client(config)

def fetch_perf(client, cycle_start, cycle_end, prev_start, prev_end):
    """Fetch performance data for given periods with per-day queries."""
    att_data = []
    for period_start, period_end in [(cycle_start, cycle_end), (prev_start, prev_end)]:
        d_start = datetime.strptime(period_start, "%Y%m%d")
        d_end = datetime.strptime(period_end, "%Y%m%d")
        d_curr = d_start
        while d_curr <= d_end:
            cond = json.dumps({"statis_date": d_curr.strftime("%Y%m%d")})
            req = qb_models.QueryDataServiceRequest(api_id="c4f429db60b3", conditions=cond)
            resp = client.query_data_service(req)
            if resp.body.result.values:
                att_data.extend(resp.body.result.values)
            d_curr += timedelta(days=1)
    return att_data

# Probe: quick check if current period has data
probe = fetch_perf(client, CYCLE_START, CYCLE_START, PREV_START, PREV_START)  # one day each
pl_probe = [r for r in probe if r.get('dept_name', '') == 'PL' and CYCLE_START <= str(r.get('statis_date', '')) <= CYCLE_START]
if not pl_probe:
    # Fallback: swap to the other period
    if PERIOD_NAME == "第二周期":
        PERIOD_NAME = "第一周期"
        CYCLE_START = f"{year}{month:02d}01"
        CYCLE_END = f"{year}{month:02d}15"
        PREV_START = f"{py}{pm:02d}16"
        PREV_END = f"{py}{pm:02d}{prev_last_day:02d}"
    else:
        PERIOD_NAME = "第二周期"
        CYCLE_START = f"{year}{month:02d}16"
        CYCLE_END = f"{year}{month:02d}{last_day:02d}"
        PREV_START = f"{year}{month:02d}01"
        PREV_END = f"{year}{month:02d}15"
    # Re-probe with new dates
    probe = fetch_perf(client, CYCLE_START, CYCLE_START, PREV_START, PREV_START)
    pl_probe = [r for r in probe if r.get('dept_name', '') == 'PL' and CYCLE_START <= str(r.get('statis_date', '')) <= CYCLE_START]
    if not pl_probe:
        print("No data found in either cycle")
        sys.exit(1)

print(f"  Period: {PERIOD_NAME} ({CYCLE_START}~{CYCLE_END})")

# ============================================================
# STEP 2: FETCH ALL DATA
# ============================================================
print("[2/6] Fetching data from Quick BI...")

att_data = fetch_perf(client, CYCLE_START, CYCLE_END, PREV_START, PREV_END)

req = qb_models.QueryDataServiceRequest(api_id="7f9969dc9020"); resp = client.query_data_service(req)
emp_data = resp.body.result.values

pl_att = [r for r in att_data if r.get('dept_name', '') == 'PL' and 'S3' not in str(r.get('overdue_level', ''))]
pl_emp = [r for r in emp_data if r.get('depart_name', '') == 'PL']
print(f"  Performance: {len(pl_att)} rows, Employee: {len(pl_emp)} rows")

# ============================================================
# STEP 3: BUILD EMPLOYEE DATA
# ============================================================
print("[3/6] Building employee data...")
emp_info = {}; working_days = defaultdict(set); scheduled_days = defaultdict(set)

for r in pl_emp:
    name = r.get('name', '').strip()
    if not name: continue
    nk = norm(name); day = str(r.get('day', '')); status = r.get('working_status', '')
    if nk not in emp_info:
        emp_info[nk] = {'name': name, 'employer_no': r.get('employer_no', ''),
                        'group': r.get('group', ''), 'is_resigned': r.get('is_resigned', ''),
                        'join_date': str(r.get('join_date', ''))}
    if status == '上班': working_days[nk].add(day)
    if status != '休息': scheduled_days[nk].add(day)

today_tenure = date.today()
for nk, info in emp_info.items():
    jd = info['join_date']
    if jd and jd != 'None' and len(jd) == 8:
        try: info['tenure'] = (today_tenure - date(int(jd[:4]), int(jd[4:6]), int(jd[6:8]))).days
        except: info['tenure'] = 0
    else: info['tenure'] = 0

all_dates = sorted(set(str(r.get('statis_date', '')) for r in pl_att if CYCLE_START <= str(r.get('statis_date', '')) <= CYCLE_END))
ACTUAL_DAYS = len(all_dates)
cycle_dates_set = set(all_dates)
cycle_year = int(CYCLE_START[:4]); cycle_month = int(CYCLE_START[4:6])
print(f"  Available dates: {ACTUAL_DAYS} ({all_dates[0]}~{all_dates[-1]})")

# ============================================================
# STEP 4: AGGREGATE PERFORMANCE
# ============================================================
print("[4/7] Calculating achievement rates...")
curr = defaultdict(lambda: {'repay': 0, 'enter_x_target': 0, 'perf_dates': set()})
prev = defaultdict(lambda: {'repay': 0, 'enter_x_target': 0, 'perf_dates': set()})

for r in pl_att:
    name = r.get('staff_name', '').strip()
    if not name or name.startswith('Training'): continue
    nk = norm(name); d = str(r.get('statis_date', ''))
    repay = float(r.get('today_repay_amount', 0) or 0)
    enter = float(r.get('today_enter_collect_amount', 0) or 0)
    t_str = r.get('target', 0)
    try: target = float(t_str) if t_str else 0
    except: target = 0

    wd_set = working_days.get(nk, set())
    if CYCLE_START <= d <= CYCLE_END and d in wd_set:
        curr[nk]['repay'] += repay; curr[nk]['enter_x_target'] += enter * target
        curr[nk]['perf_dates'].add(d)
    elif PREV_START <= d <= PREV_END and d in wd_set:
        prev[nk]['repay'] += repay; prev[nk]['enter_x_target'] += enter * target
        prev[nk]['perf_dates'].add(d)

# Calculate results
results = []
for nk, info in emp_info.items():
    if info.get('is_resigned', '') == '离职': continue
    c = curr.get(nk, {}); p = prev.get(nk, {})
    # Attendance rate = actual work days / scheduled days (excludes rest days)
    perf_dates = c.get('perf_dates', set())
    attend_days = len(perf_dates)
    sdays = scheduled_days.get(nk, set()) & cycle_dates_set
    scheduled_count = len(sdays)
    attend_rate = attend_days / scheduled_count if scheduled_count > 0 else 0
    attend_rate_cycle = attend_rate
    achieve = c['repay'] / c['enter_x_target'] if c.get('enter_x_target', 0) > 0 else 0
    prev_achieve = p['repay'] / p['enter_x_target'] if p.get('enter_x_target', 0) > 0 else None
    progress = achieve - prev_achieve if prev_achieve is not None else None
    composite = achieve * 0.8 + attend_rate * 0.2

    results.append({'name': info['name'], 'employer_no': info['employer_no'], 'group': info['group'],
                    'tenure': info['tenure'], 'join_date': info['join_date'], 'attend_days': attend_days,
                    'attend_rate': attend_rate, 'attend_rate_cycle': attend_rate_cycle,
                    'achieve_rate': achieve, 'composite': composite, 'prev_achieve': prev_achieve,
                    'progress': progress})

results.sort(key=lambda x: x['composite'], reverse=True)

# Awards
champions = [r for r in results if r['tenure']>=30 and r['achieve_rate']>=1.0 and r['attend_days']>=7 and r['attend_rate']>=0.90]
progress_list = [r for r in results if r['tenure']>=30 and r['attend_days']>=7 and r['attend_rate']>=0.90 and r['achieve_rate']>=0.90 and r['prev_achieve'] is not None and r['prev_achieve'] >= 0.70]
progress_list.sort(key=lambda x: x['progress'], reverse=True)
newcomers = []
for r in results:
    if 4 <= r['tenure'] <= 30 and r['attend_rate'] >= 0.90:
        rate = r['achieve_rate']
        if rate >= 1.20: reward = '1000 MXN礼包'
        elif rate >= 1.10: reward = '800 MXN礼包'
        elif rate >= 1.00: reward = '500 MXN礼包'
        else: reward = None
        newcomers.append({**r, 'reward': reward, 'n_eligible': rate >= 1.0})
newcomers.sort(key=lambda x: x['achieve_rate'], reverse=True)
newcomer_winners = [n for n in newcomers if n['n_eligible']]

# ============================================================
# FINAL WINNERS (同周期不重复获奖)
# ============================================================
champion_winners = champions[:3]
winner_names = {r['name'] for r in champion_winners}

progress_winners = []
for r in progress_list:
    if r['name'] not in winner_names:
        progress_winners.append(r)
        winner_names.add(r['name'])
    if len(progress_winners) >= 2:
        break

final_newcomer_winners = []
for r in newcomer_winners:
    if r['name'] not in winner_names:
        final_newcomer_winners.append(r)
        winner_names.add(r['name'])

print(f"  Champion Winners: {len(champion_winners)}, Progress Winners: {len(progress_winners)}, Newcomer Winners: {len(final_newcomer_winners)}")

# ============================================================
# STEP 5: VALIDATE
# ============================================================
print("[5/7] Validating...")
if PERIOD_NAME == "第一周期":
    checks = [
        ('Gabriela Jimenez', 1.00, 1.30), ('Lourdes Nunez', 1.00, 1.35),
        ('Luz Lopez', 1.00, 1.40), ('Jovanna Santoyo', 1.05, 1.55),
        ('Laura Bautista', 0.95, 1.20), ('Sandra Contreras', 1.00, 1.35),
    ]
else:
    checks = [
        ('Gabriela Jimenez', 0.65, 1.30), ('Lourdes Nunez', 0.70, 1.30),
        ('Luz Lopez', 0.70, 1.30), ('Jovanna Santoyo', 0.70, 1.30),
        ('Laura Bautista', 0.70, 1.30), ('Sandra Contreras', 0.70, 1.30),
    ]
errors = []
for check_name, lo, hi in checks:
    nk = norm(check_name)
    wdays = {d for d in working_days[nk] if CYCLE_START<=d<=CYCLE_END}
    sdays = {d for d in scheduled_days[nk] if CYCLE_START<=d<=CYCLE_END}
    achieve = curr[nk]['repay'] / curr[nk]['enter_x_target'] if curr[nk].get('enter_x_target', 0) > 0 else 0
    ar = len(wdays) / len(sdays) if len(sdays) > 0 else 0
    if not (lo <= achieve <= hi): errors.append(f'{check_name}: achieve={achieve:.2%} out of range')
    if ar < 0.75: errors.append(f'{check_name}: attend_rate={ar:.0%} too low')
    print(f'  {check_name}: achieve={achieve:.2%} attend={len(wdays)}/{len(sdays)}d')

if errors:
    print(f'  VALIDATION FAILED: {errors}')
    sys.exit(1)
print("  VALIDATION PASSED")

# ============================================================
# STEP 6: GENERATE EXCEL
# ============================================================
print("[6/7] Generating Excel report...")

hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_font = Font(bold=True, size=11, color="FFFFFF", name="Microsoft YaHei")
title_font = Font(bold=True, size=14, color="1F4E79", name="Microsoft YaHei")
sub_font = Font(bold=True, size=12, color="2E75B6", name="Microsoft YaHei")
normal_font = Font(size=10.5, name="Microsoft YaHei")
note_font = Font(italic=True, size=10, color="666666", name="Microsoft YaHei")
warn_font = Font(italic=True, size=10, color="CC0000", name="Microsoft YaHei")
border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
gold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

wb = openpyxl.Workbook()

def set_hdr(ws, row, n):
    for c in range(1, n+1):
        cell = ws.cell(row=row, column=c); cell.font = hdr_font; cell.fill = hdr_fill
        cell.border = border; cell.alignment = center

def set_cell(ws, row, col, value, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value); cell.font = font or normal_font
    cell.border = border; cell.alignment = align or left_align
    if fill: cell.fill = fill

def set_row(ws, row, values, font=None, fill=None):
    for c, v in enumerate(values, 1): set_cell(ws, row, c, v, font=font, fill=fill)

# ---- Sheet 1: Report Summary ----
ws = wb.active; ws.title = "激励活动报告"
for c, w in enumerate([4, 28, 12, 12, 16, 22, 28], 1): ws.column_dimensions[chr(64+c)].width = w

row = 1
set_cell(ws, row, 1, f"{cycle_month}月冲刺奖励活动 — PL团队分析", font=title_font)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
cycle_start_text = f"{int(CYCLE_START[4:6])}月{int(CYCLE_START[6:8])}日"
cycle_end_text = f"{int(CYCLE_END[4:6])}月{int(CYCLE_END[6:8])}日"
set_cell(ws, row, 1, f"{PERIOD_NAME}：{cycle_year}年{cycle_start_text}-{cycle_end_text}", font=sub_font)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
set_cell(ws, row, 1, f"数据截至{int(CYCLE_END[4:6])}月{int(CYCLE_END[6:8])}日 | {PERIOD_NAME} | PL团队在职{len(results)}人 | {datetime.now().strftime('%Y-%m-%d %H:%M')}", font=note_font)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 2
set_cell(ws, row, 1, "最终发奖名单", font=sub_font)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
for i, r in enumerate(champion_winners):
    prize = CHAMPION_PRIZES[i]
    set_cell(ws, row, 1, f"🏆 冠军奖：{r['name']} - {prize} MXN")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
for i, r in enumerate(progress_winners):
    prize = PROGRESS_PRIZES[i]
    set_cell(ws, row, 1, f"📈 最佳进步奖：{r['name']} - {prize} MXN")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
for r in final_newcomer_winners:
    reward = r["reward"].replace("礼包", "")
    set_cell(ws, row, 1, f"🎁 新人达标奖：{r['name']} - {reward}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
row += 2
row += 1
set_cell(ws, row, 1, "预算统计", font=sub_font)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
champ_budget = sum(CHAMPION_PRIZES[:len(champion_winners)])
prog_budget = sum(PROGRESS_PRIZES[:len(progress_winners)])
newc_budget = sum(int(w['reward'].replace(' MXN礼包','').replace(' MXN','')) for w in final_newcomer_winners)
total_budget = champ_budget + prog_budget + newc_budget
set_cell(ws, row, 1, f"冠军奖：{' + '.join(str(p) for p in CHAMPION_PRIZES[:len(champion_winners)])} = {champ_budget} MXN")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7); row += 1
if progress_winners:
    set_cell(ws, row, 1, f"最佳进步奖：{' + '.join(str(p) for p in PROGRESS_PRIZES[:len(progress_winners)])} = {prog_budget} MXN")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7); row += 1
if final_newcomer_winners:
    set_cell(ws, row, 1, f"新人达标奖：{' + '.join(str(int(w['reward'].replace(' MXN礼包','').replace(' MXN',''))) for w in final_newcomer_winners)} = {newc_budget} MXN")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7); row += 1
set_cell(ws, row, 1, f"预算合计：{total_budget} MXN", font=Font(bold=True, size=11, color="CC0000", name="Microsoft YaHei"))
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 2
set_cell(ws, row, 1, "核心发现", font=Font(bold=True, size=12, color="CC0000", name="Microsoft YaHei"))

# Champion
ctitle = "🏆 一、冠军员工奖（当前排名）"
set_cell(ws, row, 1, ctitle, font=sub_font); ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
set_cell(ws, row, 1, "条件：入职≥30天 | 达成率≥100% | 出勤率≥90% | 出勤≥7天 | 无严重违规", font=note_font)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
set_cell(ws, row, 1, "计算：综合评分 = 达成率×80% + 出勤率×20% | 出勤率 = 实际上班天数 ÷ 应出勤天数（除休息外）", font=note_font)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
if champion_winners:
    p = champion_winners
    ptxt = f"获奖：第1名 {p[0]['name']} 1000 MXN"
    if len(p)>=2: ptxt += f" | 第2名 {p[1]['name']} 800 MXN"
    if len(p)>=3: ptxt += f" | 第3名 {p[2]['name']} 500 MXN"
    set_cell(ws, row, 1, ptxt, font=Font(size=10, color="006600", name="Microsoft YaHei"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 2

hdr = ['排名', '姓名', '达成率', '出勤天数', '出勤率', '综合评分', '达标状态']
for c, h in enumerate(hdr, 1): set_cell(ws, row, c, h, font=hdr_font, fill=None, align=center)
ws.cell(row=row, column=1).fill = hdr_fill  # tricky due to set_cell override
for c in range(1,8): ws.cell(row=row, column=c).fill = hdr_fill
row += 1

champion_show = sorted(
    results,
    key=lambda x: x['composite'],
    reverse=True
)[:10]
for i, r in enumerate(champion_show[:12]):
    rank = i + 1
    gaps = []
    if r['tenure'] < 30: gaps.append(f"入职仅{r['tenure']}天")
    if r['achieve_rate'] < 1.0: gaps.append(f"达成差{1.0-r['achieve_rate']:.1%}")
    if r['attend_days'] < 7: gaps.append(f"出勤仅{r['attend_days']}天")
    if r['attend_rate'] < 0.90: gaps.append(f"出勤率差{0.90-r['attend_rate']:.1%}")
    if not gaps: gaps.append("✅ 达标")
    vals = [rank, r['name'], f"{r['achieve_rate']:.2%}", r['attend_days'],
            f"{r['attend_rate']:.2%}", f"{r['composite']:.4f}", gaps[0]]
    set_row(ws, row, vals, fill=green_fill if '✅' in gaps[0] else None)
    if rank <= 3:
        for c in range(1, 8): ws.cell(row=row, column=c).fill = gold_fill
    row += 1

row += 2

# Progress
ptitle = "📈 二、最佳进步奖（当前排名）"
set_cell(ws, row, 1, ptitle, font=sub_font); ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
set_cell(ws, row, 1, "条件：入职≥30天 | 出勤率≥90% | 出勤≥7天 | 达成率≥90% | 上期达成率≥70%", font=note_font)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
if progress_winners:
    pp = progress_winners
    pptxt = f"获奖：第1名 {pp[0]['name']} 进步{pp[0]['progress']:+.2%} 800 MXN"
    if len(pp)>=2: pptxt += f" | 第2名 {pp[1]['name']} 进步{pp[1]['progress']:+.2%} 500 MXN"
    set_cell(ws, row, 1, pptxt, font=Font(size=10, color="006600", name="Microsoft YaHei"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 2

hdr2 = ['排名', '姓名', '本期达成率', '上期达成率', '进步值', '出勤天数', '达标状态']
for c, h in enumerate(hdr2, 1): set_cell(ws, row, c, h, font=hdr_font, fill=None, align=center)
for c in range(1,8): ws.cell(row=row, column=c).fill = hdr_fill
row += 1

candidates = sorted(
    [r for r in results if r['progress'] is not None],
    key=lambda x: (1 if (x['prev_achieve'] is not None and x['prev_achieve'] >= 0.70) else 0, x['progress']),
    reverse=True
)[:10]
for i, r in enumerate(candidates[:10]):
    rank = i + 1
    gaps = []
    if r['prev_achieve'] is not None and r['prev_achieve'] < 0.70: gaps.append("上周期未达70%")
    if r['tenure'] < 30: gaps.append(f"入职仅{r['tenure']}天")
    if r['achieve_rate'] < MIN_PROGRESS_RATE: gaps.append(f"达成率差{MIN_PROGRESS_RATE-r['achieve_rate']:.1%}")
    if r['attend_days'] < MIN_ATTEND_DAYS: gaps.append(f"出勤仅{r['attend_days']}天")
    if r['attend_rate'] < MIN_ATTEND_RATE: gaps.append(f"出勤率差{MIN_ATTEND_RATE-r['attend_rate']:.1%}")
    if r['prev_achieve'] is None: gaps.append("无上期数据")
    status = "✅ 达标" if not gaps else gaps[0]
    vals = [rank, r['name'], f"{r['achieve_rate']:.2%}", f"{r['prev_achieve']:.2%}" if r['prev_achieve'] is not None else "N/A",
            f"{r['progress']:+.2%}" if r['progress'] is not None else "N/A", r['attend_days'], status]
    set_row(ws, row, vals, fill=green_fill if status.startswith("✅") else None)
    row += 1

row += 2

# Newcomer
ntitle = f"🎁 三、新人达标奖：{len(final_newcomer_winners)}人获奖" if newcomer_winners else "🎁 三、新人达标奖：暂无人达标"
set_cell(ws, row, 1, ntitle, font=sub_font); ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
set_cell(ws, row, 1, "条件：入职4-30天 | 出勤率≥90% | 达成100%→500MXN, 110%→800MXN, 120%→1000MXN | 不限人数", font=note_font)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
if final_newcomer_winners:
    wtxt = '获奖：' + ' | '.join(f"{w['name']} {w['reward']}" for w in final_newcomer_winners)
    set_cell(ws, row, 1, wtxt, font=Font(size=10, color="006600", name="Microsoft YaHei"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 2

hdr3 = ['排名', '姓名', '小组', '入职天数', '出勤天数', '出勤率', '达成率', '达标等级', '奖励', '达标状态']
for c, h in enumerate(hdr3, 1): set_cell(ws, row, c, h, font=hdr_font, fill=None, align=center)
for c in range(1,12): ws.cell(row=row, column=c).fill = hdr_fill
row += 1

newcomer_show1 = sorted([r for r in results if 4 <= r['tenure'] <= 30], key=lambda x: x['achieve_rate'], reverse=True)
for idx, r in enumerate(newcomer_show1, start=1):
    rate = r['achieve_rate']
    lvl = '120%+' if rate>=1.20 else ('110%+' if rate>=1.10 else ('100%+' if rate>=1.00 else '未达标'))
    rwd = '1000 MXN' if rate>=1.20 else ('800 MXN' if rate>=1.10 else ('500 MXN' if rate>=1.00 else '—'))
    gaps = []
    if r['tenure'] < 4: gaps.append(f"入职仅{r['tenure']}天")
    if r['attend_rate'] < MIN_ATTEND_RATE: gaps.append(f"出勤率差{MIN_ATTEND_RATE-r['attend_rate']:.1%}")
    if rate < 1.00: gaps.append(f"达成率差{1.00-rate:.1%}")
    status = "✅ 获奖" if not gaps else gaps[0]
    vals = [idx, r['name'], r['group'], r['tenure'], r['attend_days'],
            f"{r['attend_rate']:.2%}", f"{rate:.2%}", lvl, rwd, status]
    set_row(ws, row, vals, fill=green_fill if status.startswith("✅") else None)
    row += 1

row += 2
notes = ["⚠️ 说明:",
         f"1. 周期进度：出勤率 = 实际上班天数 ÷ 应出勤天数（除休息外）",
         "2. 达成率 = 累计回收金额 / 累计(入催金额×目标), 仅计算出勤日, 与BI看板口径一致",
         "3. 同周期不重复获奖, 同一员工只发最高奖项",
         "4. 并列情况按活动规则对应名次奖金平均分配",
         "5. 需要合规数据确认无严重违规后最终发奖",
         f"6. 出勤率 = 实际上班天数 ÷ 应出勤天数（除休息外）, {PERIOD_NAME} {int(CYCLE_START[4:6])}月{int(CYCLE_START[6:8])}日-{int(CYCLE_END[4:6])}月{int(CYCLE_END[6:8])}日"]
for note in notes:
    set_cell(ws, row, 1, note, font=warn_font if note.startswith("⚠️") else note_font)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

# ---- Sheet 2: Champion Full Ranking ----
ws2 = wb.create_sheet("01_冠军奖")
h = ['排名','姓名','小组','入职天数','出勤天数','出勤率','达成率','综合评分','上期达成率','进步值','达标状态']
ws2.append(h); set_hdr(ws2, 1, len(h))
for i, r in enumerate(results):
    q = r['tenure']>=30 and r['achieve_rate']>=MIN_CHAMPION_RATE and r['attend_days']>=MIN_ATTEND_DAYS and r['attend_rate']>=MIN_ATTEND_RATE
    gaps = []
    if r['tenure'] < 30: gaps.append(f"入职仅{r['tenure']}天")
    if r['achieve_rate'] < MIN_CHAMPION_RATE: gaps.append(f"达成率差{MIN_CHAMPION_RATE-r['achieve_rate']:.1%}")
    if r['attend_days'] < MIN_ATTEND_DAYS: gaps.append(f"出勤仅{r['attend_days']}天")
    if r['attend_rate'] < MIN_ATTEND_RATE: gaps.append(f"出勤率差{MIN_ATTEND_RATE-r['attend_rate']:.1%}")
    status = "✅ 达标" if not gaps else gaps[0]
    vals = [i+1, r['name'], r['group'], r['tenure'], r['attend_days'],
            f"{r['attend_rate']:.2%}", f"{r['achieve_rate']:.2%}",
            f"{r['composite']:.4f}", f"{r['prev_achieve']:.2%}" if r['prev_achieve'] else 'N/A',
            f"{r['progress']:+.2%}" if r['progress'] is not None else 'N/A', status]
    set_row(ws2, i+2, vals)
for i, w in enumerate([8,22,12,10,10,12,12,12,14,12,20]): ws2.column_dimensions[chr(65+i)].width = w

# ---- Sheet 3: Progress ----
ws3 = wb.create_sheet("02_最佳进步奖")
h = ['排名','姓名','小组','入职天数','本期达成率','上期达成率','进步值','出勤天数','出勤率','达标状态']
ws3.append(h); set_hdr(ws3, 1, len(h))
sorted_prog = sorted(results, key=lambda x: (1 if (x['prev_achieve'] is not None and x['prev_achieve'] >= 0.70) else 0, x['progress'] if x['progress'] else -999), reverse=True)
for i, r in enumerate(sorted_prog):
    gaps = []
    if r['prev_achieve'] is not None and r['prev_achieve'] < 0.70: gaps.append("上周期未达70%")
    if r['tenure'] < 30: gaps.append(f"入职仅{r['tenure']}天")
    if r['achieve_rate'] < MIN_PROGRESS_RATE: gaps.append(f"达成率差{MIN_PROGRESS_RATE-r['achieve_rate']:.1%}")
    if r['attend_days'] < MIN_ATTEND_DAYS: gaps.append(f"出勤仅{r['attend_days']}天")
    if r['attend_rate'] < MIN_ATTEND_RATE: gaps.append(f"出勤率差{MIN_ATTEND_RATE-r['attend_rate']:.1%}")
    if r['prev_achieve'] is None: gaps.append("无上期数据")
    status = "✅ 达标" if not gaps else gaps[0]
    vals = [i+1, r['name'], r['group'], r['tenure'],
            f"{r['achieve_rate']:.2%}",
            f"{r['prev_achieve']:.2%}" if r['prev_achieve'] else 'N/A',
            f"{r['progress']:+.2%}" if r['progress'] is not None else 'N/A',
            r['attend_days'], f"{r['attend_rate']:.2%}", status]
    set_row(ws3, i+2, vals)
for i, w in enumerate([8,22,12,10,14,14,12,10,12,20]): ws3.column_dimensions[chr(65+i)].width = w

# ---- Sheet 4: Newcomer ----
ws4 = wb.create_sheet("03_新人达标奖")
h = ['排名','姓名','小组','入职天数','出勤天数','出勤率','达成率','达标等级','奖励','达标状态']
ws4.append(h); set_hdr(ws4, 1, len(h))
newcomer_show4 = sorted([r for r in results if 4 <= r['tenure'] <= 30], key=lambda x: x['achieve_rate'], reverse=True)
for idx, r in enumerate(newcomer_show4, start=1):
    rate = r['achieve_rate']
    lvl = '120%+' if rate>=1.20 else ('110%+' if rate>=1.10 else ('100%+' if rate>=1.00 else '未达标'))
    rwd = '1000 MXN' if rate>=1.20 else ('800 MXN' if rate>=1.10 else ('500 MXN' if rate>=1.00 else '—'))
    gaps = []
    if r['tenure'] < 4: gaps.append(f"入职仅{r['tenure']}天")
    if r['attend_rate'] < MIN_ATTEND_RATE: gaps.append(f"出勤率差{MIN_ATTEND_RATE-r['attend_rate']:.1%}")
    if rate < 1.00: gaps.append(f"达成率差{1.00-rate:.1%}")
    status = "✅ 获奖" if not gaps else gaps[0]
    vals = [idx, r['name'], r['group'], r['tenure'], r['attend_days'],
            f"{r['attend_rate']:.2%}", f"{rate:.2%}", lvl, rwd, status]
    set_row(ws4, ws4.max_row+1, vals)
for i, w in enumerate([8,22,12,10,10,12,12,10,12,20]): ws4.column_dimensions[chr(65+i)].width = w

out_name = f'{cycle_year}年{cycle_month:02d}月冲刺奖励_PL团队分析_{PERIOD_NAME}.xlsx'
home_dir = os.path.expanduser('~')
out_path = os.path.join(home_dir, out_name)
wb.save(out_path)
# Copy to desktop
import shutil
desktop = os.path.expanduser('~/Desktop')
desktop_path = os.path.join(desktop, out_name)
try:
    shutil.copy2(out_path, desktop_path)
    os.remove(out_path)
    out_path = desktop_path
except:
    pass
print(f"  Saved: {out_path}")

# ============================================================
# STEP 7: SEND TO FEISHU
# ============================================================
def get_fs_token():
    r = requests.post('https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal',
        json={'app_id': FEISHU_APP_ID, 'app_secret': FEISHU_APP_SECRET}, timeout=15)
    return r.json()['tenant_access_token']

def send_fs_text(token, target_id, target_type, text):
    r = requests.post('https://open.feishu.cn/open-apis/im/v1/messages',
        params={'receive_id_type': target_type},
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
        json={'receive_id': target_id, 'msg_type': 'text', 'content': json.dumps({'text': text})},
        timeout=15)
    return r.json().get('code') == 0

def send_fs_file(token, target_id, target_type, file_path, file_name):
    with open(file_path, 'rb') as f:
        r = requests.post('https://open.feishu.cn/open-apis/im/v1/files',
            headers={'Authorization': f'Bearer {token}'},
            files={'file': (file_name, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            data={'file_type': 'stream', 'file_name': file_name}, timeout=30)
    fk = r.json().get('data', {}).get('file_key')
    if fk:
        r = requests.post('https://open.feishu.cn/open-apis/im/v1/messages',
            params={'receive_id_type': target_type},
            headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
            json={'receive_id': target_id, 'msg_type': 'file', 'content': json.dumps({'file_key': fk})},
            timeout=15)
        return r.json().get('code') == 0
    return False

args = set(sys.argv[1:])
if '--send-user' in args or '--send-group' in args:
    print("[7/7] Sending to Feishu...")
    token = get_fs_token()

    # Build report text
    text = f'{cycle_month}月冲刺奖励活动 — PL团队分析报告\n'
    
    text += '================================\n\n'
    text += f'[一] 冠军员工奖: {len(champion_winners)}人获奖\n'
    text += f'条件: 入职>=30天 | 达成率>=100% | 出勤率>=90% | 出勤>=7天\n'
    if champions:
        p = champion_winners
        prizes = ['1000 MXN', '800 MXN', '500 MXN']
        for i, pp in enumerate(p):
            text += '%d. %s - achieve=%.2f%% attend=%dd rate=%.0f%% %s\n' % (
                i+1, pp['name'], pp['achieve_rate']*100, pp['attend_days'], pp['attend_rate']*100, prizes[i])
        others = [r['name'] for r in champions[3:]]
        if others: text += '其余达标: %s\n' % ', '.join(others)
    else:
        text += '暂无人完全达标\n'
    text += f'\n[二] 最佳进步奖: {len(progress_winners)}人获奖\n'
    if progress_winners:
        pp = progress_winners
        text += '1. %s - progress %+.2f%% | 800 MXN\n' % (pp[0]['name'], pp[0]['progress']*100)
        if len(pp)>=2: text += '2. %s - progress %+.2f%% | 500 MXN\n' % (pp[1]['name'], pp[1]['progress']*100)
    else:
        text += '暂无人达标\n'
    text += f'\n[三] 新人达标奖: {len(final_newcomer_winners)}人获奖\n'
    if final_newcomer_winners:
        for w in final_newcomer_winners:
            text += '%s(入职%dd, 达成%.2f%%) - %s\n' % (w['name'], w['tenure'], w['achieve_rate']*100, w['reward'])
    else:
        text += '暂无人达标\n'
    text += '\n================================\n'
    text += f'出勤率 = 实际上班天数 ÷ 应出勤天数（除休息外）\n'
    text += '达成率 = 累计回收金额 / 累计(入催金额×目标) | 与BI看板口径一致\n'
    text += '同周期不重复获奖, 只发最高奖项\n'
    text += '详细数据见下方Excel文件\n'

    if '--send-user' in args:
        ok = send_fs_text(token, USER_OPEN_ID, 'open_id', text)
        print('  Text to user: %s' % ('OK' if ok else 'FAIL'))
        ok = send_fs_file(token, USER_OPEN_ID, 'open_id', out_path, f'{cycle_year}年{cycle_month:02d}月冲刺奖励_PL团队分析_{PERIOD_NAME}.xlsx')
        print('  File to user: %s' % ('OK' if ok else 'FAIL'))

    if '--send-group' in args:
        ok = send_fs_text(token, CHAT_ID, 'chat_id', text)
        print('  Text to group: %s' % ('OK' if ok else 'FAIL'))
        ok = send_fs_file(token, CHAT_ID, 'chat_id', out_path, f'{cycle_year}年{cycle_month:02d}月冲刺奖励_PL团队分析_{PERIOD_NAME}.xlsx')
        print('  File to group: %s' % ('OK' if ok else 'FAIL'))
        time.sleep(1)
        mention_text = f'<at id={YANG_OPEN_ID}></at>总监，以上是{cycle_month}月冲刺奖励活动{PERIOD_NAME}分析报告，请查收。'
        ok = send_fs_text(token, CHAT_ID, 'chat_id', mention_text)
        print('  @Yang to group: %s' % ('OK' if ok else 'FAIL'))

print("\nDone!")
